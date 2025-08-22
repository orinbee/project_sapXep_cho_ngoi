import tkinter as tk
from tkinter import filedialog, messagebox, colorchooser, scrolledtext, TclError, font as tkfont, simpledialog
import ttkbootstrap as ttk
from ttkbootstrap.constants import *
import pandas as pd
from openpyxl import Workbook
import random
import sqlite3
import json
from datetime import datetime
from PIL import ImageGrab
import collections

# Yêu cầu cài đặt thư viện: pip install fpdf2 matplotlib
from fpdf import FPDF
from matplotlib.font_manager import findfont, FontProperties

# --- LỚP XỬ LÝ TẠO FILE PDF ---
class PDFGenerator(FPDF):
    def __init__(self, class_name, teacher_name, school_year, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.class_name = class_name
        self.teacher_name = teacher_name
        self.school_year = school_year
        
        self.font_name = "SystemFont"
        try:
            regular_path = findfont(FontProperties(family="Times New Roman", style="normal", weight="normal"))
            self.add_font(self.font_name, "", regular_path, uni=True)
            
            bold_path = findfont(FontProperties(family="Times New Roman", style="normal", weight="bold"))
            self.add_font(self.font_name, "B", bold_path, uni=True)
            
            italic_path = findfont(FontProperties(family="Times New Roman", style="italic", weight="normal"))
            self.add_font(self.font_name, "I", italic_path, uni=True)
            
            print(f"Sử dụng font hệ thống 'Times New Roman' thành công.")
            
        except Exception:
            try:
                self.add_font("DejaVu", "", "DejaVuSans.ttf", uni=True)
                self.font_name = "DejaVu"
                print("Sử dụng font cục bộ: DejaVuSans.ttf")
            except RuntimeError as e:
                raise RuntimeError("Không tìm thấy font 'Times New Roman' trong hệ thống và cũng không có 'DejaVuSans.ttf' trong thư mục ứng dụng.") from e
        
        self.set_font(self.font_name, "", 14)

    def header(self):
        self.set_font(self.font_name, "B", 20)
        title = f"SƠ ĐỒ CHỖ NGỒI LỚP {self.class_name.upper()}"
        self.cell(0, 10, title, 0, 1, 'C')
        
        self.set_font(self.font_name, "", 11)
        info_line = f"Năm học: {self.school_year}  |  GVCN: {self.teacher_name}"
        self.cell(0, 8, info_line, 0, 1, 'C')
        self.ln(8)

    def footer(self):
        self.set_y(-15)
        self.set_font(self.font_name, "I", 8)
        export_date = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
        self.cell(0, 10, f'Trang {self.page_no()}/{{nb}}  |  Xuất ngày: {export_date}', 0, 0, 'C')

    def draw_seating_chart(self, students, num_teams, num_tables_per_team, colors_hex):
        margin = 10
        drawable_width = self.w - 2 * margin
        drawable_height = self.h - 45

        team_gap = 10
        total_team_width = drawable_width - (num_teams - 1) * team_gap
        team_width = total_team_width / num_teams
        seat_width = (team_width - 5) / 2

        max_seat_height = drawable_height / num_tables_per_team - 5
        seat_height = min(25, max_seat_height) 
        
        def hex_to_rgb(hex_color):
            hex_color = hex_color.lstrip('#')
            return tuple(int(hex_color[i:i+2], 16) for i in (0, 2, 4))

        for i, student_name in enumerate(students):
            team_idx = i // (num_tables_per_team * 2)
            table_in_team_idx = (i % (num_tables_per_team * 2)) // 2
            seat_idx = i % 2

            x = margin + team_idx * (team_width + team_gap) + seat_idx * (seat_width + 5)
            y = 40 + table_in_team_idx * (seat_height + 5)

            color_rgb = hex_to_rgb(colors_hex[team_idx % len(colors_hex)])
            self.set_fill_color(*color_rgb)
            self.rect(x, y, seat_width, seat_height, 'DF')

            self.set_xy(x, y + seat_height/2 - 4)
            self.set_font(self.font_name, "", 9)
            self.set_text_color(0, 0, 0)
            self.multi_cell(seat_width, 4, student_name, 0, 'C')
            
# --- LỚP ỨNG DỤNG CHÍNH ---
class SeatArrangementApp:
    # --- CÁC HẰNG SỐ CHO GIAO DIỆN ---
    RECT_WIDTH = 180
    RECT_HEIGHT = 75
    X_GAP = 10
    Y_GAP = 20
    TEAM_X_GAP = 40
    
    # --- HẰNG SỐ MỚI CHO BÀN GIÁO VIÊN ---
    TEACHER_DESK_WIDTH = 250
    TEACHER_DESK_HEIGHT = 70
    TEACHER_AREA_GAP = 50 # Khoảng cách trống xung quanh bàn giáo viên
    
    def __init__(self, root):
        self.root = root
        self.root.title("Chương trình quản lý sắp xếp chỗ ngồi dành cho học sinh ")
        self.root.geometry("1366x1080")
        
        self.students_data = []
        self.students = []
        self.undo_stack = []
        self.redo_stack = []

        self.colors = ["#FFA07A", "#7FFFD4", "#87CEFA", "#FFD700", "#98FB98", "#F08080", "#E0FFFF"]
        self.seat_positions = {}
        self.text_positions = {}
        self.dragged_item = None
        self.start_coords = None
        self.tooltip = None
        self.num_teams = 4
        self.num_tables = 5
        
        self.current_class_id = None
        self.current_class_name = None
        self.is_dirty = False

        self.init_db()
        self.load_ui()
        self.root.protocol("WM_DELETE_WINDOW", self.on_closing)
        self.update_status("Chào mừng! Vui lòng chọn hoặc thêm một lớp học để bắt đầu.")

    # ... (init_db, on_closing, update_status, _set_dirty giữ nguyên)
    def init_db(self):
        self.conn = sqlite3.connect("seat_arrangements_multi_class.db")
        self.cursor = self.conn.cursor()
        
        self.cursor.execute('''
            CREATE TABLE IF NOT EXISTS classes (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL UNIQUE,
                students_data_json TEXT,
                current_arrangement_json TEXT,
                num_teams INTEGER,
                num_tables INTEGER,
                last_modified TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')

        self.cursor.execute('''
            CREATE TABLE IF NOT EXISTS arrangements_history (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                class_id INTEGER NOT NULL,
                arrangement TEXT,
                timestamp TEXT,
                FOREIGN KEY (class_id) REFERENCES classes (id) ON DELETE CASCADE
            )
        ''')
        self.conn.commit()

    def on_closing(self):
        if self.is_dirty:
            if not messagebox.askyesno("Thoát", "Bạn có những thay đổi chưa được lưu. Bạn có chắc chắn muốn thoát?"):
                return
        self.conn.close()
        self.root.destroy()

    def update_status(self, message):
        self.status_bar.config(text=message)
        
    def _set_dirty(self, dirty_status=True):
        if not self.current_class_id: return
        self.is_dirty = dirty_status
        
        current_text = self.class_selector.get()
        if dirty_status and not current_text.endswith('*'):
            self.class_selector.set(current_text + ' *')
        elif not dirty_status and current_text.endswith('*'):
            self.class_selector.set(current_text[:-2])

    def load_ui(self):
        # ... (Toàn bộ hàm load_ui giữ nguyên như phiên bản trước)
        main_pane = ttk.PanedWindow(self.root, orient=HORIZONTAL)
        main_pane.pack(fill=BOTH, expand=True)

        control_panel = ttk.Frame(main_pane, padding=10, width=320)
        control_panel.pack_propagate(False)
        main_pane.add(control_panel, weight=1)

        canvas_container = ttk.Frame(main_pane, padding=(0, 10, 10, 0))
        main_pane.add(canvas_container, weight=4)

        class_lf = ttk.LabelFrame(control_panel, text=" Quản Lý Lớp Học ", padding=10, bootstyle=PRIMARY)
        class_lf.pack(fill=X, pady=(0, 15))
        
        self.class_selector = ttk.Combobox(class_lf, state="readonly", values=[])
        self.class_selector.pack(fill=X, pady=(0, 5))
        self.class_selector.bind("<<ComboboxSelected>>", self._on_class_selected)

        class_btn_frame = ttk.Frame(class_lf)
        class_btn_frame.pack(fill=X)
        ttk.Button(class_btn_frame, text="Thêm Lớp Mới", command=self._add_new_class, bootstyle="success-outline").pack(side=LEFT, expand=True, fill=X, padx=(0,2))
        ttk.Button(class_btn_frame, text="Xóa Lớp Này", command=self._delete_class, bootstyle="danger-outline").pack(side=LEFT, expand=True, fill=X, padx=(2,0))

        ttk.Button(class_lf, text="💾 LƯU TRẠNG THÁI LỚP", command=self._save_class_state, bootstyle="primary").pack(fill=X, pady=5)

        theme_frame = ttk.Frame(control_panel); theme_frame.pack(fill=X, pady=(0, 15))
        ttk.Label(theme_frame, text="Chế độ Sáng / Tối:").pack(side=LEFT)
        self.theme_var = tk.BooleanVar(value=False)
        theme_switch = ttk.Checkbutton(theme_frame, bootstyle="switch", variable=self.theme_var, command=self.toggle_theme); theme_switch.pack(side=LEFT, padx=10)

        lf1 = ttk.LabelFrame(control_panel, text=" 1. Tải Dữ Liệu Học Sinh ", padding=10, bootstyle=DEFAULT); lf1.pack(fill=X, pady=(0, 10))
        load_btn = ttk.Button(lf1, text="📂 Tải File Excel cho Lớp Này", command=self.load_students, bootstyle=DEFAULT); load_btn.pack(fill=X)

        lf2 = ttk.LabelFrame(control_panel, text=" 2. Cấu Hình Lớp ", padding=10, bootstyle=DEFAULT); lf2.pack(fill=X, pady=10)
        ttk.Label(lf2, text="Số tổ:").grid(row=0, column=0, padx=5, pady=5, sticky=W)
        self.team_spinbox = ttk.Spinbox(lf2, from_=1, to=20, width=5); self.team_spinbox.grid(row=0, column=1, padx=5, pady=5, sticky=W); self.team_spinbox.set(self.num_teams)
        ttk.Label(lf2, text="Số bàn mỗi tổ:").grid(row=1, column=0, padx=5, pady=5, sticky=W)
        self.table_spinbox = ttk.Spinbox(lf2, from_=1, to=30, width=5); self.table_spinbox.grid(row=1, column=1, padx=5, pady=5, sticky=W); self.table_spinbox.set(self.num_tables)
        apply_config_btn = ttk.Button(lf2, text="Áp dụng cấu hình", command=self.apply_team_table_config, bootstyle="secondary-outline"); apply_config_btn.grid(row=2, column=0, columnspan=2, pady=10, sticky=EW)

        lf3 = ttk.LabelFrame(control_panel, text=" 3. Sắp Xếp & Tùy Chỉnh ", padding=10, bootstyle=DEFAULT); lf3.pack(fill=X, pady=10)
        sort_btn = ttk.Button(lf3, text="🎲 Tùy chọn Sắp xếp", command=self.show_sort_options); sort_btn.pack(fill=X, pady=(0,5))
        undo_redo_frame = ttk.Frame(lf3); undo_redo_frame.pack(fill=X, pady=(0,5))
        self.undo_btn = ttk.Button(undo_redo_frame, text="↩️ Hoàn tác", command=self.undo, bootstyle="secondary-outline", state="disabled"); self.undo_btn.pack(side=LEFT, expand=True, fill=X, padx=(0,2))
        self.redo_btn = ttk.Button(undo_redo_frame, text="↪️ Làm lại", command=self.redo, bootstyle="secondary-outline", state="disabled"); self.redo_btn.pack(side=LEFT, expand=True, fill=X, padx=(2,0))

        lf4 = ttk.LabelFrame(control_panel, text=" 4. Xuất & Báo Cáo ", padding=10, bootstyle=DEFAULT); lf4.pack(fill=X, pady=10)
        save_excel_btn = ttk.Button(lf4, text="Xuất Sơ Đồ ra Excel", command=self.save_results, bootstyle="info-outline"); save_excel_btn.pack(fill=X, pady=(0, 5))
        save_pdf_btn = ttk.Button(lf4, text="🖨️ Xuất Sơ Đồ ra PDF (Để in)", command=self.export_to_pdf, bootstyle="info"); save_pdf_btn.pack(fill=X, pady=(5, 5))
        save_img_btn = ttk.Button(lf4, text="Chụp Ảnh Sơ Đồ", command=self.save_as_image, bootstyle="info-outline"); save_img_btn.pack(fill=X)
        
        lf5 = ttk.LabelFrame(control_panel, text=" Công Cụ Khác ", padding=10, bootstyle=DEFAULT); lf5.pack(fill=X, pady=10)
        color_btn = ttk.Button(lf5, text="🎨 Đổi Màu Tổ", command=self.change_team_colors, bootstyle="secondary-outline"); color_btn.pack(fill=X, pady=(0, 5))
        history_btn = ttk.Button(lf5, text="📜 Xem Lịch Sử Sắp Xếp của Lớp", command=self.view_history, bootstyle="secondary-outline"); history_btn.pack(fill=X)

        ttk.Label(canvas_container, text="Sơ Đồ Lớp Học", font=("Arial", 16, "bold")).pack(pady=(0, 5))
        canvas_frame = ttk.Frame(canvas_container); canvas_frame.pack(fill=BOTH, expand=True)
        self.canvas = tk.Canvas(canvas_frame, bg='white', relief="solid", bd=1)
        h_scroll = ttk.Scrollbar(canvas_frame, orient=HORIZONTAL, command=self.canvas.xview, bootstyle="round"); v_scroll = ttk.Scrollbar(canvas_frame, orient=VERTICAL, command=self.canvas.yview, bootstyle="round")
        self.canvas.config(xscrollcommand=h_scroll.set, yscrollcommand=v_scroll.set)
        h_scroll.pack(side=BOTTOM, fill=X); v_scroll.pack(side=RIGHT, fill=Y); self.canvas.pack(side=LEFT, fill=BOTH, expand=True)
        random_pick_btn = ttk.Button(lf5, text="✨ Gọi Tên Ngẫu Nhiên", command=self.pick_random_student)
        random_pick_btn.pack(fill=X, pady=(5, 0))

        
        self.status_bar = ttk.Label(self.root, text="Sẵn sàng", padding=5, font=("Arial", 9), anchor=W, bootstyle=INVERSE); self.status_bar.pack(side=BOTTOM, fill=X)
        
        self._load_class_list_to_selector()
        
    def export_to_pdf(self):
        # ... (Toàn bộ hàm export_to_pdf giữ nguyên như phiên bản trước)
        if not self.current_class_id or not self.students:
            messagebox.showwarning("Chưa có dữ liệu", "Vui lòng chọn một lớp và tải dữ liệu học sinh để xuất ra PDF.")
            return

        teacher_name = simpledialog.askstring("Thông tin bổ sung", "Nhập tên Giáo viên Chủ nhiệm:", parent=self.root)
        if teacher_name is None: return
        
        school_year = simpledialog.askstring("Thông tin bổ sung", "Nhập Năm học (VD: 2025-2026):", parent=self.root)
        if school_year is None: return

        file_path = filedialog.asksaveasfilename(
            title=f"Lưu sơ đồ PDF cho lớp {self.current_class_name}",
            defaultextension=".pdf",
            filetypes=[("PDF Files", "*.pdf")]
        )
        if not file_path:
            return

        self.update_status("Đang tạo file PDF, vui lòng đợi...")
        try:
            pdf = PDFGenerator(
                orientation='L', unit='mm', format='A4',
                class_name=self.current_class_name,
                teacher_name=teacher_name,
                school_year=school_year
            )
            pdf.alias_nb_pages()
            pdf.add_page()
            pdf.draw_seating_chart(
                students=self.students,
                num_teams=self.num_teams,
                num_tables_per_team=self.num_tables,
                colors_hex=self.colors
            )
            pdf.output(file_path)
            
            self.update_status(f"Đã xuất PDF thành công: {file_path}")
            messagebox.showinfo("Thành công", f"Đã xuất sơ đồ lớp {self.current_class_name} ra file PDF thành công!")
        
        except Exception as e:
            if isinstance(e, RuntimeError) and ("TTF Font file not found" in str(e) or "DejaVuSans.ttf" in str(e)):
                 messagebox.showerror(
                     "Lỗi Thiếu Font Chữ", 
                     "Không tìm thấy font chữ cần thiết để hỗ trợ tiếng Việt.\n\n"
                     "Giải pháp:\n"
                     "1. Đảm bảo font 'Times New Roman' đã được cài đặt trên máy của bạn.\n"
                     "2. (Nếu cách 1 không được) Tải và đặt file 'DejaVuSans.ttf' vào cùng thư mục với ứng dụng."
                 )
            else:
                messagebox.showerror("Lỗi", f"Không thể tạo file PDF: {e}")
            self.update_status("Lỗi khi tạo file PDF.")
            
    def _wrap_text(self, text, font_object, max_width):
        # ... (Hàm _wrap_text giữ nguyên)
        words = text.split()
        lines = []
        current_line = ""
        for word in words:
            separator = " " if current_line else ""
            test_line = current_line + separator + word
            
            if font_object.measure(test_line) <= max_width:
                current_line = test_line
            else:
                lines.append(current_line)
                current_line = word
        lines.append(current_line)
        return "\n".join(lines)

    # --- HÀM arrange_seats ĐƯỢC CẬP NHẬT ĐỂ VẼ BÀN GIÁO VIÊN ---
        # --- THAY THẾ TOÀN BỘ HÀM arrange_seats CŨ BẰNG HÀM NÀY ---
    def arrange_seats(self):
        self.canvas.delete("all")
        self.seat_positions.clear()
        self.text_positions.clear()
        
        if not self.students: 
            bbox = self.canvas.bbox("all")
            if bbox: self.canvas.config(scrollregion=bbox)
            return

        text_font = tkfont.Font(family="Arial", size=11, weight="bold")
        
        # --- LOGIC MỚI: TẠO KHOẢNG TRỐNG PHÍA TRÊN CHO BÀN GIÁO VIÊN ---
        # Đẩy tất cả bàn học sinh xuống dưới một khoảng
        top_offset = self.TEACHER_DESK_HEIGHT + 100
        
        student_map = {s['Học sinh']: s for s in self.students_data}
        for i, student_name in enumerate(self.students):
            team_index = i // (self.num_tables * 2)
            table_in_team_index = (i % (self.num_tables * 2)) // 2
            seat_index = i % 2
            student_info = student_map.get(student_name)
            if not student_info: continue
            
            # Tính toán vị trí x, không cần dịch chuyển các tổ nữa
            team_visual_width = self.RECT_WIDTH * 2 + self.X_GAP + self.TEAM_X_GAP
            x_start = 20 + team_index * team_visual_width
            
            # Áp dụng khoảng trống phía trên cho vị trí y
            y_start = top_offset + table_in_team_index * (self.RECT_HEIGHT + self.Y_GAP)
            
            x = x_start + seat_index * (self.RECT_WIDTH + self.X_GAP)
            y = y_start
            
            # ... (Phần vẽ chỗ ngồi học sinh giữ nguyên)
            outline_color = "#E53935" if student_info.get('GhiChu') == 'Cần ngồi trước' else ("#FFFFFF" if self.theme_var.get() else "#000000")
            outline_width = 3 if student_info.get('GhiChu') == 'Cần ngồi trước' else 1
            text_color = "white" if self.theme_var.get() else "black"
            
            rect = self.canvas.create_rectangle(x, y, x + self.RECT_WIDTH, y + self.RECT_HEIGHT, fill=self.colors[team_index % len(self.colors)], tags="rect", outline=outline_color, width=outline_width)
            
            wrapped_name = self._wrap_text(student_name, text_font, self.RECT_WIDTH - 10)
            
            text = self.canvas.create_text(
                x + self.RECT_WIDTH / 2, y + self.RECT_HEIGHT / 2, 
                text=wrapped_name, 
                font=text_font, 
                tags="text", 
                fill=text_color,
                justify=tk.CENTER
            )

            self.seat_positions[rect] = student_info
            self.text_positions[rect] = text
            self.canvas.tag_bind(rect, "<Button-1>", self.start_drag)
            self.canvas.tag_bind(rect, "<B1-Motion>", self.on_drag)
            self.canvas.tag_bind(rect, "<ButtonRelease-1>", self.stop_drag)
            self.canvas.tag_bind(rect, "<Enter>", lambda e, r=rect: self.show_tooltip(e, r))
            self.canvas.tag_bind(rect, "<Leave>", self.hide_tooltip)
        
        # --- VẼ BÀN GIÁO VIÊN VÀO VỊ TRÍ MỚI (PHÍA TRÊN TỔ 1) ---
        if self.num_teams > 0:
            # Tính toán vị trí trung tâm của Tổ 1
            team_1_start_x = 20
            team_1_width = self.RECT_WIDTH * 2 + self.X_GAP
            team_1_center_x = team_1_start_x + team_1_width / 2

            # Đặt bàn giáo viên ở giữa Tổ 1 và ở lề trên cùng
            teacher_desk_x = team_1_center_x - (self.TEACHER_DESK_WIDTH / 2)
            teacher_desk_y = 20 

            self.canvas.create_rectangle(
                teacher_desk_x, teacher_desk_y,
                teacher_desk_x + self.TEACHER_DESK_WIDTH,
                teacher_desk_y + self.TEACHER_DESK_HEIGHT,
                fill="#DEB887", # Màu BurlyWood, giống màu gỗ hơn
                outline=("white" if self.theme_var.get() else "black")
            )
            
            self.canvas.create_text(
                teacher_desk_x + self.TEACHER_DESK_WIDTH / 2,
                teacher_desk_y + self.TEACHER_DESK_HEIGHT / 2,
                text="Bàn Giáo Viên",
                font=text_font,
                fill=("white" if self.theme_var.get() else "black")
            )
            
        bbox = self.canvas.bbox("all")
        if bbox: self.canvas.config(scrollregion=bbox)
    def pick_random_student(self):
        """
        Chọn ngẫu nhiên một học sinh từ danh sách, làm nổi bật chỗ ngồi trên sơ đồ
        và hiển thị thông báo. Hiệu ứng nổi bật sẽ tự động tắt sau một khoảng thời gian.
        """
        if not self.students:
            messagebox.showwarning("Chưa có dữ liệu", "Không có học sinh nào trong danh sách để lựa chọn.")
            return

        # 1. Chọn ngẫu nhiên một học sinh
        random_student_name = random.choice(self.students)
        self.update_status(f"Đang chọn ngẫu nhiên... Kết quả là: {random_student_name}!")

        # 2. Tìm kiếm đối tượng (rectangle) tương ứng trên canvas
        target_rect_id = None
        for rect_id, student_info in self.seat_positions.items():
            if student_info['Học sinh'] == random_student_name:
                target_rect_id = rect_id
                break
        
        # 3. Tạo hiệu ứng làm nổi bật và hiển thị thông báo
        if target_rect_id:
            # Lưu lại màu gốc
            original_color = self.canvas.itemcget(target_rect_id, "fill")
            highlight_color = "#FFD700"  # Màu vàng gold

            # Hàm nhỏ để hoàn lại màu sắc ban đầu
            def revert_highlight():
                self.canvas.itemconfig(target_rect_id, fill=original_color)

            # Bắt đầu làm nổi bật
            self.canvas.itemconfig(target_rect_id, fill=highlight_color)
            
            # Đưa hình chữ nhật và chữ lên lớp trên cùng để đảm bảo nhìn thấy rõ
            text_item = self.text_positions.get(target_rect_id)
            if text_item:
                self.canvas.tag_raise(target_rect_id)
                self.canvas.tag_raise(text_item)
            
            # Hiển thị thông báo cho giáo viên
            messagebox.showinfo(
                "Học sinh được chọn",
                f"🌟 Chúc mừng em: {random_student_name}! 🌟",
                parent=self.root
            )

            # Lên lịch để tắt hiệu ứng nổi bật sau 2 giây (2000 mili giây)
            self.root.after(2000, revert_highlight)
        else:
            # Trường hợp dự phòng nếu không tìm thấy học sinh trên sơ đồ
            messagebox.showinfo(
                 "Học sinh được chọn",
                f"Học sinh được chọn là: {random_student_name} (không tìm thấy trên sơ đồ)."
            )

    # --- CÁC HÀM CÒN LẠI GIỮ NGUYÊN ---
    # (Toàn bộ các hàm còn lại không có thay đổi so với phiên bản trước đó)
    def toggle_theme(self):
        if self.theme_var.get():
            self.root.style.theme_use('darkly')
            self.canvas.config(bg="#303030")
        else:
            self.root.style.theme_use('litera')
            self.canvas.config(bg="white")
        if self.students: self.arrange_seats()

    def _save_state_for_undo(self):
        if not self.students:
            return
        self.undo_stack.append(self.students.copy())
        self.redo_stack.clear()
        self._update_undo_redo_buttons()

    def _update_undo_redo_buttons(self):
        self.undo_btn.config(state="normal" if self.undo_stack else "disabled")
        self.redo_btn.config(state="normal" if self.redo_stack else "disabled")

    def undo(self):
        if not self.undo_stack:
            return
        self.redo_stack.append(self.students.copy())
        self.students = self.undo_stack.pop()
        self._set_dirty()
        self._sync_data_and_ui()
        self.update_status("Đã hoàn tác hành động.")

    def redo(self):
        if not self.redo_stack:
            return
        self.undo_stack.append(self.students.copy())
        self.students = self.redo_stack.pop()
        self._set_dirty()
        self._sync_data_and_ui()
        self.update_status("Đã làm lại hành động.")

    def _sync_data_and_ui(self):
        if self.students_data:
            student_map = {s['Học sinh']: s for s in self.students_data}
            self.students_data = [student_map[name] for name in self.students if name in student_map]
        
        self.arrange_seats()
        self._update_undo_redo_buttons()

    def apply_sort(self, window):
        self._save_state_for_undo()
        sort_method = self.sort_method_var.get()
        if sort_method == "intelligent": self.students = AdvancedSorter(self.students_data).arrange()
        elif sort_method == "random": random.shuffle(self.students)
        elif sort_method == "height":
            df = pd.DataFrame(self.students_data)
            self.students = df.sort_values(by='Chiều cao', ascending=False)['Học sinh'].tolist()
        elif sort_method == "gender":
            df = pd.DataFrame(self.students_data)
            males = df[df['Giới tính'] == 'Nam']['Học sinh'].tolist(); random.shuffle(males)
            females = df[df['Giới tính'] == 'Nữ']['Học sinh'].tolist(); random.shuffle(females)
            self.students = []
            i, j = 0, 0
            while i < len(males) or j < len(females):
                if i < len(males): self.students.append(males[i]); i += 1
                if j < len(females): self.students.append(females[j]); j += 1
        
        self._set_dirty()
        self._sync_data_and_ui()
        self.update_status(f"Đã áp dụng sắp xếp theo phương pháp: {sort_method}.")
        window.destroy()

    def stop_drag(self, event):
        if not self.dragged_item: return
        drop_x = self.canvas.canvasx(event.x)
        drop_y = self.canvas.canvasy(event.y)
        overlapping_items = self.canvas.find_overlapping(drop_x, drop_y, drop_x, drop_y)
        target_item = None
        for item in overlapping_items:
            if item in self.seat_positions and item != self.dragged_item:
                target_item = item
                break
        if target_item:
            self._save_state_for_undo()
            
            dragged_info = self.seat_positions[self.dragged_item]
            target_info = self.seat_positions[target_item]
            dragged_index = self.students.index(dragged_info['Học sinh'])
            target_index = self.students.index(target_info['Học sinh'])
            self.students[dragged_index], self.students[target_index] = self.students[target_index], self.students[dragged_index]
            self.update_status(f"Đã hoán đổi vị trí: {dragged_info['Học sinh']} và {target_info['Học sinh']}.")
            self._set_dirty()
            self._sync_data_and_ui()
        else: 
            self.update_status("Thao tác kéo thả bị hủy.")
            self.arrange_seats()
            
        self.dragged_item = None

    def restore_arrangement(self, arrangement_json):
        if not self.students_data:
            messagebox.showerror("Lỗi", "Không thể khôi phục khi chưa có danh sách học sinh.")
            return
        
        self._save_state_for_undo()
        
        restored_student_names = json.loads(arrangement_json)
        if set(restored_student_names) != {s['Học sinh'] for s in self.students_data}:
            messagebox.showwarning("Cảnh báo", "Danh sách học sinh trong lịch sử không khớp với danh sách hiện tại.")
        
        self.students = restored_student_names
        
        if self.num_teams * self.num_tables * 2 < len(self.students):
            messagebox.showwarning("Cảnh báo", f"Số ghế hiện tại không đủ.")
        
        self._set_dirty()
        self._sync_data_and_ui()
        self.update_status("Đã khôi phục sơ đồ từ lịch sử.")
        messagebox.showinfo("Thành công", "Đã khôi phục sơ đồ từ lịch sử!")
    
    def _load_class_list_to_selector(self):
        try:
            self.cursor.execute("SELECT id, name FROM classes ORDER BY name")
            self.class_list = self.cursor.fetchall()
            class_names = [row[1] for row in self.class_list]
            self.class_selector['values'] = class_names
            if not class_names:
                self.class_selector.set("Chưa có lớp nào. Hãy thêm một lớp.")
        except sqlite3.Error as e:
            messagebox.showerror("Lỗi Database", f"Không thể tải danh sách lớp: {e}")

    def _on_class_selected(self, event=None):
        if self.is_dirty:
            if not messagebox.askyesno("Cảnh báo", "Bạn có thay đổi chưa lưu ở lớp hiện tại. Bạn có muốn chuyển lớp và hủy các thay đổi đó?"):
                self.class_selector.set(self.current_class_name)
                return
        
        selected_name = self.class_selector.get()
        selected_class = next((c for c in self.class_list if c[1] == selected_name), None)

        if selected_class:
            self.current_class_id, self.current_class_name = selected_class
            try:
                self.cursor.execute("SELECT * FROM classes WHERE id=?", (self.current_class_id,))
                class_data = self.cursor.fetchone()
                
                self.students_data = []
                self.students = []
                
                if class_data and class_data[2]:
                    self.students_data = json.loads(class_data[2])
                if class_data and class_data[3]:
                    self.students = json.loads(class_data[3])
                
                self.num_teams = class_data[4] if class_data[4] else 1
                self.num_tables = class_data[5] if class_data[5] else 3
                
                self.team_spinbox.set(self.num_teams)
                self.table_spinbox.set(self.num_tables)
                
                self.undo_stack.clear()
                self.redo_stack.clear()
                self._set_dirty(False)
                self._sync_data_and_ui()
                self.update_status(f"Đã tải dữ liệu cho lớp: {self.current_class_name}")

            except sqlite3.Error as e:
                messagebox.showerror("Lỗi Database", f"Không thể tải dữ liệu lớp: {e}")
            except (json.JSONDecodeError, TypeError):
                messagebox.showwarning("Dữ liệu lỗi", "Dữ liệu của lớp này có thể bị lỗi. Vui lòng tải lại file Excel.")
                self._clear_canvas_and_data()

    def _add_new_class(self):
        class_name = simpledialog.askstring("Thêm Lớp Mới", "Nhập tên lớp học:", parent=self.root)
        if not class_name or not class_name.strip():
            return

        class_name = class_name.strip()
        try:
            self.cursor.execute("INSERT INTO classes (name) VALUES (?)", (class_name,))
            self.conn.commit()
            self.update_status(f"Đã tạo lớp mới: {class_name}")
            self._load_class_list_to_selector()
            self.class_selector.set(class_name)
            self._on_class_selected()
        except sqlite3.IntegrityError:
            messagebox.showerror("Lỗi", f"Tên lớp '{class_name}' đã tồn tại. Vui lòng chọn tên khác.")
        except sqlite3.Error as e:
            messagebox.showerror("Lỗi Database", f"Không thể tạo lớp mới: {e}")

    def _save_class_state(self):
        if not self.current_class_id:
            messagebox.showwarning("Chưa chọn lớp", "Vui lòng chọn một lớp để lưu.")
            return

        try:
            students_data_json = json.dumps(self.students_data)
            current_arrangement_json = json.dumps(self.students)
            self.num_teams = int(self.team_spinbox.get())
            self.num_tables = int(self.table_spinbox.get())
            
            self.cursor.execute("""
                UPDATE classes 
                SET students_data_json=?, current_arrangement_json=?, num_teams=?, num_tables=?, last_modified=CURRENT_TIMESTAMP
                WHERE id=?
            """, (students_data_json, current_arrangement_json, self.num_teams, self.num_tables, self.current_class_id))
            self.conn.commit()
            self._set_dirty(False)
            self.update_status(f"Đã lưu thành công trạng thái của lớp: {self.current_class_name}")
            messagebox.showinfo("Thành công", f"Đã lưu trạng thái của lớp '{self.current_class_name}'.")
        except (sqlite3.Error, TclError, ValueError) as e:
             messagebox.showerror("Lỗi", f"Không thể lưu trạng thái lớp: {e}")

    def _delete_class(self):
        if not self.current_class_id:
            messagebox.showwarning("Chưa chọn lớp", "Vui lòng chọn một lớp để xóa.")
            return
        
        if messagebox.askyesno("Xác nhận Xóa", f"Bạn có chắc chắn muốn xóa vĩnh viễn lớp '{self.current_class_name}' và toàn bộ lịch sử của nó không?"):
            try:
                class_name_to_delete = self.current_class_name
                self.cursor.execute("DELETE FROM classes WHERE id=?", (self.current_class_id,))
                self.conn.commit()
                self._clear_canvas_and_data()
                self.current_class_id = None
                self.current_class_name = None
                self._load_class_list_to_selector()
                self.update_status(f"Đã xóa lớp: {class_name_to_delete}")
            except sqlite3.Error as e:
                messagebox.showerror("Lỗi Database", f"Không thể xóa lớp: {e}")

    def _clear_canvas_and_data(self):
        self.students = []
        self.students_data = []
        self.undo_stack.clear()
        self.redo_stack.clear()
        self._sync_data_and_ui()
        
    def apply_team_table_config(self):
        try:
            self.num_teams = int(self.team_spinbox.get())
            self.num_tables = int(self.table_spinbox.get())
            if self.num_teams <= 0 or self.num_tables <= 0:
                raise ValueError("Số tổ và số bàn phải lớn hơn 0!")
            total_seats = self.num_teams * self.num_tables * 2
            if self.students and total_seats < len(self.students):
                messagebox.showwarning("Cảnh báo", f"Tổng số ghế ({total_seats}) nhỏ hơn số học sinh ({len(self.students)})!", title="Cấu hình không hợp lệ")
                return
            if self.students:
                self.arrange_seats()
                self._set_dirty()
            self.update_status(f"Đã thiết lập: {self.num_teams} tổ, {self.num_tables} bàn mỗi tổ.")
        except (ValueError, TclError):
            messagebox.showerror("Lỗi", "Số tổ và số bàn phải là các số hợp lệ.", title="Lỗi dữ liệu")
            self.update_status("Lỗi: Dữ liệu cấu hình không hợp lệ.")

    def show_sort_options(self):
        if not self.students_data:
            messagebox.showwarning("Cảnh báo", "Vui lòng nhập danh sách học sinh trước!", title="Chưa có dữ liệu")
            return
        
        sort_window = ttk.Toplevel(self.root)
        sort_window.title("Tùy chọn sắp xếp")
        sort_window.transient(self.root)
        sort_window.grab_set()
        
        container = ttk.Frame(sort_window, padding=20)
        container.pack(fill=BOTH, expand=True)

        ttk.Label(container, text="Chọn phương pháp sắp xếp:", font=("Arial", 12)).pack(pady=10)
        self.sort_method_var = tk.StringVar(value="intelligent")

        style = ttk.Style()
        style.configure('TRadiobutton', font=('Arial', 10), padding=(0,5))
        
        ttk.Radiobutton(container, text="Thông minh (Đề xuất)", variable=self.sort_method_var, value="intelligent", style='TRadiobutton').pack(fill=X, padx=20, pady=2)
        ttk.Radiobutton(container, text="Ngẫu nhiên", variable=self.sort_method_var, value="random", style='TRadiobutton').pack(fill=X, padx=20, pady=2)
        ttk.Radiobutton(container, text="Xen kẽ nam-nữ", variable=self.sort_method_var, value="gender", style='TRadiobutton').pack(fill=X, padx=20, pady=2)
        ttk.Radiobutton(container, text="Theo chiều cao (cao trước, thấp sau)", variable=self.sort_method_var, value="height", style='TRadiobutton').pack(fill=X, padx=20, pady=2)
        
        button_frame = ttk.Frame(container)
        button_frame.pack(pady=20)
        ttk.Button(button_frame, text="Áp dụng", command=lambda: self.apply_sort(sort_window), bootstyle=SUCCESS).pack(side=LEFT, padx=10)
        ttk.Button(button_frame, text="Hủy", command=sort_window.destroy, bootstyle="secondary-outline").pack(side=LEFT, padx=10)
        
    def load_students(self):
        if not self.current_class_id:
            messagebox.showwarning("Chưa chọn lớp", "Vui lòng chọn hoặc thêm một lớp trước khi tải danh sách học sinh.")
            return
        
        self.undo_stack.clear()
        self.redo_stack.clear()
        
        file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx;*.xls")])
        if not file_path: return
        try:
            df = pd.read_excel(file_path)
            required_columns = ['Học sinh', 'Giới tính', 'Chiều cao', 'DiemTB']
            if not all(col in df.columns for col in required_columns):
                raise ValueError("File Excel phải có các cột: 'Học sinh', 'Giới tính', 'Chiều cao', 'DiemTB'.\nCác cột tùy chọn: 'GhiChu', 'KhongNgoiCanh'.")
            
            df['GhiChu'] = df['GhiChu'].fillna('')
            df['KhongNgoiCanh'] = df['KhongNgoiCanh'].fillna('')

            for index, row in df.iterrows():
                if pd.isna(row['Chiều cao']) or not isinstance(row['Chiều cao'], (int, float)): raise ValueError(f"Dòng {index + 2}: 'Chiều cao' phải là một con số.")
                if pd.isna(row['DiemTB']) or not isinstance(row['DiemTB'], (int, float)): raise ValueError(f"Dòng {index + 2}: 'DiemTB' phải là một con số.")
                if row['Giới tính'] not in ['Nam', 'Nữ']: raise ValueError(f"Dòng {index + 2}: 'Giới tính' phải là 'Nam' hoặc 'Nữ'.")

            self.students_data = df.to_dict('records')
            self.students = df['Học sinh'].tolist()
            
            self.update_status(f"Đã tải {len(self.students)} HS cho lớp {self.current_class_name}. Nhấn 'Lưu Trạng Thái' để ghi nhớ.")
            random.shuffle(self.students)
            self._set_dirty()
            self._sync_data_and_ui()
        except Exception as e:
            messagebox.showerror("Lỗi tải file", f"Đã xảy ra lỗi: {str(e)}", title="Lỗi")
            self.update_status(f"Lỗi tải file Excel cho lớp {self.current_class_name}.")

    def change_team_colors(self):
        if self.num_teams == 0:
            messagebox.showinfo("Thông báo", "Vui lòng thiết lập số tổ trước.")
            return
        
        color_window = ttk.Toplevel(self.root)
        color_window.title("Đổi màu tổ")
        
        for i in range(self.num_teams):
            frame = ttk.Frame(color_window, padding=5)
            frame.pack(fill=X)
            while i >= len(self.colors): self.colors.append("#FFFFFF")
            color_preview = tk.Label(frame, text="    ", bg=self.colors[i], relief="solid", borderwidth=1)
            color_preview.pack(side=LEFT, padx=5)
            ttk.Label(frame, text=f"Màu cho tổ {i + 1}:").pack(side=LEFT, padx=5)
            ttk.Button(frame, text="Chọn màu", bootstyle="outline", command=lambda idx=i, p=color_preview: self.choose_color(idx, p)).pack(side=LEFT, padx=5)

    def view_history(self):
        if not self.current_class_id:
            messagebox.showwarning("Chưa chọn lớp", "Vui lòng chọn một lớp để xem lịch sử.")
            return

        history_window = ttk.Toplevel(self.root)
        history_window.title(f"Lịch sử sắp xếp - Lớp {self.current_class_name}")
        history_window.geometry("1050x1000")
        history_window.transient(self.root); history_window.grab_set()
        
        left_frame = ttk.Frame(history_window, padding=5); left_frame.pack(side=LEFT, fill=Y)
        right_frame = ttk.Frame(history_window, padding=5); right_frame.pack(side=RIGHT, fill=BOTH, expand=True)
        ttk.Label(left_frame, text="Các phiên đã lưu", font=("Arial", 12, "bold")).pack(pady=5)
        list_frame = ttk.Frame(left_frame); list_frame.pack(fill=BOTH, expand=True)
        history_listbox = tk.Listbox(list_frame, width=30, font=("Arial", 10)); scrollbar = ttk.Scrollbar(list_frame, orient="vertical", command=history_listbox.yview, bootstyle="round")
        history_listbox.config(yscrollcommand=scrollbar.set); scrollbar.pack(side=RIGHT, fill=Y); history_listbox.pack(side=LEFT, fill=BOTH, expand=True)
        ttk.Label(right_frame, text="Xem trước Sơ đồ", font=("Arial", 12, "bold")).pack(pady=5)
        preview_text = scrolledtext.ScrolledText(right_frame, wrap=tk.WORD, state=tk.DISABLED, font=("Arial", 10)); preview_text.pack(fill=BOTH, expand=True, pady=5)
        button_frame = ttk.Frame(right_frame); button_frame.pack(fill=X, pady=5)
        restore_btn = ttk.Button(button_frame, text="Khôi phục phiên này", state=tk.DISABLED, bootstyle=SUCCESS); restore_btn.pack(side=LEFT, padx=5, expand=True, fill=X)
        delete_btn = ttk.Button(button_frame, text="Xóa mục này", state=tk.DISABLED, bootstyle=DANGER); delete_btn.pack(side=LEFT, padx=5, expand=True, fill=X)
        ttk.Button(right_frame, text="Xóa toàn bộ lịch sử của lớp này", bootstyle="danger-outline", command=lambda: self.clear_history(history_window)).pack(fill=X, pady=10)

        self.cursor.execute("SELECT id, timestamp, arrangement FROM arrangements_history WHERE class_id=? ORDER BY timestamp DESC", (self.current_class_id,))
        arrangements_data = self.cursor.fetchall()
        
        if not arrangements_data: history_listbox.insert(tk.END, "Không có lịch sử nào."); return
        history_map = {index: {'id': arr_id, 'json': arr_json} for index, (arr_id, _, arr_json) in enumerate(arrangements_data)}
        for index, (arr_id, timestamp, _) in enumerate(arrangements_data): history_listbox.insert(tk.END, f"ID {arr_id}: {timestamp}")
        def on_history_select(event):
            selected_indices = history_listbox.curselection()
            if not selected_indices: return
            selected_index = selected_indices[0]; data = history_map.get(selected_index)
            if not data: return
            preview_text.config(state=tk.NORMAL); preview_text.delete('1.0', tk.END)
            students_list = json.loads(data['json'])
            preview_content = f"--- SƠ ĐỒ CHI TIẾT (ID: {data['id']}) ---\n\n"
            for i, student_name in enumerate(students_list):
                team = i // (self.num_tables * 2) + 1; table = (i % (self.num_tables * 2)) // 2 + 1; seat = "A" if i % 2 == 0 else "B"
                preview_content += f"Tổ {team} - Bàn {table} - Ghế {seat}: {student_name}\n"
                if seat == "B": preview_content += "-"*20 + "\n"
            preview_text.insert('1.0', preview_content); preview_text.config(state=tk.DISABLED)
            restore_btn.config(state=tk.NORMAL, command=lambda: (self.restore_arrangement(data['json']), history_window.destroy()))
            delete_btn.config(state=tk.NORMAL, command=lambda: self.delete_history_entry(data['id'], history_window))
        history_listbox.bind('<<ListboxSelect>>', on_history_select)

    def save_results(self):
        if not self.current_class_id or not self.students:
            messagebox.showwarning("Cảnh báo", "Không có dữ liệu để xuất hoặc chưa chọn lớp.", title="Lỗi")
            return
        
        try:
            arrangement_to_save = json.dumps(self.students)
            self.cursor.execute("INSERT INTO arrangements_history (class_id, arrangement, timestamp) VALUES (?, ?, datetime('now', 'localtime'))",
                                (self.current_class_id, arrangement_to_save))
            self.conn.commit()
            self.update_status(f"Đã lưu một bản ghi vào lịch sử của lớp {self.current_class_name}.")
        except sqlite3.Error as e:
            messagebox.showerror("Lỗi DB", f"Không thể lưu vào lịch sử: {e}")
            return
            
        wb = Workbook(); ws = wb.active
        ws.append(["Tổ", "Bàn", "Ghế", "Học sinh", "Giới tính", "Chiều cao", "Điểm TB", "Ghi Chú"])
        for i, student_name in enumerate(self.students):
            team = i // (self.num_tables * 2) + 1; table = (i % (self.num_tables * 2)) // 2 + 1; seat = i % 2 + 1
            student_info = next((s for s in self.students_data if s['Học sinh'] == student_name), None)
            if student_info: ws.append([team, table, seat, student_info['Học sinh'], student_info.get('Giới tính'), student_info.get('Chiều cao'), student_info.get('DiemTB'), student_info.get('GhiChu')])
        
        file_path = filedialog.asksaveasfilename(
            title=f"Lưu sơ đồ lớp {self.current_class_name}",
            defaultextension=".xlsx", 
            filetypes=[("Excel Files", "*.xlsx")]
        )
        if file_path:
            try:
                wb.save(file_path)
                messagebox.showinfo("Thành công", f"Đã xuất sơ đồ lớp {self.current_class_name} ra file Excel và lưu vào lịch sử!")
            except Exception as e:
                messagebox.showerror("Lỗi", f"Không thể lưu file: {e}")

    def save_as_image(self):
        if not self.students: messagebox.showwarning("Cảnh báo", "Không có sơ đồ để chụp ảnh!", title="Lỗi"); return
        file_path = filedialog.asksaveasfilename(defaultextension=".png", filetypes=[("PNG Files", "*.png")])
        if file_path:
            x = self.canvas.winfo_rootx(); y = self.canvas.winfo_rooty()
            x1 = x + self.canvas.winfo_width(); y1 = y + self.canvas.winfo_height()
            ImageGrab.grab().crop((x, y, x1, y1)).save(file_path)
            self.update_status(f"Đã lưu ảnh sơ đồ thành công: {file_path}")
            messagebox.showinfo("Thành công", "Lưu ảnh thành công!")

    def start_drag(self, event):
        items = self.canvas.find_overlapping(event.x, event.y, event.x, event.y);
        if items:
            closest_item = items[-1]
            if "rect" in self.canvas.gettags(closest_item):
                self.dragged_item = closest_item; self.canvas.tag_raise(self.dragged_item)
                text_item = self.text_positions.get(self.dragged_item)
                if text_item: self.canvas.tag_raise(text_item)
                self.update_status(f"Đang di chuyển học sinh: {self.seat_positions[self.dragged_item]['Học sinh']}")

    def on_drag(self, event):
        if self.dragged_item:
            canvas_x = self.canvas.canvasx(event.x); canvas_y = self.canvas.canvasy(event.y)
            text_item = self.text_positions.get(self.dragged_item)
            self.canvas.coords(self.dragged_item, canvas_x - self.RECT_WIDTH/2, canvas_y - self.RECT_HEIGHT/2, canvas_x + self.RECT_WIDTH/2, canvas_y + self.RECT_HEIGHT/2)
            if text_item: self.canvas.coords(text_item, canvas_x, canvas_y)

    def show_tooltip(self, event, rect_id):
        if self.tooltip: self.tooltip.destroy()
        student_info = self.seat_positions.get(rect_id)
        if not student_info: return
        text = (f"Tên: {student_info['Học sinh']}\n"
                f"Giới tính: {student_info.get('Giới tính', 'N/A')}\n"
                f"Chiều cao: {student_info.get('Chiều cao', 'N/A')} cm\n"
                f"Điểm TB: {student_info.get('DiemTB', 'N/A')}\n"
                f"Ghi chú: {student_info.get('GhiChu', 'Không')}")
        self.tooltip = ttk.Toplevel(self.root); self.tooltip.wm_overrideredirect(True); self.tooltip.wm_geometry(f"+{event.x_root + 15}+{event.y_root + 10}")
        ttk.Label(self.tooltip, text=text, justify=LEFT, padding=5, background="#FFFFE0", relief="solid", borderwidth=1).pack()

    def hide_tooltip(self, event):
        if self.tooltip: self.tooltip.destroy(); self.tooltip = None

    def choose_color(self, team_index, preview_label):
        color_code = colorchooser.askcolor(title=f"Chọn màu cho tổ {team_index + 1}")
        if color_code[1]:
            self.colors[team_index] = color_code[1]
            preview_label.config(bg=self.colors[team_index])
            if self.students: self.arrange_seats(); self.update_status(f"Đã đổi màu cho tổ {team_index + 1}.")
    
    def delete_history_entry(self, arrangement_id, window):
        if messagebox.askyesno("Xác nhận", f"Bạn có chắc muốn xóa vĩnh viễn mục lịch sử ID: {arrangement_id}?"):
            self.cursor.execute("DELETE FROM arrangements_history WHERE id = ?", (arrangement_id,)); self.conn.commit()
            self.update_status(f"Đã xóa mục lịch sử ID: {arrangement_id}.")
            window.destroy(); self.view_history()

    def clear_history(self, window=None):
        if not self.current_class_id: return
        if messagebox.askyesno("Xác nhận", f"Bạn có chắc chắn muốn xóa TOÀN BỘ lịch sử của lớp '{self.current_class_name}' không?"):
            self.cursor.execute("DELETE FROM arrangements_history WHERE class_id=?", (self.current_class_id,)); self.conn.commit()
            self.update_status(f"Đã xóa toàn bộ lịch sử của lớp {self.current_class_name}.")
            if window: window.destroy(); self.view_history()


if __name__ == "__main__":
    root = ttk.Window(themename="litera")
    app = SeatArrangementApp(root)
    root.mainloop()