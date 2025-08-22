# --- START OF FILE mainv11.py ---

import tkinter as tk
from tkinter import filedialog, messagebox, colorchooser, scrolledtext, TclError, font as tkfont, simpledialog
import ttkbootstrap as ttk
from ttkbootstrap.constants import *
import pandas as pd
from openpyxl import Workbook
import random
import sqlite3
import json
from datetime import datetime
from PIL import ImageGrab
import collections

# Yêu cầu cài đặt thư viện: pip install fpdf2 matplotlib
from fpdf import FPDF
from matplotlib.font_manager import findfont, FontProperties

# --- LỚP XỬ LÝ TẠO FILE PDF ---
# ... (Toàn bộ lớp PDFGenerator giữ nguyên, không thay đổi)
class PDFGenerator(FPDF):
    def __init__(self, class_name, teacher_name, school_year, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.class_name = class_name
        self.teacher_name = teacher_name
        self.school_year = school_year
        
        self.font_name = "SystemFont"
        try:
            regular_path = findfont(FontProperties(family="Times New Roman", style="normal", weight="normal"))
            self.add_font(self.font_name, "", regular_path, uni=True)
            
            bold_path = findfont(FontProperties(family="Times New Roman", style="normal", weight="bold"))
            self.add_font(self.font_name, "B", bold_path, uni=True)
            
            italic_path = findfont(FontProperties(family="Times New Roman", style="italic", weight="normal"))
            self.add_font(self.font_name, "I", italic_path, uni=True)
            
            print(f"Sử dụng font hệ thống 'Times New Roman' thành công.")
            
        except Exception:
            try:
                self.add_font("DejaVu", "", "DejaVuSans.ttf", uni=True)
                self.font_name = "DejaVu"
                print("Sử dụng font cục bộ: DejaVuSans.ttf")
            except RuntimeError as e:
                raise RuntimeError("Không tìm thấy font 'Times New Roman' trong hệ thống và cũng không có 'DejaVuSans.ttf' trong thư mục ứng dụng.") from e
        
        self.set_font(self.font_name, "", 14)

    def header(self):
        self.set_font(self.font_name, "B", 20)
        title = f"SƠ ĐỒ CHỖ NGỒI LỚP {self.class_name.upper()}"
        self.cell(0, 10, title, 0, 1, 'C')
        
        self.set_font(self.font_name, "", 11)
        info_line = f"Năm học: {self.school_year}  |  GVCN: {self.teacher_name}"
        self.cell(0, 8, info_line, 0, 1, 'C')
        self.ln(8)

    def footer(self):
        self.set_y(-15)
        self.set_font(self.font_name, "I", 8)
        export_date = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
        self.cell(0, 10, f'Trang {self.page_no()}/{{nb}}  |  Xuất ngày: {export_date}', 0, 0, 'C')

    def draw_seating_chart(self, students, num_teams, num_tables_per_team, colors_hex):
        margin = 10
        drawable_width = self.w - 2 * margin
        drawable_height = self.h - 45

        team_gap = 10
        total_team_width = drawable_width - (num_teams - 1) * team_gap
        team_width = total_team_width / num_teams
        seat_width = (team_width - 5) / 2

        max_seat_height = drawable_height / num_tables_per_team - 5
        seat_height = min(25, max_seat_height) 
        
        def hex_to_rgb(hex_color):
            hex_color = hex_color.lstrip('#')
            return tuple(int(hex_color[i:i+2], 16) for i in (0, 2, 4))

        for i, student_name in enumerate(students):
            team_idx = i // (num_tables_per_team * 2)
            table_in_team_idx = (i % (num_tables_per_team * 2)) // 2
            seat_idx = i % 2

            x = margin + team_idx * (team_width + team_gap) + seat_idx * (seat_width + 5)
            y = 40 + table_in_team_idx * (seat_height + 5)

            color_rgb = hex_to_rgb(colors_hex[team_idx % len(colors_hex)])
            self.set_fill_color(*color_rgb)
            self.rect(x, y, seat_width, seat_height, 'DF')

            self.set_xy(x, y + seat_height/2 - 4)
            self.set_font(self.font_name, "", 9)
            self.set_text_color(0, 0, 0)
            self.multi_cell(seat_width, 4, student_name, 0, 'C')


# <<--- MỚI: LỚP XỬ LÝ SẮP XẾP THÔNG MINH --- >>
# THAY THẾ TOÀN BỘ LỚP AdvancedSorter CŨ BẰNG LỚP NÀY

class AdvancedSorter:
    def __init__(self, students_data):
        self.students_data = students_data
        self.student_map = {s['Học sinh']: s for s in students_data}

    def _get_partner_name(self, arrangement, index):
        partner_idx = index + 1 if index % 2 == 0 else index - 1
        if 0 <= partner_idx < len(arrangement):
            return arrangement[partner_idx]
        return None

    def _calculate_score(self, arrangement):
        score = 0
        for i, student_name in enumerate(arrangement):
            student_data = self.student_map.get(student_name)
            if not student_data: continue

            partner_name = self._get_partner_name(arrangement, i)
            if not partner_name: continue

            # Phạt nặng nếu vi phạm "Không ngồi cạnh"
            if partner_name in student_data.get('dont_sit_next_to', []):
                score -= 100
            
            # Thưởng điểm nếu thỏa mãn "Muốn ngồi cạnh"
            if partner_name in student_data.get('sit_next_to', []):
                score += 50
        
        # Thêm các tiêu chí khác ở đây nếu muốn, ví dụ:
        # - Phạt nếu 2 bạn nam ngồi cạnh nhau (để ưu tiên nam nữ)
        # - Thưởng nếu bạn học giỏi ngồi cạnh bạn học yếu
        return score

    def arrange(self, initial_arrangement):
        best_arrangement = list(initial_arrangement)
        best_score = self._calculate_score(best_arrangement)

        # Thử tối ưu trong một số lần lặp nhất định
        # Tăng số lần lặp nếu có nhiều ràng buộc phức tạp
        num_iterations = 200 * len(best_arrangement) 

        for _ in range(num_iterations):
            current_arrangement = list(best_arrangement)
            
            # Chọn ngẫu nhiên 2 học sinh để thử hoán đổi
            idx1, idx2 = random.sample(range(len(current_arrangement)), 2)
            
            current_arrangement[idx1], current_arrangement[idx2] = current_arrangement[idx2], current_arrangement[idx1]
            
            new_score = self._calculate_score(current_arrangement)
            
            # Nếu sơ đồ mới tốt hơn, giữ lại nó
            if new_score > best_score:
                best_score = new_score
                best_arrangement = current_arrangement

        print(f"Tối ưu hóa hoàn tất. Điểm số cuối cùng: {best_score}")
        return best_arrangement
# --- LỚP ỨNG DỤNG CHÍNH ---
class SeatArrangementApp:
    # --- CÁC HẰNG SỐ CHO GIAO DIỆN ---
    RECT_WIDTH = 180
    RECT_HEIGHT = 75
    X_GAP = 100
    Y_GAP = 40
    TEAM_X_GAP = 70
    
    TEACHER_DESK_WIDTH = 250
    TEACHER_DESK_HEIGHT = 70
    
    def __init__(self, root):
        self.root = root
        self.root.title("Chương trình quản lý sắp xếp chỗ ngồi dành cho học sinh ")
        self.root.geometry("1366x1080")
        
        self.students_data = []
        self.students = []
        self.undo_stack = []
        self.redo_stack = []

        self.colors = ["#FFA07A", "#7FFFD4", "#87CEFA", "#FFD700", "#98FB98", "#F08080", "#E0FFFF"]
        self.seat_positions = {}
        self.text_positions = {}
        self.dragged_item = None
        self.tooltip = None
        self.num_teams = 4
        self.num_tables = 5
        
        self.current_class_id = None
        self.current_class_name = None
        self.is_dirty = False

        self.init_db()
        self.load_ui()
        self.root.protocol("WM_DELETE_WINDOW", self.on_closing)
        self.update_status("Chào mừng! Vui lòng chọn hoặc thêm một lớp học để bắt đầu.")

    def init_db(self):
        # ... (init_db giữ nguyên)
        self.conn = sqlite3.connect("seat_arrangements_multi_class.db")
        self.cursor = self.conn.cursor()
        
        self.cursor.execute('''
            CREATE TABLE IF NOT EXISTS classes (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL UNIQUE,
                students_data_json TEXT,
                current_arrangement_json TEXT,
                num_teams INTEGER,
                num_tables INTEGER,
                last_modified TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')

        self.cursor.execute('''
            CREATE TABLE IF NOT EXISTS arrangements_history (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                class_id INTEGER NOT NULL,
                arrangement TEXT,
                timestamp TEXT,
                FOREIGN KEY (class_id) REFERENCES classes (id) ON DELETE CASCADE
            )
        ''')
        self.conn.commit()

    def on_closing(self):
        # ... (on_closing giữ nguyên)
        if self.is_dirty:
            if not messagebox.askyesno("Thoát", "Bạn có những thay đổi chưa được lưu. Bạn có chắc chắn muốn thoát?"):
                return
        self.conn.close()
        self.root.destroy()
    # Đặt các hàm mới này vào bên trong lớp SeatArrangementApp

    # <<--- MỚI: CÁC HÀM QUẢN LÝ SĨ SỐ TRỰC TIẾP --- >>
    def open_roster_manager(self):
        """Mở cửa sổ quản lý sĩ số lớp học."""
        if not self.current_class_id:
            messagebox.showwarning("Chưa chọn lớp", "Vui lòng chọn một lớp để quản lý sĩ số.")
            return
        if not self.students_data:
            if not messagebox.askyesno("Chưa có dữ liệu", "Lớp này chưa có dữ liệu học sinh. Bạn có muốn tạo mới?"):
                return

        win = ttk.Toplevel(self.root)
        win.title(f"Quản lý Sĩ số - Lớp {self.current_class_name}")
        win.geometry("900x600")
        win.transient(self.root); win.grab_set()

        # --- Treeview để hiển thị dữ liệu ---
        tree_frame = ttk.Frame(win, padding=10)
        tree_frame.pack(fill=BOTH, expand=True)

        columns = ("name", "gender", "height", "score", "notes")
        tree = ttk.Treeview(tree_frame, columns=columns, show="headings")
        
        # Định nghĩa các cột
        tree.heading("name", text="Họ và Tên")
        tree.heading("gender", text="Giới tính")
        tree.heading("height", text="Chiều cao (cm)")
        tree.heading("score", text="Điểm TB")
        tree.heading("notes", text="Ghi Chú")
        
        tree.column("name", width=250)
        tree.column("gender", width=80, anchor=CENTER)
        tree.column("height", width=120, anchor=E)
        tree.column("score", width=100, anchor=E)
        tree.column("notes", width=250)

        # Thêm scrollbar
        scrollbar = ttk.Scrollbar(tree_frame, orient=VERTICAL, command=tree.yview, bootstyle='round')
        tree.configure(yscrollcommand=scrollbar.set)
        
        tree.pack(side=LEFT, fill=BOTH, expand=True)
        scrollbar.pack(side=RIGHT, fill=Y)

        # Đổ dữ liệu vào Treeview
        for student in self.students_data:
            values = (
                student.get('Học sinh', ''),
                student.get('Giới tính', 'Nữ'),
                student.get('Chiều cao', 160),
                student.get('DiemTB', 0.0),
                student.get('GhiChu', '')
            )
            tree.insert("", tk.END, values=values)

        # --- Các nút hành động ---
        btn_frame = ttk.Frame(win, padding=(10, 0, 10, 10))
        btn_frame.pack(fill=X)

        ttk.Button(btn_frame, text="➕ Thêm Học sinh", bootstyle="success-outline", command=lambda: self._add_student_to_roster(tree)).pack(side=LEFT, padx=5)
        ttk.Button(btn_frame, text="❌ Xóa Học sinh đã chọn", bootstyle="danger-outline", command=lambda: self._delete_student_from_roster(tree)).pack(side=LEFT, padx=5)
        ttk.Button(btn_frame, text="💾 Lưu và Đóng", bootstyle="primary", command=lambda: self._save_roster_changes(tree, win)).pack(side=RIGHT, padx=5)

        # --- Logic chỉnh sửa trực tiếp trên ô ---
        def on_double_click(event):
            region = tree.identify("region", event.x, event.y)
            if region != "cell": return

            item_id = tree.identify_row(event.y)
            column_id = tree.identify_column(event.x)
            
            x, y, width, height = tree.bbox(item_id, column_id)

            # Lấy giá trị hiện tại
            value = tree.item(item_id, "values")[int(column_id.replace('#','')) - 1]
            
            entry = ttk.Entry(tree_frame)
            entry.place(x=x, y=y, width=width, height=height)
            entry.insert(0, value)
            entry.focus_set()

            def on_save_edit(e):
                new_value = entry.get()
                tree.set(item_id, column_id, new_value)
                entry.destroy()
            
            entry.bind("<Return>", on_save_edit)
            entry.bind("<FocusOut>", on_save_edit)

        tree.bind("<Double-1>", on_double_click)

    def _add_student_to_roster(self, tree):
        """Thêm một dòng học sinh mới với giá trị mặc định vào Treeview."""
        default_values = ("Học sinh Mới", "Nữ", 160, 5.0, "")
        new_item = tree.insert("", tk.END, values=default_values)
        tree.selection_set(new_item)
        tree.see(new_item) # Cuộn đến dòng mới

    def _delete_student_from_roster(self, tree):
        """Xóa các học sinh đã được chọn khỏi Treeview."""
        selected_items = tree.selection()
        if not selected_items:
            messagebox.showwarning("Chưa chọn", "Vui lòng chọn ít nhất một học sinh để xóa.", parent=tree.winfo_toplevel())
            return
        
        if messagebox.askyesno("Xác nhận Xóa", f"Bạn có chắc chắn muốn xóa {len(selected_items)} học sinh đã chọn không?"):
            for item in selected_items:
                tree.delete(item)

    def _save_roster_changes(self, tree, window):
        """Lưu lại toàn bộ dữ liệu từ Treeview vào self.students_data."""
        new_students_data = []
        all_student_names = set()

        try:
            for item_id in tree.get_children():
                values = tree.item(item_id, "values")
                name = str(values[0]).strip()

                # --- VALIDATION DỮ LIỆU ---
                if not name:
                    raise ValueError("Tên học sinh không được để trống.")
                if name in all_student_names:
                    raise ValueError(f"Tên học sinh '{name}' bị trùng lặp.")
                all_student_names.add(name)

                gender = str(values[1])
                if gender not in ['Nam', 'Nữ']:
                    raise ValueError(f"Giới tính của '{name}' phải là 'Nam' hoặc 'Nữ'.")
                
                height = float(values[2])
                score = float(values[3])
                notes = str(values[4])

                # --- Tìm dữ liệu cũ để bảo toàn các ràng buộc ---
                old_student_info = next((s for s in self.students_data if s['Học sinh'] == name), {})
                
                new_students_data.append({
                    'Học sinh': name,
                    'Giới tính': gender,
                    'Chiều cao': height,
                    'DiemTB': score,
                    'GhiChu': notes,
                    'sit_next_to': old_student_info.get('sit_next_to', []),
                    'dont_sit_next_to': old_student_info.get('dont_sit_next_to', [])
                })
            
            # --- Cập nhật dữ liệu chính của ứng dụng ---
            self.students_data = new_students_data
            self.students = [s['Học sinh'] for s in self.students_data] # Cập nhật lại danh sách tên
            
            self._set_dirty()
            self._sync_data_and_ui() # Vẽ lại sơ đồ
            self.update_status(f"Đã cập nhật sĩ số lớp. Hiện có {len(self.students)} học sinh.")
            window.destroy()

        except ValueError as e:
            messagebox.showerror("Lỗi Dữ liệu", str(e), parent=window)

    def update_status(self, message):
        self.status_bar.config(text=message)
        
    def _set_dirty(self, dirty_status=True):
        # ... (_set_dirty giữ nguyên)
        if not self.current_class_id: return
        self.is_dirty = dirty_status
        
        current_text = self.class_selector.get()
        if dirty_status and not current_text.endswith('*'):
            self.class_selector.set(current_text + ' *')
        elif not dirty_status and current_text.endswith('*'):
            self.class_selector.set(current_text[:-2])

    def load_ui(self):
        # ... (Toàn bộ hàm load_ui giữ nguyên như phiên bản trước, chỉ thêm nút Gọi Tên)
        main_pane = ttk.PanedWindow(self.root, orient=HORIZONTAL)
        main_pane.pack(fill=BOTH, expand=True)

        control_panel = ttk.Frame(main_pane, padding=10, width=320)
        control_panel.pack_propagate(False)
        main_pane.add(control_panel, weight=1)

        canvas_container = ttk.Frame(main_pane, padding=(0, 10, 10, 0))
        main_pane.add(canvas_container, weight=4)

        class_lf = ttk.LabelFrame(control_panel, text=" Quản Lý Lớp Học ", padding=10, bootstyle=PRIMARY)
        class_lf.pack(fill=X, pady=(0, 15))
        
        self.class_selector = ttk.Combobox(class_lf, state="readonly", values=[])
        self.class_selector.pack(fill=X, pady=(0, 5))
        self.class_selector.bind("<<ComboboxSelected>>", self._on_class_selected)

        class_btn_frame = ttk.Frame(class_lf)
        class_btn_frame.pack(fill=X)
        ttk.Button(class_btn_frame, text="Thêm Lớp Mới", command=self._add_new_class, bootstyle="success-outline").pack(side=LEFT, expand=True, fill=X, padx=(0,2))
        ttk.Button(class_btn_frame, text="Xóa Lớp Này", command=self._delete_class, bootstyle="danger-outline").pack(side=LEFT, expand=True, fill=X, padx=(2,0))

        ttk.Button(class_lf, text="💾 LƯU TRẠNG THÁI LỚP", command=self._save_class_state, bootstyle="primary").pack(fill=X, pady=5)

        theme_frame = ttk.Frame(control_panel); theme_frame.pack(fill=X, pady=(0, 15))
        ttk.Label(theme_frame, text="Chế độ Sáng / Tối:").pack(side=LEFT)
        self.theme_var = tk.BooleanVar(value=False)
        theme_switch = ttk.Checkbutton(theme_frame, bootstyle="switch", variable=self.theme_var, command=self.toggle_theme); theme_switch.pack(side=LEFT, padx=10)

        lf1 = ttk.LabelFrame(control_panel, text=" 1. Tải Dữ Liệu Học Sinh ", padding=10, bootstyle=DEFAULT); lf1.pack(fill=X, pady=(0, 10))
        load_btn = ttk.Button(lf1, text="📂 Tải File Excel cho Lớp Này", command=self.load_students, bootstyle=DEFAULT); load_btn.pack(fill=X)
        roster_btn = ttk.Button(lf1, text="👨‍🎓 Quản lý Sĩ số Lớp", command=self.open_roster_manager, bootstyle="info")
        roster_btn.pack(fill=X, pady=(5, 0))
        lf2 = ttk.LabelFrame(control_panel, text=" 2. Cấu Hình Lớp ", padding=10, bootstyle=DEFAULT); lf2.pack(fill=X, pady=10)
        ttk.Label(lf2, text="Số tổ:").grid(row=0, column=0, padx=5, pady=5, sticky=W)
        self.team_spinbox = ttk.Spinbox(lf2, from_=1, to=20, width=5); self.team_spinbox.grid(row=0, column=1, padx=5, pady=5, sticky=W); self.team_spinbox.set(self.num_teams)
        ttk.Label(lf2, text="Số bàn mỗi tổ:").grid(row=1, column=0, padx=5, pady=5, sticky=W)
        self.table_spinbox = ttk.Spinbox(lf2, from_=1, to=30, width=5); self.table_spinbox.grid(row=1, column=1, padx=5, pady=5, sticky=W); self.table_spinbox.set(self.num_tables)
        apply_config_btn = ttk.Button(lf2, text="Áp dụng cấu hình", command=self.apply_team_table_config, bootstyle="secondary-outline"); apply_config_btn.grid(row=2, column=0, columnspan=2, pady=10, sticky=EW)

        lf3 = ttk.LabelFrame(control_panel, text=" 3. Sắp Xếp & Tùy Chỉnh ", padding=10, bootstyle=DEFAULT); lf3.pack(fill=X, pady=10)
        sort_btn = ttk.Button(lf3, text="🎲 Tùy chọn Sắp xếp", command=self.show_sort_options); sort_btn.pack(fill=X, pady=(0,5))
        undo_redo_frame = ttk.Frame(lf3); undo_redo_frame.pack(fill=X, pady=(0,5))
        self.undo_btn = ttk.Button(undo_redo_frame, text="↩️ Hoàn tác", command=self.undo, bootstyle="secondary-outline", state="disabled"); self.undo_btn.pack(side=LEFT, expand=True, fill=X, padx=(0,2))
        self.redo_btn = ttk.Button(undo_redo_frame, text="↪️ Làm lại", command=self.redo, bootstyle="secondary-outline", state="disabled"); self.redo_btn.pack(side=LEFT, expand=True, fill=X, padx=(2,0))

        lf4 = ttk.LabelFrame(control_panel, text=" 4. Xuất & Báo Cáo ", padding=10, bootstyle=DEFAULT); lf4.pack(fill=X, pady=10)
        save_excel_btn = ttk.Button(lf4, text="Xuất Sơ Đồ ra Excel", command=self.save_results, bootstyle="info-outline"); save_excel_btn.pack(fill=X, pady=(0, 5))
        save_pdf_btn = ttk.Button(lf4, text="🖨️ Xuất Sơ Đồ ra PDF (Để in)", command=self.export_to_pdf, bootstyle="info"); save_pdf_btn.pack(fill=X, pady=(5, 5))
        save_img_btn = ttk.Button(lf4, text="Chụp Ảnh Sơ Đồ", command=self.save_as_image, bootstyle="info-outline"); save_img_btn.pack(fill=X)
        
        lf5 = ttk.LabelFrame(control_panel, text=" Công Cụ Khác ", padding=10, bootstyle=DEFAULT); lf5.pack(fill=X, pady=10)
        color_btn = ttk.Button(lf5, text="🎨 Đổi Màu Tổ", command=self.change_team_colors, bootstyle="secondary-outline"); color_btn.pack(fill=X, pady=(0, 5))
        history_btn = ttk.Button(lf5, text="📜 Xem Lịch Sử Sắp Xếp của Lớp", command=self.view_history, bootstyle="secondary-outline"); history_btn.pack(fill=X, pady=(5,0))
        random_pick_btn = ttk.Button(lf5, text="✨ Gọi Tên Ngẫu Nhiên", command=self.pick_random_student, bootstyle="success")
        random_pick_btn.pack(fill=X, pady=(5, 0))
        
        ttk.Label(canvas_container, text="Sơ Đồ Lớp Học", font=("Arial", 16, "bold")).pack(pady=(0, 5))
        canvas_frame = ttk.Frame(canvas_container); canvas_frame.pack(fill=BOTH, expand=True)
        self.canvas = tk.Canvas(canvas_frame, bg='white', relief="solid", bd=1)
        h_scroll = ttk.Scrollbar(canvas_frame, orient=HORIZONTAL, command=self.canvas.xview, bootstyle="round"); v_scroll = ttk.Scrollbar(canvas_frame, orient=VERTICAL, command=self.canvas.yview, bootstyle="round")
        self.canvas.config(xscrollcommand=h_scroll.set, yscrollcommand=v_scroll.set)
        h_scroll.pack(side=BOTTOM, fill=X); v_scroll.pack(side=RIGHT, fill=Y); self.canvas.pack(side=LEFT, fill=BOTH, expand=True)
        
        self.status_bar = ttk.Label(self.root, text="Sẵn sàng", padding=5, font=("Arial", 9), anchor=W, bootstyle=INVERSE); self.status_bar.pack(side=BOTTOM, fill=X)
        
        self._load_class_list_to_selector()
        
    def export_to_pdf(self):
        # ... (export_to_pdf giữ nguyên)
        if not self.current_class_id or not self.students:
            messagebox.showwarning("Chưa có dữ liệu", "Vui lòng chọn một lớp và tải dữ liệu học sinh để xuất ra PDF.")
            return

        teacher_name = simpledialog.askstring("Thông tin bổ sung", "Nhập tên Giáo viên Chủ nhiệm:", parent=self.root)
        if teacher_name is None: return
        
        school_year = simpledialog.askstring("Thông tin bổ sung", "Nhập Năm học (VD: 2025-2026):", parent=self.root)
        if school_year is None: return

        file_path = filedialog.asksaveasfilename(
            title=f"Lưu sơ đồ PDF cho lớp {self.current_class_name}",
            defaultextension=".pdf",
            filetypes=[("PDF Files", "*.pdf")]
        )
        if not file_path:
            return

        self.update_status("Đang tạo file PDF, vui lòng đợi...")
        try:
            pdf = PDFGenerator(
                orientation='L', unit='mm', format='A4',
                class_name=self.current_class_name,
                teacher_name=teacher_name,
                school_year=school_year
            )
            pdf.alias_nb_pages()
            pdf.add_page()
            pdf.draw_seating_chart(
                students=self.students,
                num_teams=self.num_teams,
                num_tables_per_team=self.num_tables,
                colors_hex=self.colors
            )
            pdf.output(file_path)
            
            self.update_status(f"Đã xuất PDF thành công: {file_path}")
            messagebox.showinfo("Thành công", f"Đã xuất sơ đồ lớp {self.current_class_name} ra file PDF thành công!")
        
        except Exception as e:
            if isinstance(e, RuntimeError) and ("TTF Font file not found" in str(e) or "DejaVuSans.ttf" in str(e)):
                 messagebox.showerror(
                     "Lỗi Thiếu Font Chữ", 
                     "Không tìm thấy font chữ cần thiết để hỗ trợ tiếng Việt.\n\n"
                     "Giải pháp:\n"
                     "1. Đảm bảo font 'Times New Roman' đã được cài đặt trên máy của bạn.\n"
                     "2. (Nếu cách 1 không được) Tải và đặt file 'DejaVuSans.ttf' vào cùng thư mục với ứng dụng."
                 )
            else:
                messagebox.showerror("Lỗi", f"Không thể tạo file PDF: {e}")
            self.update_status("Lỗi khi tạo file PDF.")
            
    def _wrap_text(self, text, font_object, max_width):
        # ... (_wrap_text giữ nguyên)
        words = text.split()
        lines = []
        current_line = ""
        for word in words:
            separator = " " if current_line else ""
            test_line = current_line + separator + word
            
            if font_object.measure(test_line) <= max_width:
                current_line = test_line
            else:
                lines.append(current_line)
                current_line = word
        lines.append(current_line)
        return "\n".join(lines)

    def arrange_seats(self):
        self.canvas.delete("all")
        self.seat_positions.clear()
        self.text_positions.clear()
        
        if not self.students: 
            bbox = self.canvas.bbox("all")
            if bbox: self.canvas.config(scrollregion=bbox)
            return

        text_font = tkfont.Font(family="Arial", size=11, weight="bold")
        
        top_offset = self.TEACHER_DESK_HEIGHT + 100
        
        student_map = {s['Học sinh']: s for s in self.students_data}
        for i, student_name in enumerate(self.students):
            team_index = i // (self.num_tables * 2)
            table_in_team_index = (i % (self.num_tables * 2)) // 2
            seat_index = i % 2
            student_info = student_map.get(student_name)
            if not student_info: continue
            
            team_visual_width = self.RECT_WIDTH * 2 + self.X_GAP + self.TEAM_X_GAP
            x_start = 20 + team_index * team_visual_width
            
            y_start = top_offset + table_in_team_index * (self.RECT_HEIGHT + self.Y_GAP)
            
            x = x_start + seat_index * (self.RECT_WIDTH + self.X_GAP)
            y = y_start
            
            outline_color = "#E53935" if student_info.get('GhiChu') == 'Cần ngồi trước' else ("#FFFFFF" if self.theme_var.get() else "#000000")
            outline_width = 3 if student_info.get('GhiChu') == 'Cần ngồi trước' else 1
            text_color = "white" if self.theme_var.get() else "black"
            
            rect = self.canvas.create_rectangle(x, y, x + self.RECT_WIDTH, y + self.RECT_HEIGHT, fill=self.colors[team_index % len(self.colors)], tags="rect", outline=outline_color, width=outline_width)
            
            wrapped_name = self._wrap_text(student_name, text_font, self.RECT_WIDTH - 10)
            
            text = self.canvas.create_text(
                x + self.RECT_WIDTH / 2, y + self.RECT_HEIGHT / 2, 
                text=wrapped_name, 
                font=text_font, 
                tags="text", 
                fill=text_color,
                justify=tk.CENTER
            )

            self.seat_positions[rect] = student_info
            self.text_positions[rect] = text
            self.canvas.tag_bind(rect, "<Button-1>", self.start_drag)
            self.canvas.tag_bind(rect, "<B1-Motion>", self.on_drag)
            self.canvas.tag_bind(rect, "<ButtonRelease-1>", self.stop_drag)
            self.canvas.tag_bind(rect, "<Enter>", lambda e, r=rect: self.show_tooltip(e, r))
            self.canvas.tag_bind(rect, "<Leave>", self.hide_tooltip)
            # <<--- MỚI: BIND SỰ KIỆN CHUỘT PHẢI --- >>
            self.canvas.tag_bind(rect, "<Button-3>", lambda e, r=rect: self.show_context_menu(e, r))
        
        # --- VẼ BÀN GIÁO VIÊN ---
        if self.num_teams > 0:
            team_1_start_x = 20
            team_1_width = self.RECT_WIDTH * 2 + self.X_GAP
            team_1_center_x = team_1_start_x + team_1_width / 2

            teacher_desk_x = team_1_center_x - (self.TEACHER_DESK_WIDTH / 2)
            teacher_desk_y = 20 

            self.canvas.create_rectangle(
                teacher_desk_x, teacher_desk_y,
                teacher_desk_x + self.TEACHER_DESK_WIDTH,
                teacher_desk_y + self.TEACHER_DESK_HEIGHT,
                fill="#DEB887",
                outline=("white" if self.theme_var.get() else "black")
            )
            
            self.canvas.create_text(
                teacher_desk_x + self.TEACHER_DESK_WIDTH / 2,
                teacher_desk_y + self.TEACHER_DESK_HEIGHT / 2,
                text="Bàn Giáo Viên",
                font=text_font,
                fill=("white" if self.theme_var.get() else "black")
            )
            
        bbox = self.canvas.bbox("all")
        if bbox: self.canvas.config(scrollregion=bbox)

    # <<--- MỚI: CÁC HÀM XỬ LÝ RÀNG BUỘC CHỖ NGỒI --- >>
    def show_context_menu(self, event, rect_id):
        student_info = self.seat_positions.get(rect_id)
        if not student_info: return

        context_menu = tk.Menu(self.root, tearoff=0)
        context_menu.add_command(
            label=f"Ràng buộc cho: {student_info['Học sinh']}",
            command=lambda: self.open_constraint_window(student_info)
        )
        try:
            context_menu.tk_popup(event.x_root, event.y_root)
        finally:
            context_menu.grab_release()

    # Trong lớp SeatArrangementApp
# THAY THẾ TOÀN BỘ HÀM open_constraint_window CŨ BẰNG HÀM NÀY

    def open_constraint_window(self, student_info):
        student_name = student_info['Học sinh']
        # Tạo map từ tên sang index để truy cập nhanh
        other_students_map = {s['Học sinh']: i for i, s in enumerate(self.students_data) if s['Học sinh'] != student_name}
        other_students_list = list(other_students_map.keys())

        win = ttk.Toplevel(self.root)
        win.title(f"Ràng buộc cho {student_name}")
        win.transient(self.root); win.grab_set()
        
        main_frame = ttk.Frame(win, padding=15)
        main_frame.pack(fill=BOTH, expand=True)

        # --- Khung "Ngồi cạnh" ---
        sit_next_lf = ttk.LabelFrame(main_frame, text=" Chọn bạn MUỐN ngồi cạnh ", padding=10)
        sit_next_lf.pack(fill=BOTH, expand=True, pady=5)
        sit_next_lb = tk.Listbox(sit_next_lf, selectmode=tk.MULTIPLE, height=8, exportselection=False)
        # ... (scrollbar cho sit_next_lb giữ nguyên)
        sit_next_scroll = ttk.Scrollbar(sit_next_lf, orient=VERTICAL, command=sit_next_lb.yview, bootstyle='round')
        sit_next_lb.config(yscrollcommand=sit_next_scroll.set)
        sit_next_scroll.pack(side=RIGHT, fill=Y)
        sit_next_lb.pack(side=LEFT, fill=BOTH, expand=True)


        # --- Khung "Không ngồi cạnh" ---
        dont_sit_lf = ttk.LabelFrame(main_frame, text=" Chọn bạn KHÔNG MUỐN ngồi cạnh ", padding=10)
        dont_sit_lf.pack(fill=BOTH, expand=True, pady=5)
        dont_sit_lb = tk.Listbox(dont_sit_lf, selectmode=tk.MULTIPLE, height=8, exportselection=False)
        # ... (scrollbar cho dont_sit_lb giữ nguyên)
        dont_sit_scroll = ttk.Scrollbar(dont_sit_lf, orient=VERTICAL, command=dont_sit_lb.yview, bootstyle='round')
        dont_sit_lb.config(yscrollcommand=dont_sit_scroll.set)
        dont_sit_scroll.pack(side=RIGHT, fill=Y)
        dont_sit_lb.pack(side=LEFT, fill=BOTH, expand=True)

        # --- Hàm xử lý sự kiện chọn để ngăn mâu thuẫn ---
        def on_sit_next_select(event):
            selected_indices = sit_next_lb.curselection()
            for idx in selected_indices:
                selected_name = sit_next_lb.get(idx)
                if selected_name in other_students_map:
                    dont_sit_lb.selection_clear(other_students_list.index(selected_name))

        def on_dont_sit_select(event):
            selected_indices = dont_sit_lb.curselection()
            for idx in selected_indices:
                selected_name = dont_sit_lb.get(idx)
                if selected_name in other_students_map:
                    sit_next_lb.selection_clear(other_students_list.index(selected_name))
        
        sit_next_lb.bind("<<ListboxSelect>>", on_sit_next_select)
        dont_sit_lb.bind("<<ListboxSelect>>", on_dont_sit_select)

        # --- Nút bấm ---
        btn_frame = ttk.Frame(main_frame)
        btn_frame.pack(fill=X, pady=(10, 0))
        save_cmd = lambda: self._save_constraints(student_info, sit_next_lb, dont_sit_lb, win)
        ttk.Button(btn_frame, text="Lưu thay đổi", command=save_cmd, bootstyle=SUCCESS).pack(side=LEFT, expand=True, padx=5)
        ttk.Button(btn_frame, text="Hủy", command=win.destroy, bootstyle="secondary-outline").pack(side=LEFT, expand=True, padx=5)

        # --- Tải dữ liệu vào listbox ---
        current_sit_next = student_info.get('sit_next_to', [])
        current_dont_sit = student_info.get('dont_sit_next_to', [])
        
        for i, s_name in enumerate(other_students_list):
            sit_next_lb.insert(tk.END, s_name)
            dont_sit_lb.insert(tk.END, s_name)
            if s_name in current_sit_next:
                sit_next_lb.selection_set(i)
            if s_name in current_dont_sit:
                dont_sit_lb.selection_set(i)

    # Trong lớp SeatArrangementApp
# THAY THẾ TOÀN BỘ HÀM _save_constraints CŨ BẰNG HÀM NÀY

    def _save_constraints(self, student_info, sit_next_lb, dont_sit_lb, window):
        student_name = student_info['Học sinh']
        
        # Lấy danh sách mới từ lựa chọn của người dùng
        selected_sit_next_indices = sit_next_lb.curselection()
        new_sit_next_list = {sit_next_lb.get(i) for i in selected_sit_next_indices}
        
        selected_dont_sit_indices = dont_sit_lb.curselection()
        new_dont_sit_list = {dont_sit_lb.get(i) for i in selected_dont_sit_indices}

        # --- LOGIC CẬP NHẬT TƯƠNG HỖ ---
        for other_student in self.students_data:
            other_name = other_student['Học sinh']
            if other_name == student_name: continue

            # Lấy danh sách cũ của học sinh kia
            other_sit_next = set(other_student.get('sit_next_to', []))
            other_dont_sit = set(other_student.get('dont_sit_next_to', []))

            # Xử lý "Muốn ngồi cạnh"
            if other_name in new_sit_next_list:
                other_sit_next.add(student_name) # Thêm student_name vào danh sách muốn ngồi cạnh của other_name
            else:
                other_sit_next.discard(student_name) # Xóa nếu không còn được chọn

            # Xử lý "Không muốn ngồi cạnh"
            if other_name in new_dont_sit_list:
                other_dont_sit.add(student_name) # Thêm student_name vào danh sách không muốn của other_name
            else:
                other_dont_sit.discard(student_name)

            other_student['sit_next_to'] = sorted(list(other_sit_next))
            other_student['dont_sit_next_to'] = sorted(list(other_dont_sit))

        # Cập nhật cho chính học sinh đang được sửa
        for student in self.students_data:
            if student['Học sinh'] == student_name:
                student['sit_next_to'] = sorted(list(new_sit_next_list))
                student['dont_sit_next_to'] = sorted(list(new_dont_sit_list))
                break
        
        self._set_dirty()
        self.update_status(f"Đã cập nhật ràng buộc cho {student_name} và các bạn liên quan.")
        messagebox.showinfo("Thành công", "Đã lưu ràng buộc. Hãy nhấn 'Sắp xếp thông minh' để áp dụng.", parent=window)
        window.destroy()

    def pick_random_student(self):
        # ... (pick_random_student giữ nguyên)
        if not self.students:
            messagebox.showwarning("Chưa có dữ liệu", "Không có học sinh nào trong danh sách để lựa chọn.")
            return

        random_student_name = random.choice(self.students)
        self.update_status(f"Đang chọn ngẫu nhiên... Kết quả là: {random_student_name}!")

        target_rect_id = None
        for rect_id, student_info in self.seat_positions.items():
            if student_info['Học sinh'] == random_student_name:
                target_rect_id = rect_id
                break
        
        if target_rect_id:
            original_color = self.canvas.itemcget(target_rect_id, "fill")
            highlight_color = "#FFD700"

            def revert_highlight():
                self.canvas.itemconfig(target_rect_id, fill=original_color)

            self.canvas.itemconfig(target_rect_id, fill=highlight_color)
            
            text_item = self.text_positions.get(target_rect_id)
            if text_item:
                self.canvas.tag_raise(target_rect_id)
                self.canvas.tag_raise(text_item)
            
            messagebox.showinfo(
                "Học sinh được chọn",
                f"🌟 Chúc mừng em: {random_student_name}! 🌟",
                parent=self.root
            )

            self.root.after(2000, revert_highlight)
        else:
            messagebox.showinfo(
                 "Học sinh được chọn",
                f"Học sinh được chọn là: {random_student_name} (không tìm thấy trên sơ đồ)."
            )

    def toggle_theme(self):
        # ... (toggle_theme giữ nguyên)
        if self.theme_var.get():
            self.root.style.theme_use('darkly')
            self.canvas.config(bg="#303030")
        else:
            self.root.style.theme_use('litera')
            self.canvas.config(bg="white")
        if self.students: self.arrange_seats()

    def _save_state_for_undo(self):
        # ... (các hàm undo/redo giữ nguyên)
        if not self.students:
            return
        self.undo_stack.append(self.students.copy())
        self.redo_stack.clear()
        self._update_undo_redo_buttons()

    def _update_undo_redo_buttons(self):
        self.undo_btn.config(state="normal" if self.undo_stack else "disabled")
        self.redo_btn.config(state="normal" if self.redo_stack else "disabled")

    def undo(self):
        if not self.undo_stack: return
        self.redo_stack.append(self.students.copy())
        self.students = self.undo_stack.pop()
        self._set_dirty(); self._sync_data_and_ui()
        self.update_status("Đã hoàn tác hành động.")

    def redo(self):
        if not self.redo_stack: return
        self.undo_stack.append(self.students.copy())
        self.students = self.redo_stack.pop()
        self._set_dirty(); self._sync_data_and_ui()
        self.update_status("Đã làm lại hành động.")

    def _sync_data_and_ui(self):
        if self.students_data:
            student_map = {s['Học sinh']: s for s in self.students_data}
            self.students_data = [student_map[name] for name in self.students if name in student_map]
        
        self.arrange_seats()
        self._update_undo_redo_buttons()

    # <<--- MỚI: CẬP NHẬT HÀM APPLY_SORT ĐỂ SỬ DỤNG ADVANCEDSORTER --- >>
    def apply_sort(self, window):
        self._save_state_for_undo()
        sort_method = self.sort_method_var.get()
        
        base_arrangement = self.students.copy()
        
        if sort_method == "random": 
            random.shuffle(base_arrangement)
        elif sort_method == "height":
            df = pd.DataFrame(self.students_data)
            base_arrangement = df.sort_values(by='Chiều cao', ascending=False)['Học sinh'].tolist()
        elif sort_method == "gender":
            df = pd.DataFrame(self.students_data)
            males = df[df['Giới tính'] == 'Nam']['Học sinh'].tolist(); random.shuffle(males)
            females = df[df['Giới tính'] == 'Nữ']['Học sinh'].tolist(); random.shuffle(females)
            base_arrangement = []
            i, j = 0, 0
            while i < len(males) or j < len(females):
                if i < len(males): base_arrangement.append(males[i]); i += 1
                if j < len(females): base_arrangement.append(females[j]); j += 1
        
        # Luôn áp dụng sắp xếp thông minh sau các bước trên nếu được chọn
        if sort_method == "intelligent":
            # Tạo một sắp xếp cơ sở tốt (ví dụ theo chiều cao) trước khi tối ưu
            df = pd.DataFrame(self.students_data)
            base_arrangement = df.sort_values(by='Chiều cao', ascending=False)['Học sinh'].tolist()
            
            sorter = AdvancedSorter(self.students_data)
            self.students = sorter.arrange(base_arrangement)
        else:
            self.students = base_arrangement

        self._set_dirty()
        self._sync_data_and_ui()
        self.update_status(f"Đã áp dụng sắp xếp theo phương pháp: {sort_method}.")
        window.destroy()

    def stop_drag(self, event):
        # ... (stop_drag giữ nguyên)
        if not self.dragged_item: return
        drop_x = self.canvas.canvasx(event.x)
        drop_y = self.canvas.canvasy(event.y)
        overlapping_items = self.canvas.find_overlapping(drop_x, drop_y, drop_x, drop_y)
        target_item = None
        for item in overlapping_items:
            if item in self.seat_positions and item != self.dragged_item:
                target_item = item
                break
        if target_item:
            self._save_state_for_undo()
            
            dragged_info = self.seat_positions[self.dragged_item]
            target_info = self.seat_positions[target_item]
            dragged_index = self.students.index(dragged_info['Học sinh'])
            target_index = self.students.index(target_info['Học sinh'])
            self.students[dragged_index], self.students[target_index] = self.students[target_index], self.students[dragged_index]
            self.update_status(f"Đã hoán đổi vị trí: {dragged_info['Học sinh']} và {target_info['Học sinh']}.")
            self._set_dirty()
            self._sync_data_and_ui()
        else: 
            self.update_status("Thao tác kéo thả bị hủy.")
            self.arrange_seats()
            
        self.dragged_item = None

    def restore_arrangement(self, arrangement_json):
        # ... (restore_arrangement giữ nguyên)
        if not self.students_data:
            messagebox.showerror("Lỗi", "Không thể khôi phục khi chưa có danh sách học sinh.")
            return
        
        self._save_state_for_undo()
        
        restored_student_names = json.loads(arrangement_json)
        if set(restored_student_names) != {s['Học sinh'] for s in self.students_data}:
            messagebox.showwarning("Cảnh báo", "Danh sách học sinh trong lịch sử không khớp với danh sách hiện tại.")
        
        self.students = restored_student_names
        
        if self.num_teams * self.num_tables * 2 < len(self.students):
            messagebox.showwarning("Cảnh báo", f"Số ghế hiện tại không đủ.")
        
        self._set_dirty()
        self._sync_data_and_ui()
        self.update_status("Đã khôi phục sơ đồ từ lịch sử.")
        messagebox.showinfo("Thành công", "Đã khôi phục sơ đồ từ lịch sử!")
    
    def _load_class_list_to_selector(self):
        # ... (_load_class_list_to_selector giữ nguyên)
        try:
            self.cursor.execute("SELECT id, name FROM classes ORDER BY name")
            self.class_list = self.cursor.fetchall()
            class_names = [row[1] for row in self.class_list]
            self.class_selector['values'] = class_names
            if not class_names:
                self.class_selector.set("Chưa có lớp nào. Hãy thêm một lớp.")
        except sqlite3.Error as e:
            messagebox.showerror("Lỗi Database", f"Không thể tải danh sách lớp: {e}")

    def _on_class_selected(self, event=None):
        # ... (_on_class_selected giữ nguyên)
        if self.is_dirty:
            if not messagebox.askyesno("Cảnh báo", "Bạn có thay đổi chưa lưu ở lớp hiện tại. Bạn có muốn chuyển lớp và hủy các thay đổi đó?"):
                self.class_selector.set(self.current_class_name)
                return
        
        selected_name = self.class_selector.get()
        selected_class = next((c for c in self.class_list if c[1] == selected_name), None)

        if selected_class:
            self.current_class_id, self.current_class_name = selected_class
            try:
                self.cursor.execute("SELECT * FROM classes WHERE id=?", (self.current_class_id,))
                class_data = self.cursor.fetchone()
                
                self.students_data = []
                self.students = []
                
                if class_data and class_data[2]:
                    self.students_data = json.loads(class_data[2])
                if class_data and class_data[3]:
                    self.students = json.loads(class_data[3])
                
                self.num_teams = class_data[4] if class_data[4] else 1
                self.num_tables = class_data[5] if class_data[5] else 3
                
                self.team_spinbox.set(self.num_teams)
                self.table_spinbox.set(self.num_tables)
                
                self.undo_stack.clear()
                self.redo_stack.clear()
                self._set_dirty(False)
                self._sync_data_and_ui()
                self.update_status(f"Đã tải dữ liệu cho lớp: {self.current_class_name}")

            except sqlite3.Error as e:
                messagebox.showerror("Lỗi Database", f"Không thể tải dữ liệu lớp: {e}")
            except (json.JSONDecodeError, TypeError):
                messagebox.showwarning("Dữ liệu lỗi", "Dữ liệu của lớp này có thể bị lỗi. Vui lòng tải lại file Excel.")
                self._clear_canvas_and_data()

    def _add_new_class(self):
        # ... (_add_new_class giữ nguyên)
        class_name = simpledialog.askstring("Thêm Lớp Mới", "Nhập tên lớp học:", parent=self.root)
        if not class_name or not class_name.strip(): return
        class_name = class_name.strip()
        try:
            self.cursor.execute("INSERT INTO classes (name) VALUES (?)", (class_name,))
            self.conn.commit()
            self.update_status(f"Đã tạo lớp mới: {class_name}")
            self._load_class_list_to_selector()
            self.class_selector.set(class_name)
            self._on_class_selected()
        except sqlite3.IntegrityError:
            messagebox.showerror("Lỗi", f"Tên lớp '{class_name}' đã tồn tại. Vui lòng chọn tên khác.")
        except sqlite3.Error as e:
            messagebox.showerror("Lỗi Database", f"Không thể tạo lớp mới: {e}")

    def _save_class_state(self):
        # ... (_save_class_state giữ nguyên)
        if not self.current_class_id:
            messagebox.showwarning("Chưa chọn lớp", "Vui lòng chọn một lớp để lưu.")
            return

        try:
            students_data_json = json.dumps(self.students_data, ensure_ascii=False, indent=2)
            current_arrangement_json = json.dumps(self.students, ensure_ascii=False)
            self.num_teams = int(self.team_spinbox.get())
            self.num_tables = int(self.table_spinbox.get())
            
            self.cursor.execute("""
                UPDATE classes 
                SET students_data_json=?, current_arrangement_json=?, num_teams=?, num_tables=?, last_modified=CURRENT_TIMESTAMP
                WHERE id=?
            """, (students_data_json, current_arrangement_json, self.num_teams, self.num_tables, self.current_class_id))
            self.conn.commit()
            self._set_dirty(False)
            self.update_status(f"Đã lưu thành công trạng thái của lớp: {self.current_class_name}")
            messagebox.showinfo("Thành công", f"Đã lưu trạng thái của lớp '{self.current_class_name}'.")
        except (sqlite3.Error, TclError, ValueError) as e:
             messagebox.showerror("Lỗi", f"Không thể lưu trạng thái lớp: {e}")

    def _delete_class(self):
        # ... (_delete_class giữ nguyên)
        if not self.current_class_id:
            messagebox.showwarning("Chưa chọn lớp", "Vui lòng chọn một lớp để xóa.")
            return
        
        if messagebox.askyesno("Xác nhận Xóa", f"Bạn có chắc chắn muốn xóa vĩnh viễn lớp '{self.current_class_name}' và toàn bộ lịch sử của nó không?"):
            try:
                class_name_to_delete = self.current_class_name
                self.cursor.execute("DELETE FROM classes WHERE id=?", (self.current_class_id,))
                self.conn.commit()
                self._clear_canvas_and_data()
                self.current_class_id = None
                self.current_class_name = None
                self._load_class_list_to_selector()
                self.update_status(f"Đã xóa lớp: {class_name_to_delete}")
            except sqlite3.Error as e:
                messagebox.showerror("Lỗi Database", f"Không thể xóa lớp: {e}")

    def _clear_canvas_and_data(self):
        # ... (_clear_canvas_and_data giữ nguyên)
        self.students = []; self.students_data = []
        self.undo_stack.clear(); self.redo_stack.clear()
        self._sync_data_and_ui()
        
    def apply_team_table_config(self):
        # ... (apply_team_table_config giữ nguyên)
        try:
            self.num_teams = int(self.team_spinbox.get())
            self.num_tables = int(self.table_spinbox.get())
            if self.num_teams <= 0 or self.num_tables <= 0:
                raise ValueError("Số tổ và số bàn phải lớn hơn 0!")
            total_seats = self.num_teams * self.num_tables * 2
            if self.students and total_seats < len(self.students):
                messagebox.showwarning("Cảnh báo", f"Tổng số ghế ({total_seats}) nhỏ hơn số học sinh ({len(self.students)})!", title="Cấu hình không hợp lệ")
                return
            if self.students:
                self.arrange_seats()
                self._set_dirty()
            self.update_status(f"Đã thiết lập: {self.num_teams} tổ, {self.num_tables} bàn mỗi tổ.")
        except (ValueError, TclError):
            messagebox.showerror("Lỗi", "Số tổ và số bàn phải là các số hợp lệ.", title="Lỗi dữ liệu")
            self.update_status("Lỗi: Dữ liệu cấu hình không hợp lệ.")

    def show_sort_options(self):
        # ... (show_sort_options giữ nguyên)
        if not self.students_data:
            messagebox.showwarning("Cảnh báo", "Vui lòng nhập danh sách học sinh trước!", title="Chưa có dữ liệu")
            return
        
        sort_window = ttk.Toplevel(self.root)
        sort_window.title("Tùy chọn sắp xếp")
        sort_window.transient(self.root); sort_window.grab_set()
        
        container = ttk.Frame(sort_window, padding=20); container.pack(fill=BOTH, expand=True)
        ttk.Label(container, text="Chọn phương pháp sắp xếp:", font=("Arial", 12)).pack(pady=10)
        self.sort_method_var = tk.StringVar(value="intelligent")

        style = ttk.Style(); style.configure('TRadiobutton', font=('Arial', 10), padding=(0,5))
        
        ttk.Radiobutton(container, text="Thông minh (ưu tiên ràng buộc)", variable=self.sort_method_var, value="intelligent", style='TRadiobutton').pack(fill=X, padx=20, pady=2)
        ttk.Radiobutton(container, text="Ngẫu nhiên", variable=self.sort_method_var, value="random", style='TRadiobutton').pack(fill=X, padx=20, pady=2)
        ttk.Radiobutton(container, text="Xen kẽ nam-nữ", variable=self.sort_method_var, value="gender", style='TRadiobutton').pack(fill=X, padx=20, pady=2)
        ttk.Radiobutton(container, text="Theo chiều cao (cao trước, thấp sau)", variable=self.sort_method_var, value="height", style='TRadiobutton').pack(fill=X, padx=20, pady=2)
        
        button_frame = ttk.Frame(container); button_frame.pack(pady=20)
        ttk.Button(button_frame, text="Áp dụng", command=lambda: self.apply_sort(sort_window), bootstyle=SUCCESS).pack(side=LEFT, padx=10)
        ttk.Button(button_frame, text="Hủy", command=sort_window.destroy, bootstyle="secondary-outline").pack(side=LEFT, padx=10)
        
    # <<--- MỚI: CẬP NHẬT HÀM LOAD_STUDENTS ĐỂ ĐỌC RÀNG BUỘC TỪ EXCEL --- >>
    def load_students(self):
        if not self.current_class_id:
            messagebox.showwarning("Chưa chọn lớp", "Vui lòng chọn hoặc thêm một lớp trước khi tải danh sách học sinh.")
            return
        
        file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx;*.xls")])
        if not file_path: return
        try:
            df = pd.read_excel(file_path)
            required_columns = ['Học sinh', 'Giới tính', 'Chiều cao', 'DiemTB']
            if not all(col in df.columns for col in required_columns):
                raise ValueError("File Excel phải có các cột: 'Học sinh', 'Giới tính', 'Chiều cao', 'DiemTB'.\nCác cột tùy chọn: 'GhiChu', 'KhongNgoiCanh', 'NgoiCanh'.")
            
            # Chuẩn hóa các cột ràng buộc
            df['GhiChu'] = df.get('GhiChu', pd.Series(index=df.index, dtype='object')).fillna('')
            df['KhongNgoiCanh'] = df.get('KhongNgoiCanh', pd.Series(index=df.index, dtype='object')).fillna('')
            df['NgoiCanh'] = df.get('NgoiCanh', pd.Series(index=df.index, dtype='object')).fillna('')

            for index, row in df.iterrows():
                if pd.isna(row['Chiều cao']) or not isinstance(row['Chiều cao'], (int, float)): raise ValueError(f"Dòng {index + 2}: 'Chiều cao' phải là một con số.")
                if pd.isna(row['DiemTB']) or not isinstance(row['DiemTB'], (int, float)): raise ValueError(f"Dòng {index + 2}: 'DiemTB' phải là một con số.")
                if row['Giới tính'] not in ['Nam', 'Nữ']: raise ValueError(f"Dòng {index + 2}: 'Giới tính' phải là 'Nam' hoặc 'Nữ'.")

            self.students_data = df.to_dict('records')
            
            # Chuyển đổi chuỗi ràng buộc thành danh sách
            for student in self.students_data:
                dont_sit_str = student.get('KhongNgoiCanh', '')
                student['dont_sit_next_to'] = [name.strip() for name in str(dont_sit_str).split(',') if name.strip()]
                
                sit_next_str = student.get('NgoiCanh', '')
                student['sit_next_to'] = [name.strip() for name in str(sit_next_str).split(',') if name.strip()]

            self.students = df['Học sinh'].tolist()
            
            self.update_status(f"Đã tải {len(self.students)} HS cho lớp {self.current_class_name}. Nhấn 'Lưu Trạng Thái' để ghi nhớ.")
            random.shuffle(self.students)
            self.undo_stack.clear(); self.redo_stack.clear()
            self._set_dirty()
            self._sync_data_and_ui()
        except Exception as e:
            messagebox.showerror("Lỗi tải file", f"Đã xảy ra lỗi: {str(e)}", title="Lỗi")
            self.update_status(f"Lỗi tải file Excel cho lớp {self.current_class_name}.")

    # ... (Các hàm còn lại từ change_team_colors đến hết giữ nguyên không thay đổi)
    def change_team_colors(self):
        if self.num_teams == 0:
            messagebox.showinfo("Thông báo", "Vui lòng thiết lập số tổ trước.")
            return
        
        color_window = ttk.Toplevel(self.root)
        color_window.title("Đổi màu tổ")
        
        for i in range(self.num_teams):
            frame = ttk.Frame(color_window, padding=5)
            frame.pack(fill=X)
            while i >= len(self.colors): self.colors.append("#FFFFFF")
            color_preview = tk.Label(frame, text="    ", bg=self.colors[i], relief="solid", borderwidth=1)
            color_preview.pack(side=LEFT, padx=5)
            ttk.Label(frame, text=f"Màu cho tổ {i + 1}:").pack(side=LEFT, padx=5)
            ttk.Button(frame, text="Chọn màu", bootstyle="outline", command=lambda idx=i, p=color_preview: self.choose_color(idx, p)).pack(side=LEFT, padx=5)

    def view_history(self):
        if not self.current_class_id:
            messagebox.showwarning("Chưa chọn lớp", "Vui lòng chọn một lớp để xem lịch sử.")
            return

        history_window = ttk.Toplevel(self.root)
        history_window.title(f"Lịch sử sắp xếp - Lớp {self.current_class_name}")
        history_window.geometry("1050x1000")
        history_window.transient(self.root); history_window.grab_set()
        
        left_frame = ttk.Frame(history_window, padding=5); left_frame.pack(side=LEFT, fill=Y)
        right_frame = ttk.Frame(history_window, padding=5); right_frame.pack(side=RIGHT, fill=BOTH, expand=True)
        ttk.Label(left_frame, text="Các phiên đã lưu", font=("Arial", 12, "bold")).pack(pady=5)
        list_frame = ttk.Frame(left_frame); list_frame.pack(fill=BOTH, expand=True)
        history_listbox = tk.Listbox(list_frame, width=30, font=("Arial", 10)); scrollbar = ttk.Scrollbar(list_frame, orient="vertical", command=history_listbox.yview, bootstyle="round")
        history_listbox.config(yscrollcommand=scrollbar.set); scrollbar.pack(side=RIGHT, fill=Y); history_listbox.pack(side=LEFT, fill=BOTH, expand=True)
        ttk.Label(right_frame, text="Xem trước Sơ đồ", font=("Arial", 12, "bold")).pack(pady=5)
        preview_text = scrolledtext.ScrolledText(right_frame, wrap=tk.WORD, state=tk.DISABLED, font=("Arial", 10)); preview_text.pack(fill=BOTH, expand=True, pady=5)
        button_frame = ttk.Frame(right_frame); button_frame.pack(fill=X, pady=5)
        restore_btn = ttk.Button(button_frame, text="Khôi phục phiên này", state=tk.DISABLED, bootstyle=SUCCESS); restore_btn.pack(side=LEFT, padx=5, expand=True, fill=X)
        delete_btn = ttk.Button(button_frame, text="Xóa mục này", state=tk.DISABLED, bootstyle=DANGER); delete_btn.pack(side=LEFT, padx=5, expand=True, fill=X)
        ttk.Button(right_frame, text="Xóa toàn bộ lịch sử của lớp này", bootstyle="danger-outline", command=lambda: self.clear_history(history_window)).pack(fill=X, pady=10)

        self.cursor.execute("SELECT id, timestamp, arrangement FROM arrangements_history WHERE class_id=? ORDER BY timestamp DESC", (self.current_class_id,))
        arrangements_data = self.cursor.fetchall()
        
        if not arrangements_data: history_listbox.insert(tk.END, "Không có lịch sử nào."); return
        history_map = {index: {'id': arr_id, 'json': arr_json} for index, (arr_id, _, arr_json) in enumerate(arrangements_data)}
        for index, (arr_id, timestamp, _) in enumerate(arrangements_data): history_listbox.insert(tk.END, f"ID {arr_id}: {timestamp}")
        def on_history_select(event):
            selected_indices = history_listbox.curselection()
            if not selected_indices: return
            selected_index = selected_indices[0]; data = history_map.get(selected_index)
            if not data: return
            preview_text.config(state=tk.NORMAL); preview_text.delete('1.0', tk.END)
            students_list = json.loads(data['json'])
            preview_content = f"--- SƠ ĐỒ CHI TIẾT (ID: {data['id']}) ---\n\n"
            for i, student_name in enumerate(students_list):
                team = i // (self.num_tables * 2) + 1; table = (i % (self.num_tables * 2)) // 2 + 1; seat = "A" if i % 2 == 0 else "B"
                preview_content += f"Tổ {team} - Bàn {table} - Ghế {seat}: {student_name}\n"
                if seat == "B": preview_content += "-"*20 + "\n"
            preview_text.insert('1.0', preview_content); preview_text.config(state=tk.DISABLED)
            restore_btn.config(state=tk.NORMAL, command=lambda: (self.restore_arrangement(data['json']), history_window.destroy()))
            delete_btn.config(state=tk.NORMAL, command=lambda: self.delete_history_entry(data['id'], history_window))
        history_listbox.bind('<<ListboxSelect>>', on_history_select)

    def save_results(self):
        if not self.current_class_id or not self.students:
            messagebox.showwarning("Cảnh báo", "Không có dữ liệu để xuất hoặc chưa chọn lớp.", title="Lỗi")
            return
        
        try:
            arrangement_to_save = json.dumps(self.students)
            self.cursor.execute("INSERT INTO arrangements_history (class_id, arrangement, timestamp) VALUES (?, ?, datetime('now', 'localtime'))",
                                (self.current_class_id, arrangement_to_save))
            self.conn.commit()
            self.update_status(f"Đã lưu một bản ghi vào lịch sử của lớp {self.current_class_name}.")
        except sqlite3.Error as e:
            messagebox.showerror("Lỗi DB", f"Không thể lưu vào lịch sử: {e}")
            return
            
        wb = Workbook(); ws = wb.active
        ws.append(["Tổ", "Bàn", "Ghế", "Học sinh", "Giới tính", "Chiều cao", "Điểm TB", "Ghi Chú"])
        for i, student_name in enumerate(self.students):
            team = i // (self.num_tables * 2) + 1; table = (i % (self.num_tables * 2)) // 2 + 1; seat = i % 2 + 1
            student_info = next((s for s in self.students_data if s['Học sinh'] == student_name), None)
            if student_info: ws.append([team, table, seat, student_info['Học sinh'], student_info.get('Giới tính'), student_info.get('Chiều cao'), student_info.get('DiemTB'), student_info.get('GhiChu')])
        
        file_path = filedialog.asksaveasfilename(
            title=f"Lưu sơ đồ lớp {self.current_class_name}",
            defaultextension=".xlsx", 
            filetypes=[("Excel Files", "*.xlsx")]
        )
        if file_path:
            try:
                wb.save(file_path)
                messagebox.showinfo("Thành công", f"Đã xuất sơ đồ lớp {self.current_class_name} ra file Excel và lưu vào lịch sử!")
            except Exception as e:
                messagebox.showerror("Lỗi", f"Không thể lưu file: {e}")

    def save_as_image(self):
        if not self.students: messagebox.showwarning("Cảnh báo", "Không có sơ đồ để chụp ảnh!", title="Lỗi"); return
        file_path = filedialog.asksaveasfilename(defaultextension=".png", filetypes=[("PNG Files", "*.png")])
        if file_path:
            x = self.canvas.winfo_rootx(); y = self.canvas.winfo_rooty()
            x1 = x + self.canvas.winfo_width(); y1 = y + self.canvas.winfo_height()
            ImageGrab.grab().crop((x, y, x1, y1)).save(file_path)
            self.update_status(f"Đã lưu ảnh sơ đồ thành công: {file_path}")
            messagebox.showinfo("Thành công", "Lưu ảnh thành công!")

    def start_drag(self, event):
        items = self.canvas.find_overlapping(event.x, event.y, event.x, event.y);
        if items:
            closest_item = items[-1]
            if "rect" in self.canvas.gettags(closest_item):
                self.dragged_item = closest_item; self.canvas.tag_raise(self.dragged_item)
                text_item = self.text_positions.get(self.dragged_item)
                if text_item: self.canvas.tag_raise(text_item)
                self.update_status(f"Đang di chuyển học sinh: {self.seat_positions[self.dragged_item]['Học sinh']}")

    def on_drag(self, event):
        if self.dragged_item:
            canvas_x = self.canvas.canvasx(event.x); canvas_y = self.canvas.canvasy(event.y)
            text_item = self.text_positions.get(self.dragged_item)
            self.canvas.coords(self.dragged_item, canvas_x - self.RECT_WIDTH/2, canvas_y - self.RECT_HEIGHT/2, canvas_x + self.RECT_WIDTH/2, canvas_y + self.RECT_HEIGHT/2)
            if text_item: self.canvas.coords(text_item, canvas_x, canvas_y)

    def show_tooltip(self, event, rect_id):
        if self.tooltip: self.tooltip.destroy()
        student_info = self.seat_positions.get(rect_id)
        if not student_info: return
        
        # <<--- MỚI: HIỂN THỊ RÀNG BUỘC TRONG TOOLTIP --- >>
        sit_next_text = ", ".join(student_info.get('sit_next_to', [])) or "Không"
        dont_sit_text = ", ".join(student_info.get('dont_sit_next_to', [])) or "Không"
        
        text = (f"Tên: {student_info['Học sinh']}\n"
                f"Giới tính: {student_info.get('Giới tính', 'N/A')}\n"
                f"Điểm TB: {student_info.get('DiemTB', 'N/A')}\n"
                f"Ghi chú: {student_info.get('GhiChu', 'Không')}\n"
                f"--- Ràng buộc ---\n"
                f"Ngồi cạnh: {sit_next_text}\n"
                f"Không ngồi cạnh: {dont_sit_text}")
                
        self.tooltip = ttk.Toplevel(self.root); self.tooltip.wm_overrideredirect(True); self.tooltip.wm_geometry(f"+{event.x_root + 15}+{event.y_root + 10}")
        ttk.Label(self.tooltip, text=text, justify=LEFT, padding=5, background="#FFFFE0", relief="solid", borderwidth=1).pack()

    def hide_tooltip(self, event):
        if self.tooltip: self.tooltip.destroy(); self.tooltip = None

    def choose_color(self, team_index, preview_label):
        color_code = colorchooser.askcolor(title=f"Chọn màu cho tổ {team_index + 1}")
        if color_code[1]:
            self.colors[team_index] = color_code[1]
            preview_label.config(bg=self.colors[team_index])
            if self.students: self.arrange_seats(); self.update_status(f"Đã đổi màu cho tổ {team_index + 1}.")
    
    def delete_history_entry(self, arrangement_id, window):
        if messagebox.askyesno("Xác nhận", f"Bạn có chắc muốn xóa vĩnh viễn mục lịch sử ID: {arrangement_id}?"):
            self.cursor.execute("DELETE FROM arrangements_history WHERE id = ?", (arrangement_id,)); self.conn.commit()
            self.update_status(f"Đã xóa mục lịch sử ID: {arrangement_id}.")
            window.destroy(); self.view_history()

    def clear_history(self, window=None):
        if not self.current_class_id: return
        if messagebox.askyesno("Xác nhận", f"Bạn có chắc chắn muốn xóa TOÀN BỘ lịch sử của lớp '{self.current_class_name}' không?"):
            self.cursor.execute("DELETE FROM arrangements_history WHERE class_id=?", (self.current_class_id,)); self.conn.commit()
            self.update_status(f"Đã xóa toàn bộ lịch sử của lớp {self.current_class_name}.")
            if window: window.destroy(); self.view_history()


if __name__ == "__main__":
    root = ttk.Window(themename="litera")
    app = SeatArrangementApp(root)
    root.mainloop()