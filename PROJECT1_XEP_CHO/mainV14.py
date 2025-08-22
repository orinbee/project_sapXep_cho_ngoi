# --- START OF FILE mainv11.py ---

import tkinter as tk
from tkinter import filedialog, messagebox, colorchooser, scrolledtext, TclError, font as tkfont, simpledialog
import ttkbootstrap as ttk
from ttkbootstrap.constants import *
import pandas as pd
from openpyxl import Workbook
import random
import sqlite3
import json

from PIL import ImageGrab
import collections
import networkx as nx
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import math
from datetime import datetime, timedelta

# Yêu cầu cài đặt thư viện: pip install fpdf2 matplotlib
from fpdf import FPDF
from matplotlib.font_manager import findfont, FontProperties

# --- LỚP XỬ LÝ TẠO FILE PDF ---
# ... (Toàn bộ lớp PDFGenerator giữ nguyên, không thay đổi)
class PDFGenerator(FPDF):
    def __init__(self, class_name, teacher_name, school_year, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.class_name = class_name
        self.teacher_name = teacher_name
        self.school_year = school_year
        
        self.font_name = "SystemFont"
        try:
            regular_path = findfont(FontProperties(family="Times New Roman", style="normal", weight="normal"))
            self.add_font(self.font_name, "", regular_path, uni=True)
            
            bold_path = findfont(FontProperties(family="Times New Roman", style="normal", weight="bold"))
            self.add_font(self.font_name, "B", bold_path, uni=True)
            
            italic_path = findfont(FontProperties(family="Times New Roman", style="italic", weight="normal"))
            self.add_font(self.font_name, "I", italic_path, uni=True)
            
            print(f"Sử dụng font hệ thống 'Times New Roman' thành công.")
            
        except Exception:
            try:
                self.add_font("DejaVu", "", "DejaVuSans.ttf", uni=True)
                self.font_name = "DejaVu"
                print("Sử dụng font cục bộ: DejaVuSans.ttf")
            except RuntimeError as e:
                raise RuntimeError("Không tìm thấy font 'Times New Roman' trong hệ thống và cũng không có 'DejaVuSans.ttf' trong thư mục ứng dụng.") from e
        
        self.set_font(self.font_name, "", 14)

    def header(self):
        self.set_font(self.font_name, "B", 20)
        title = f"SƠ ĐỒ CHỖ NGỒI LỚP {self.class_name.upper()}"
        self.cell(0, 10, title, 0, 1, 'C')
        
        self.set_font(self.font_name, "", 11)
        info_line = f"Năm học: {self.school_year}  |  GVCN: {self.teacher_name}"
        self.cell(0, 8, info_line, 0, 1, 'C')
        self.ln(8)

    def footer(self):
        self.set_y(-15)
        self.set_font(self.font_name, "I", 8)
        export_date = datetime.now().strftime("%d/%m/%Y %H:%M:%S")
        self.cell(0, 10, f'Trang {self.page_no()}/{{nb}}  |  Xuất ngày: {export_date}', 0, 0, 'C')

    def draw_seating_chart(self, students, num_teams, num_tables_per_team, colors_hex):
        margin = 10
        drawable_width = self.w - 2 * margin
        drawable_height = self.h - 45

        team_gap = 10
        total_team_width = drawable_width - (num_teams - 1) * team_gap
        team_width = total_team_width / num_teams
        seat_width = (team_width - 5) / 2

        max_seat_height = drawable_height / num_tables_per_team - 5
        seat_height = min(25, max_seat_height) 
        
        def hex_to_rgb(hex_color):
            hex_color = hex_color.lstrip('#')
            return tuple(int(hex_color[i:i+2], 16) for i in (0, 2, 4))

        for i, student_name in enumerate(students):
            team_idx = i // (num_tables_per_team * 2)
            table_in_team_idx = (i % (num_tables_per_team * 2)) // 2
            seat_idx = i % 2

            x = margin + team_idx * (team_width + team_gap) + seat_idx * (seat_width + 5)
            y = 40 + table_in_team_idx * (seat_height + 5)

            color_rgb = hex_to_rgb(colors_hex[team_idx % len(colors_hex)])
            self.set_fill_color(*color_rgb)
            self.rect(x, y, seat_width, seat_height, 'DF')

            self.set_xy(x, y + seat_height/2 - 4)
            self.set_font(self.font_name, "", 9)
            self.set_text_color(0, 0, 0)
            self.multi_cell(seat_width, 4, student_name, 0, 'C')


# <<--- MỚI: LỚP XỬ LÝ SẮP XẾP THÔNG MINH --- >>
# THAY THẾ TOÀN BỘ LỚP AdvancedSorter CŨ BẰNG LỚP NÀY

class AdvancedSorter:
    def __init__(self, students_data):
        self.students_data = students_data
        self.student_map = {s['Học sinh']: s for s in students_data}

    def _get_partner_name(self, arrangement, index):
        partner_idx = index + 1 if index % 2 == 0 else index - 1
        if 0 <= partner_idx < len(arrangement):
            return arrangement[partner_idx]
        return None

    def _calculate_score(self, arrangement):
        score = 0
        for i, student_name in enumerate(arrangement):
            student_data = self.student_map.get(student_name)
            if not student_data: continue

            partner_name = self._get_partner_name(arrangement, i)
            if not partner_name: continue

            # Phạt nặng nếu vi phạm "Không ngồi cạnh"
            if partner_name in student_data.get('dont_sit_next_to', []):
                score -= 100
            
            # Thưởng điểm nếu thỏa mãn "Muốn ngồi cạnh"
            if partner_name in student_data.get('sit_next_to', []):
                score += 50
        
        # Thêm các tiêu chí khác ở đây nếu muốn, ví dụ:
        # - Phạt nếu 2 bạn nam ngồi cạnh nhau (để ưu tiên nam nữ)
        # - Thưởng nếu bạn học giỏi ngồi cạnh bạn học yếu
        return score

    def arrange(self, initial_arrangement):
        best_arrangement = list(initial_arrangement)
        best_score = self._calculate_score(best_arrangement)

        # Thử tối ưu trong một số lần lặp nhất định
        # Tăng số lần lặp nếu có nhiều ràng buộc phức tạp
        num_iterations = 200 * len(best_arrangement) 

        for _ in range(num_iterations):
            current_arrangement = list(best_arrangement)
            
            # Chọn ngẫu nhiên 2 học sinh để thử hoán đổi
            idx1, idx2 = random.sample(range(len(current_arrangement)), 2)
            
            current_arrangement[idx1], current_arrangement[idx2] = current_arrangement[idx2], current_arrangement[idx1]
            
            new_score = self._calculate_score(current_arrangement)
            
            # Nếu sơ đồ mới tốt hơn, giữ lại nó
            if new_score > best_score:
                best_score = new_score
                best_arrangement = current_arrangement

        print(f"Tối ưu hóa hoàn tất. Điểm số cuối cùng: {best_score}")
        return best_arrangement
# --- LỚP ỨNG DỤNG CHÍNH ---
class SeatArrangementApp:
    # --- CÁC HẰNG SỐ CHO GIAO DIỆN ---
    RECT_WIDTH = 180
    RECT_HEIGHT = 75
    X_GAP = 100
    Y_GAP = 40
    TEAM_X_GAP = 70
    
    TEACHER_DESK_WIDTH = 250
    TEACHER_DESK_HEIGHT = 70
    
    def __init__(self, root):
        self.root = root
        self.root.title("Chương trình quản lý sắp xếp chỗ ngồi dành cho học sinh ")
        
        # --- MỚI: Tải Cấu hình ---
        self.CONFIG_FILE = "config.json"
        self.settings = {
            "geometry": "1366x1080",
            "theme": "litera"
        }
        self._load_config()
        self.root.geometry(self.settings.get("geometry", "1366x1080"))
        # --- KẾT THÚC PHẦN MỚI ---
        
        self.students_data = []
        self.students = []
        self.undo_stack = []
        self.students = []
        self.undo_stack = []
        self.redo_stack = []

        self.colors = ["#FFA07A", "#7FFFD4", "#87CEFA", "#FFD700", "#98FB98", "#F08080", "#E0FFFF"]
        self.seat_positions = {}
        self.text_positions = {}
        self.dragged_item = None
        self.tooltip = None
        self.num_teams = 4
        self.num_tables = 5
        
        self.current_class_id = None
        self.current_class_name = None
        self.is_dirty = False
        self.search_var = tk.StringVar()

        self.init_db()
        self.load_ui()
        self.theme_var.set(self.settings.get('theme') == 'dark')
        self.toggle_theme() # Gọi để áp dụng theme đã tải
        self.root.protocol("WM_DELETE_WINDOW", self.on_closing)

        self.root.protocol("WM_DELETE_WINDOW", self.on_closing)
        self.update_status("Chào mừng! Chọn lớp để bắt đầu. Mẹo: Nháy đúp chuột vào học sinh để xem hồ sơ, chuột phải để đặt ràng buộc.")
    def _load_config(self):
        """Tải cấu hình từ file JSON. Nếu thất bại, sử dụng cài đặt mặc định."""
        try:
            with open(self.CONFIG_FILE, 'r', encoding='utf-8') as f:
                loaded_settings = json.load(f)
                # Dùng update để nếu có thêm key mới trong tương lai,
                # các file config cũ không bị lỗi.
                self.settings.update(loaded_settings)
        except (FileNotFoundError, json.JSONDecodeError):
            # Nếu file không tồn tại hoặc bị lỗi, cứ dùng self.settings mặc định
            print("Config file not found or corrupted. Using default settings.")
            pass

    def _save_config(self):
        """Lưu cấu hình hiện tại vào file JSON."""
        try:
            with open(self.CONFIG_FILE, 'w', encoding='utf-8') as f:
                json.dump(self.settings, f, indent=4)
        except IOError as e:
            print(f"Could not save config file: {e}")
    def init_db(self):
        self.conn = sqlite3.connect("seat_arrangements_multi_class.db")
        self.cursor = self.conn.cursor()
        
        self.cursor.execute('''
            CREATE TABLE IF NOT EXISTS classes (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL UNIQUE,
                students_data_json TEXT,
                current_arrangement_json TEXT,
                num_teams INTEGER,
                num_tables INTEGER,
                last_modified TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')

        self.cursor.execute('''
            CREATE TABLE IF NOT EXISTS arrangements_history (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                class_id INTEGER NOT NULL,
                arrangement TEXT,
                timestamp TEXT,
                FOREIGN KEY (class_id) REFERENCES classes (id) ON DELETE CASCADE
            )
        ''')
        
        # <<--- CẬP NHẬT: TỰ ĐỘNG NÂNG CẤP DATABASE ĐỂ HỖ TRỢ AI REPORTING --- >>
        try:
            self.cursor.execute("ALTER TABLE arrangements_history ADD COLUMN students_data_snapshot_json TEXT")
            print("Database upgraded successfully for AI Reporting.")
        except sqlite3.OperationalError:
            pass # Cột đã tồn tại, bỏ qua
            
        self.conn.commit()

    def on_closing(self):
        if self.is_dirty:
            if not messagebox.askyesno("Thoát", "Bạn có những thay đổi chưa được lưu. Bạn có chắc chắn muốn thoát?"):
                return
                
        # --- MỚI: Lưu Cấu hình ---
        self.settings['geometry'] = self.root.winfo_geometry()
        self.settings['theme'] = 'dark' if self.theme_var.get() else 'litera'
        self._save_config()
        # --- KẾT THÚC PHẦN MỚI ---
        
        self.conn.close()
        self.root.destroy()

    def open_roster_manager(self):
        if not self.current_class_id:
            messagebox.showwarning("Chưa chọn lớp", "Vui lòng chọn một lớp để quản lý sĩ số.")
            return
        if not self.students_data:
            if not messagebox.askyesno("Chưa có dữ liệu", "Lớp này chưa có dữ liệu học sinh. Bạn có muốn tạo mới?"):
                return

        win = ttk.Toplevel(self.root); win.title(f"Quản lý Sĩ số - Lớp {self.current_class_name}"); win.geometry("900x600"); win.transient(self.root); win.grab_set()
        tree_frame = ttk.Frame(win, padding=10); tree_frame.pack(fill=BOTH, expand=True)
        columns = ("name", "gender", "height", "score", "notes"); tree = ttk.Treeview(tree_frame, columns=columns, show="headings")
        tree.heading("name", text="Họ và Tên"); tree.heading("gender", text="Giới tính"); tree.heading("height", text="Chiều cao (cm)"); tree.heading("score", text="Điểm TB"); tree.heading("notes", text="Ghi Chú")
        tree.column("name", width=250); tree.column("gender", width=80, anchor=CENTER); tree.column("height", width=120, anchor=E); tree.column("score", width=100, anchor=E); tree.column("notes", width=250)
        scrollbar = ttk.Scrollbar(tree_frame, orient=VERTICAL, command=tree.yview, bootstyle='round'); tree.configure(yscrollcommand=scrollbar.set); tree.pack(side=LEFT, fill=BOTH, expand=True); scrollbar.pack(side=RIGHT, fill=Y)
        for student in self.students_data:
            values = (student.get('Học sinh', ''), student.get('Giới tính', 'Nữ'), student.get('Chiều cao', 160), student.get('DiemTB', 0.0), student.get('GhiChu', ''))
            tree.insert("", tk.END, values=values)
        btn_frame = ttk.Frame(win, padding=(10, 0, 10, 10)); btn_frame.pack(fill=X)
        ttk.Button(btn_frame, text="➕ Thêm Học sinh", bootstyle="success-outline", command=lambda: self._add_student_to_roster(tree)).pack(side=LEFT, padx=5)
        ttk.Button(btn_frame, text="❌ Xóa Học sinh đã chọn", bootstyle="danger-outline", command=lambda: self._delete_student_from_roster(tree)).pack(side=LEFT, padx=5)
        ttk.Button(btn_frame, text="💾 Lưu và Đóng", bootstyle="primary", command=lambda: self._save_roster_changes(tree, win)).pack(side=RIGHT, padx=5)
        def on_double_click(event):
            if tree.identify("region", event.x, event.y) != "cell": return
            item_id = tree.identify_row(event.y); column_id = tree.identify_column(event.x)
            x, y, width, height = tree.bbox(item_id, column_id)
            value = tree.item(item_id, "values")[int(column_id.replace('#','')) - 1]
            entry = ttk.Entry(tree_frame); entry.place(x=x, y=y, width=width, height=height); entry.insert(0, value); entry.focus_set()
            def on_save_edit(e): tree.set(item_id, column_id, entry.get()); entry.destroy()
            entry.bind("<Return>", on_save_edit); entry.bind("<FocusOut>", on_save_edit)
        tree.bind("<Double-1>", on_double_click)

    def _add_student_to_roster(self, tree):
        new_item = tree.insert("", tk.END, values=("Học sinh Mới", "Nữ", 160, 5.0, ""))
        tree.selection_set(new_item); tree.see(new_item)

    def _delete_student_from_roster(self, tree):
        selected_items = tree.selection()
        if not selected_items: messagebox.showwarning("Chưa chọn", "Vui lòng chọn ít nhất một học sinh để xóa.", parent=tree.winfo_toplevel()); return
        if messagebox.askyesno("Xác nhận Xóa", f"Bạn có chắc chắn muốn xóa {len(selected_items)} học sinh đã chọn không?"):
            for item in selected_items: tree.delete(item)

    def _save_roster_changes(self, tree, window):
        new_students_data = []; all_student_names = set()
        try:
            for item_id in tree.get_children():
                values = tree.item(item_id, "values"); name = str(values[0]).strip()
                if not name: raise ValueError("Tên học sinh không được để trống.")
                if name in all_student_names: raise ValueError(f"Tên học sinh '{name}' bị trùng lặp.")
                all_student_names.add(name); gender = str(values[1])
                if gender not in ['Nam', 'Nữ']: raise ValueError(f"Giới tính của '{name}' phải là 'Nam' hoặc 'Nữ'.")
                height = float(values[2]); score = float(values[3]); notes = str(values[4])
                old_student_info = next((s for s in self.students_data if s['Học sinh'] == name), {})
                # Cập nhật thông tin mới nhưng giữ lại các ràng buộc và dữ liệu AI cũ
                updated_info = {
                    'Học sinh': name, 'Giới tính': gender, 'Chiều cao': height, 'DiemTB': score, 'GhiChu': notes,
                    'sit_next_to': old_student_info.get('sit_next_to', []),
                    'dont_sit_next_to': old_student_info.get('dont_sit_next_to', []),
                    'positive_marks': old_student_info.get('positive_marks', 0),
                    'negative_marks': old_student_info.get('negative_marks', 0),
                    'private_notes': old_student_info.get('private_notes', '')
                }
                new_students_data.append(updated_info)
            self.students_data = new_students_data; self.students = [s['Học sinh'] for s in self.students_data]
            self._set_dirty(); self._sync_data_and_ui()
            self.update_status(f"Đã cập nhật sĩ số lớp. Hiện có {len(self.students)} học sinh."); window.destroy()
        except ValueError as e: messagebox.showerror("Lỗi Dữ liệu", str(e), parent=window)

    def update_status(self, message):
        self.status_bar.config(text=message)
        
    def _set_dirty(self, dirty_status=True):
        if not self.current_class_id: return
        self.is_dirty = dirty_status; current_text = self.class_selector.get()
        if dirty_status and not current_text.endswith('*'): self.class_selector.set(current_text + ' *')
        elif not dirty_status and current_text.endswith('*'): self.class_selector.set(current_text[:-2])
    
    # Phần tạo hướng dẫn
    def show_help_window(self):
        win = ttk.Toplevel(self.root)
        win.title("Hướng Dẫn Sử Dụng Phần Mềm")
        win.geometry("1200x1000")
        win.transient(self.root)
        win.grab_set()

        container = ttk.Frame(win, padding=15)
        container.pack(fill=BOTH, expand=True)

        help_text_widget = scrolledtext.ScrolledText(container, wrap=tk.WORD, font=("Arial", 11))
        help_text_widget.pack(fill=BOTH, expand=True)

        # --- NỘI DUNG HƯỚNG DẪN ---
        guide_content = """
        CHÀO MỪNG BẠN ĐẾN VỚI PHẦN MỀM QUẢN LÝ SẮP XẾP CHỖ NGỒI
        ------------------------------------------------------------------------------------

        Đây là hướng dẫn nhanh giúp bạn làm chủ các tính năng của phần mềm.

        QUY TRÌNH LÀM VIỆC CƠ BẢN (DÀNH CHO NGƯỜI MỚI):
        1.  **Thêm Lớp Mới:** Nhấn nút "Thêm Lớp Mới" trong khu vực "Quản Lý Lớp Học".
        2.  **Tải Dữ Liệu:** Chuyển qua tab "Quản lý & Dữ liệu", nhấn "Tải File Excel" để nhập danh sách học sinh.
        3.  **Sắp Xếp:** Chuyển qua tab "Sắp xếp & Tùy chỉnh", nhấn "Tùy chọn Sắp xếp" và chọn một phương pháp (khuyên dùng "Thông minh").
        4.  **Lưu Trạng Thái:** Nhấn nút "💾 LƯU TRẠNG THÁI LỚP" màu xanh dương để lưu lại mọi thay đổi (sơ đồ, sĩ số, ràng buộc).
        5.  **Xuất File:** Chuyển qua tab "Báo cáo & AI" để xuất sơ đồ ra file PDF hoặc Excel.


        KHÁM PHÁ CÁC TÍNH NĂNG CHÍNH:
        --------------------------------------------------

        **I. CÁC THAO TÁC TRÊN SƠ ĐỒ LỚP HỌC (KHU VỰC BÊN PHẢI):**

        *   **Kéo & Thả:** Nhấn giữ chuột trái vào một học sinh và kéo đến vị trí một học sinh khác để hoán đổi chỗ ngồi.
        *   **Xem Thông Tin Nhanh (Tooltip):** Di chuột lên một học sinh để xem thông tin cơ bản và các ràng buộc.
        *   **[QUAN TRỌNG] XEM HỒ SƠ HỌC SINH:** **Nháy đúp chuột** vào một học sinh để mở cửa sổ "Hồ sơ Học sinh". Tại đây bạn có thể:
            -   Tích điểm thi đua (điểm cộng/trừ).
            -   Ghi chú riêng tư về học sinh đó.
            -   Xem lịch sử các bạn ngồi cùng bàn.
        *   **[QUAN TRỌNG] ĐẶT RÀNG BUỘC CHỖ NGỒI:** **Nháy chuột phải** vào một học sinh để mở cửa sổ "Ràng buộc". Tại đây bạn có thể thiết lập:
            -   Học sinh này MUỐN ngồi cạnh ai.
            -   Học sinh này KHÔNG MUỐN ngồi cạnh ai.
            (Lưu ý: Sau khi đặt ràng buộc, hãy chạy lại "Sắp xếp thông minh" để áp dụng).


        **II. BẢNG ĐIỀU KHIỂN (KHU VỰC BÊN TRÁI):**

        *   **Tab "Quản lý & Dữ liệu":**
            -   **Quản lý Sĩ số:** Cho phép bạn thêm, xóa, sửa thông tin học sinh trực tiếp trên phần mềm mà không cần file Excel.
            -   **Cấu hình Tổ/Bàn:** Thiết lập số tổ và số bàn mỗi tổ cho lớp học. Nhấn "Áp dụng" để vẽ lại sơ đồ.

        *   **Tab "Sắp xếp & Tùy chỉnh":**
            -   **Tùy chọn Sắp xếp:** Cung cấp nhiều thuật toán sắp xếp khác nhau. "Thông minh" là mạnh nhất, sẽ cố gắng thỏa mãn các ràng buộc bạn đã đặt.
            -   **Hoàn tác / Làm lại:** Quay lại hoặc tiến tới các bước thay đổi sơ đồ (kéo thả, sắp xếp...).
            -   **Đổi màu / Gọi tên:** Các công cụ hỗ trợ trực quan.

        *   **Tab "Báo cáo & AI":**
            -   **Xuất Excel & Lưu Lịch Sử:** Vừa xuất ra file Excel, vừa tạo một bản ghi sơ đồ vào lịch sử của lớp.
            -   **Xem Lịch Sử:** Xem lại, khôi phục hoặc xóa các sơ đồ đã lưu trước đó.
            -   **Phân tích Tương tác:** Vẽ biểu đồ mạng lưới xã hội của lớp, tìm ra các "nhóm bạn thân" và những học sinh cần quan tâm.
            -   **AI Gợi Ý Chia Nhóm:** Công cụ thông minh giúp chia lớp thành các nhóm học tập theo nhiều tiêu chí (hòa hợp, cân bằng...).
            -   **Báo cáo & Phân tích AI:** Tính năng cao cấp nhất, phân tích dữ liệu lịch sử để đưa ra các cảnh báo và gợi ý về xu hướng học tập của học sinh. (Cần có dữ liệu lịch sử phong phú để hoạt động tốt).


        **LƯU Ý QUAN TRỌNG:**
        ------------------------------
        -   Nút **"LƯU TRẠNG THÁI LỚP"** sẽ lưu cấu trúc lớp học (sĩ số, ràng buộc, cấu hình bàn ghế) vào database để dùng cho lần sau.
        -   Nút **"Xuất Excel & Lưu Lịch Sử"** sẽ lưu một bản ghi sơ đồ tại thời điểm đó vào lịch sử. Hãy thực hiện việc này mỗi khi bạn có một sơ đồ ưng ý.

        Chúc bạn có những trải nghiệm hiệu quả với phần mềm!
        """
    
        help_text_widget.insert(tk.END, guide_content)
        help_text_widget.config(state=tk.DISABLED) # Chuyển sang chế độ chỉ đọc

        ttk.Button(container, text="Đóng", command=win.destroy, bootstyle="secondary").pack(pady=10)

    def load_ui(self):
        main_pane = ttk.PanedWindow(self.root, orient=HORIZONTAL)
        main_pane.pack(fill=BOTH, expand=True)
        
        control_panel = ttk.Frame(main_pane, padding=10, width=350)
        control_panel.pack_propagate(False)
        main_pane.add(control_panel, weight=1)
        
        canvas_container = ttk.Frame(main_pane, padding=(0, 10, 10, 0))
        main_pane.add(canvas_container, weight=4)

        # --- PHẦN QUẢN LÝ LỚP HỌC (Giữ nguyên ở trên cùng) ---
        class_lf = ttk.LabelFrame(control_panel, text=" Quản Lý Lớp Học ", padding=10, bootstyle=PRIMARY)
        class_lf.pack(fill=X, pady=(0, 15))
        self.class_selector = ttk.Combobox(class_lf, state="readonly", values=[])
        self.class_selector.pack(fill=X, pady=(0, 5))
        self.class_selector.bind("<<ComboboxSelected>>", self._on_class_selected)
        class_btn_frame = ttk.Frame(class_lf)
        class_btn_frame.pack(fill=X)
        ttk.Button(class_btn_frame, text="Thêm Lớp Mới", command=self._add_new_class, bootstyle="success-outline").pack(side=LEFT, expand=True, fill=X, padx=(0,2))
        ttk.Button(class_btn_frame, text="Xóa Lớp Này", command=self._delete_class, bootstyle="danger-outline").pack(side=LEFT, expand=True, fill=X, padx=(2,0))
        ttk.Button(class_lf, text="💾 LƯU TRẠNG THÁI LỚP", command=self._save_class_state, bootstyle="primary").pack(fill=X, pady=5)
         # --- MỚI: Thêm Chức năng Tìm kiếm ---
        search_lf = ttk.LabelFrame(control_panel, text=" Tìm kiếm Nhanh ", padding=10)
        search_lf.pack(fill=X, pady=(0, 10))
        
        search_entry = ttk.Entry(search_lf, textvariable=self.search_var)
        search_entry.pack(fill=X, expand=True)
        search_entry.bind("<Return>", self._on_search_student) # Gán sự kiện nhấn Enter
        search_entry.bind("<FocusIn>", lambda e: search_entry.config(bootstyle="primary"))
        search_entry.bind("<FocusOut>", lambda e: search_entry.config(bootstyle="default"))
        
        # Gợi ý cho người dùng
        def on_focusout(event):
            if not self.search_var.get():
                self.search_var.set("Nhập tên học sinh rồi nhấn Enter...")
                search_entry.config(bootstyle="secondary")
        def on_focusin(event):
            if self.search_var.get() == "Nhập tên học sinh rồi nhấn Enter...":
                self.search_var.set("")
                search_entry.config(bootstyle="default")
                
        search_entry.bind("<FocusIn>", on_focusin)
        search_entry.bind("<FocusOut>", on_focusout)
        on_focusout(None) # Gọi lần đầu để hiển thị gợi ý
        # --- TẠO NOTEBOOK (GIAO DIỆN TAB) ---
        notebook = ttk.Notebook(control_panel, bootstyle="primary")
        notebook.pack(fill=BOTH, expand=True)

        # --- Tạo các Frame cho từng Tab ---
        tab1_frame = ttk.Frame(notebook, padding=10)
        tab2_frame = ttk.Frame(notebook, padding=10)
        tab3_frame = ttk.Frame(notebook, padding=10)

        notebook.add(tab1_frame, text=" Quản lý & Dữ liệu ")
        notebook.add(tab2_frame, text=" Sắp xếp & Tùy chỉnh ")
        notebook.add(tab3_frame, text=" Báo cáo & AI ")
        
        # --- TAB 1: QUẢN LÝ & DỮ LIỆU ---
        lf1 = ttk.LabelFrame(tab1_frame, text=" 1. Dữ Liệu & Cấu Hình ", padding=10)
        lf1.pack(fill=BOTH, expand=True)
        ttk.Button(lf1, text="📂 Tải File Excel cho Lớp Này", command=self.load_students).pack(fill=X)
        ttk.Button(lf1, text="👨‍🎓 Quản lý Sĩ số Lớp", command=self.open_roster_manager, bootstyle="info").pack(fill=X, pady=(5, 0))
        config_frame = ttk.Frame(lf1)
        config_frame.pack(fill=X, pady=10)
        ttk.Label(config_frame, text="Số tổ:").grid(row=0, column=0, padx=5, pady=5, sticky=W)
        self.team_spinbox = ttk.Spinbox(config_frame, from_=1, to=20, width=5)
        self.team_spinbox.grid(row=0, column=1, padx=5, pady=5, sticky=W)
        self.team_spinbox.set(self.num_teams)
        ttk.Label(config_frame, text="Số bàn/tổ:").grid(row=0, column=2, padx=5, pady=5, sticky=W)
        self.table_spinbox = ttk.Spinbox(config_frame, from_=1, to=30, width=5)
        self.table_spinbox.grid(row=0, column=3, padx=5, pady=5, sticky=W)
        self.table_spinbox.set(self.num_tables)
        ttk.Button(lf1, text="Áp dụng cấu hình", command=self.apply_team_table_config, bootstyle="secondary-outline").pack(fill=X)

        # --- TAB 2: SẮP XẾP & TÙY CHỈNH ---
        lf3 = ttk.LabelFrame(tab2_frame, text=" 2. Tác vụ Sắp xếp ", padding=10)
        lf3.pack(fill=X, pady=(0, 10))
        ttk.Button(lf3, text="🎲 Tùy chọn Sắp xếp", command=self.show_sort_options, bootstyle="success").pack(fill=X, pady=(0,5))
        undo_redo_frame = ttk.Frame(lf3)
        undo_redo_frame.pack(fill=X, pady=(0,5))
        self.undo_btn = ttk.Button(undo_redo_frame, text="↩️ Hoàn tác", command=self.undo, bootstyle="secondary-outline", state="disabled")
        self.undo_btn.pack(side=LEFT, expand=True, fill=X, padx=(0,2))
        self.redo_btn = ttk.Button(undo_redo_frame, text="↪️ Làm lại", command=self.redo, bootstyle="secondary-outline", state="disabled")
        self.redo_btn.pack(side=LEFT, expand=True, fill=X, padx=(2,0))
        
        lf_misc = ttk.LabelFrame(tab2_frame, text=" 3. Công cụ Nhanh ", padding=10)
        lf_misc.pack(fill=X, pady=10)
        misc_frame = ttk.Frame(lf_misc)
        misc_frame.pack(fill=X)
        ttk.Button(misc_frame, text="🎨 Đổi Màu", command=self.change_team_colors, bootstyle="secondary-outline").pack(side=LEFT, expand=True, fill=X, padx=(0,2))
        ttk.Button(misc_frame, text="✨ Gọi Tên", command=self.pick_random_student, bootstyle="success-outline").pack(side=LEFT, expand=True, fill=X, padx=(2,0))

        # --- TAB 3: BÁO CÁO & AI ---
        lf4 = ttk.LabelFrame(tab3_frame, text=" 4. Xuất & Báo Cáo ", padding=10)
        lf4.pack(fill=X, pady=(0, 10))
        ttk.Button(lf4, text="Xuất Excel & Lưu Lịch Sử", command=self.save_results, bootstyle="info-outline").pack(fill=X, pady=(0, 5))
        ttk.Button(lf4, text="🖨️ Xuất Sơ Đồ ra PDF", command=self.export_to_pdf, bootstyle="info").pack(fill=X, pady=(5, 5))
        ttk.Button(lf4, text="Chụp Ảnh Sơ Đồ", command=self.save_as_image, bootstyle="info-outline").pack(fill=X, pady=(5,0))
        ttk.Button(lf4, text="📜 Xem Lịch Sử Sắp Xếp", command=self.view_history, bootstyle="secondary-outline").pack(fill=X, pady=(5,0))
        
        lf5 = ttk.LabelFrame(tab3_frame, text=" 5. Phân tích Thông minh ", padding=10, bootstyle="danger")
        lf5.pack(fill=X, pady=10)
        ttk.Button(lf5, text="📊 Phân tích Tương tác Lớp học", command=self.analyze_social_network, bootstyle="primary-outline").pack(fill=X, pady=(5, 0))
        ttk.Button(lf5, text="🤖 AI Gợi Ý Chia Nhóm Học Tập", command=self.open_group_creation_tool, bootstyle="primary-outline").pack(fill=X, pady=(5, 0))
        ttk.Button(lf5, text="📈 Báo Cáo & Phân Tích AI", command=self.open_ai_report_window, bootstyle="danger").pack(fill=X, pady=(5, 0))
        
        # --- PHẦN CANVAS VÀ STATUS BAR (Giữ nguyên) ---
        theme_frame = ttk.Frame(control_panel)
        theme_frame.pack(side=BOTTOM, fill=X, pady=(10,0))
        ttk.Label(theme_frame, text="Chế độ Sáng / Tối:").pack(side=LEFT)
        self.theme_var = tk.BooleanVar(value=False)
        ttk.Checkbutton(theme_frame, bootstyle="switch", variable=self.theme_var, command=self.toggle_theme).pack(side=LEFT, padx=10)
        title_bar_frame = ttk.Frame(canvas_container)
        title_bar_frame.pack(fill=X, pady=(0, 5))
        ttk.Label(canvas_container, text="Sơ Đồ Lớp Học", font=("Arial", 16, "bold")).pack(pady=(0, 5))
        #nút hướng dẫn
        ttk.Button(title_bar_frame, text="❓ Hướng Dẫn", command=self.show_help_window, bootstyle="info-outline").pack(side=RIGHT)
        canvas_frame = ttk.Frame(canvas_container)

        canvas_frame.pack(fill=BOTH, expand=True)
        self.canvas = tk.Canvas(canvas_frame, bg='white', relief="solid", bd=1)
        h_scroll = ttk.Scrollbar(canvas_frame, orient=HORIZONTAL, command=self.canvas.xview, bootstyle="round")
        v_scroll = ttk.Scrollbar(canvas_frame, orient=VERTICAL, command=self.canvas.yview, bootstyle="round")
        self.canvas.config(xscrollcommand=h_scroll.set, yscrollcommand=v_scroll.set)
        h_scroll.pack(side=BOTTOM, fill=X)
        v_scroll.pack(side=RIGHT, fill=Y)
        self.canvas.pack(side=LEFT, fill=BOTH, expand=True)
        
        self.status_bar = ttk.Label(self.root, text="Sẵn sàng", padding=5, font=("Arial", 9), anchor=W, bootstyle=INVERSE)
        self.status_bar.pack(side=BOTTOM, fill=X)
        self._load_class_list_to_selector()
    def _on_search_student(self, event=None):
        query = self.search_var.get().strip().lower()
        if not query or query == "nhập tên học sinh rồi nhấn enter...":
            return

        found_rect_id = None
        found_student_name = ""

        # Tìm kiếm học sinh (ưu tiên khớp chính xác, sau đó đến chứa)
        for rect_id, student_info in self.seat_positions.items():
            student_name = student_info['Học sinh'].lower()
            if query == student_name: # Ưu tiên khớp hoàn toàn
                found_rect_id = rect_id
                found_student_name = student_info['Học sinh']
                break

        if not found_rect_id: # Nếu không khớp hoàn toàn, tìm kiếm chứa
            for rect_id, student_info in self.seat_positions.items():
                if query in student_info['Học sinh'].lower():
                    found_rect_id = rect_id
                    found_student_name = student_info['Học sinh']
                    break # Lấy kết quả đầu tiên

        if found_rect_id:
            self.update_status(f"Đã tìm thấy: {found_student_name}")
            
            # Highlight logic (tương tự pick_random_student)
            original_color = self.canvas.itemcget(found_rect_id, "fill")
            highlight_color = "#FFD700"  # Màu vàng nổi bật

            def revert_highlight():
                try:
                    self.canvas.itemconfig(found_rect_id, fill=original_color)
                except tk.TclError:
                    pass # Bỏ qua lỗi nếu widget đã bị hủy

            self.canvas.itemconfig(found_rect_id, fill=highlight_color)
            
            text_item = self.text_positions.get(found_rect_id)
            if text_item:
                self.canvas.tag_raise(found_rect_id)
                self.canvas.tag_raise(text_item)
            
            self.root.after(2500, revert_highlight) # Giữ highlight 2.5 giây
        else:
            self.update_status(f"Không tìm thấy học sinh nào có tên chứa '{self.search_var.get()}'")
    
        
    # <<--- MỚI: TOÀN BỘ HÀM CHO TÍNH NĂNG PHÂN TÍCH MẠNG LƯỚI XÃ HỘI --- >>
    def analyze_social_network(self):
        if not self.students_data or len(self.students_data) < 2:
            messagebox.showwarning("Chưa có dữ liệu", "Cần có ít nhất 2 học sinh và các ràng buộc 'Muốn ngồi cạnh' để phân tích.")
            return

        self.update_status("Đang phân tích mạng lưới tương tác xã hội...")
        
        G = nx.Graph()
        all_students_with_constraints = set()

        # 1. Xây dựng đồ thị từ dữ liệu `sit_next_to`
        for student_info in self.students_data:
            student_name = student_info['Học sinh']
            G.add_node(student_name)
            
            # Chỉ xét các cạnh "bạn bè" (muốn ngồi cạnh) để phân tích cộng đồng
            sit_next_to_list = student_info.get('sit_next_to', [])
            if sit_next_to_list:
                all_students_with_constraints.add(student_name)
            for friend_name in sit_next_to_list:
                # Đảm bảo bạn bè cũng có trong danh sách lớp
                if friend_name in self.students:
                    G.add_edge(student_name, friend_name)
                    all_students_with_constraints.add(friend_name)

        if not G.edges():
            messagebox.showinfo("Thông tin", "Chưa có ràng buộc 'Muốn ngồi cạnh' nào được thiết lập. Không thể phân tích mạng lưới.")
            self.update_status("Phân tích thất bại: không có dữ liệu ràng buộc.")
            return

        # 2. Phân tích đồ thị
        # Tìm các cụm/cộng đồng (nhóm bạn thân)
        try:
            communities = list(nx.algorithms.community.greedy_modularity_communities(G))
        except Exception:
            communities = [] # Xử lý trường hợp đồ thị quá đơn giản

        # Tìm các học sinh bị cô lập (không có kết nối `sit_next_to` nào)
        isolated_students = [s['Học sinh'] for s in self.students_data if s['Học sinh'] not in all_students_with_constraints]
        
        # Tìm học sinh có nhiều kết nối nhất (người có ảnh hưởng/kết nối)
        most_connected = sorted(G.degree, key=lambda x: x[1], reverse=True)
        
        # 3. Tạo báo cáo văn bản
        report_text = "--- BÁO CÁO PHÂN TÍCH TƯƠNG TÁC LỚP HỌC ---\n\n"
        if communities:
            report_text += f"🔎 Phát hiện được {len(communities)} nhóm/cộng đồng chính:\n"
            for i, group in enumerate(communities):
                report_text += f"  - Nhóm {i+1}: {', '.join(list(group))}\n"
        else:
            report_text += "🔎 Không phát hiện được nhóm/cộng đồng rõ rệt.\n"
        
        report_text += "\n"
        if isolated_students:
            report_text += f"💔 Các học sinh cần quan tâm (ít tương tác):\n"
            report_text += f"  - {', '.join(isolated_students)}\n"
        else:
            report_text += "👍 Không có học sinh nào bị cô lập hoàn toàn (dựa trên dữ liệu hiện có).\n"
            
        report_text += "\n"
        if most_connected:
            top_connectors = [name for name, degree in most_connected[:3] if degree > 0]
            if top_connectors:
                report_text += f"🔗 Các học sinh có nhiều kết nối nhất (cầu nối):\n"
                report_text += f"  - {', '.join(top_connectors)}\n"

        # 4. Trực quan hóa và hiển thị kết quả
        self._show_analysis_window(G, communities, isolated_students, report_text)
        self.update_status("Đã hoàn thành phân tích mạng lưới tương tác.")
    def open_group_creation_tool(self):
        if not self.students_data or len(self.students_data) < 2:
            messagebox.showwarning("Chưa có dữ liệu", "Cần có dữ liệu học sinh để có thể tạo nhóm.")
            return

        win = ttk.Toplevel(self.root)
        win.title("AI Gợi Ý Chia Nhóm Học Tập")
        win.geometry("800x800")
        win.transient(self.root); win.grab_set()

        container = ttk.Frame(win, padding=20); container.pack(fill=BOTH, expand=True)
        
        # --- Phần nhập liệu ---
        input_frame = ttk.Frame(container); input_frame.pack(fill=X, pady=10)
        ttk.Label(input_frame, text="Số nhóm cần tạo:", font=("Arial", 11)).pack(side=LEFT, padx=5)
        
        self.num_groups_var = tk.IntVar(value=max(1, len(self.students_data) // 5)) # Gợi ý số nhóm
        num_groups_spinbox = ttk.Spinbox(input_frame, from_=1, to=len(self.students_data), textvariable=self.num_groups_var, width=8)
        num_groups_spinbox.pack(side=LEFT, padx=5)

        # --- Phần chọn tiêu chí ---
        lf = ttk.LabelFrame(container, text=" Chọn tiêu chí chia nhóm ", padding=15, bootstyle=PRIMARY)
        lf.pack(fill=BOTH, expand=True, pady=10)

        self.grouping_criteria_var = tk.StringVar(value="balanced")
        style = ttk.Style(); style.configure('TRadiobutton', font=('Arial', 10), padding=(0,8))
        
        ttk.Radiobutton(lf, text="Hòa Hợp (Xếp bạn thân cùng nhóm)", variable=self.grouping_criteria_var, value="harmonious", style='TRadiobutton').pack(anchor=W)
        ttk.Radiobutton(lf, text="Cân Bằng (Đều về học lực, giới tính)", variable=self.grouping_criteria_var, value="balanced", style='TRadiobutton').pack(anchor=W)
        ttk.Radiobutton(lf, text="Ngẫu Nhiên Thông Minh (Tránh xung đột)", variable=self.grouping_criteria_var, value="smart_random", style='TRadiobutton').pack(anchor=W)

        # --- Nút thực thi ---
        btn_frame = ttk.Frame(container); btn_frame.pack(pady=20)
        ttk.Button(btn_frame, text="Tạo Nhóm", command=lambda: self._generate_groups(win), bootstyle=SUCCESS).pack(side=LEFT, padx=10)
        ttk.Button(btn_frame, text="Hủy", command=win.destroy, bootstyle="secondary-outline").pack(side=LEFT, padx=10)

    def _generate_groups(self, parent_window):
        num_groups = self.num_groups_var.get()
        criteria = self.grouping_criteria_var.get()

        if num_groups <= 0 or num_groups > len(self.students_data):
            messagebox.showerror("Lỗi", "Số nhóm không hợp lệ.", parent=parent_window)
            return
        
        self.update_status(f"AI đang chia nhóm theo tiêu chí '{criteria}'...")
        
        groups = []
        if criteria == "harmonious":
            groups = self._create_harmonious_groups(num_groups)
        elif criteria == "balanced":
            groups = self._create_balanced_groups(num_groups)
        elif criteria == "smart_random":
            groups = self._create_smart_random_groups(num_groups)

        self.update_status("Hoàn thành chia nhóm.")
        self._show_group_results(groups, criteria)
        parent_window.destroy()

    def _create_harmonious_groups(self, num_groups):
        G = nx.Graph()
        for student_info in self.students_data:
            G.add_node(student_info['Học sinh'])
            for friend_name in student_info.get('sit_next_to', []):
                if friend_name in self.students:
                    G.add_edge(student_info['Học sinh'], friend_name)
        
        try: communities = list(nx.algorithms.community.greedy_modularity_communities(G))
        except: communities = []
        
        groups = [[] for _ in range(num_groups)]
        placed_students = set()

        # Ưu tiên xếp các cộng đồng vào trước
        communities.sort(key=len, reverse=True) # Xếp cộng đồng lớn trước
        for i, community in enumerate(communities):
            target_group_index = i % num_groups
            groups[target_group_index].extend(list(community))
            for student in community: placed_students.add(student)

        # Xếp các học sinh còn lại
        remaining_students = [s['Học sinh'] for s in self.students_data if s['Học sinh'] not in placed_students]
        random.shuffle(remaining_students)
        for student in remaining_students:
            groups.sort(key=len) # Luôn thêm vào nhóm ít người nhất
            groups[0].append(student)

        return groups

    def _create_balanced_groups(self, num_groups):
        students_df = pd.DataFrame(self.students_data)
        students_df = students_df.sort_values(by='DiemTB', ascending=False).reset_index(drop=True)
        
        groups = [[] for _ in range(num_groups)]
        student_map = {s['Học sinh']: s for s in self.students_data}

        # Tạo ma trận xung đột để tra cứu nhanh
        conflict_map = collections.defaultdict(set)
        for s in self.students_data:
            for conflict_student in s.get('dont_sit_next_to', []):
                conflict_map[s['Học sinh']].add(conflict_student)

        # Phân phối "zig-zag" để cân bằng điểm số
        unplaced_students = []
        direction = 1
        group_idx = 0
        for _, student_row in students_df.iterrows():
            student_name = student_row['Học sinh']
            placed = False
            
            # Cố gắng đặt vào các nhóm
            initial_group_idx = group_idx
            for _ in range(num_groups):
                # Kiểm tra xung đột
                has_conflict = False
                for member in groups[group_idx]:
                    if member in conflict_map[student_name]:
                        has_conflict = True; break
                
                if not has_conflict:
                    groups[group_idx].append(student_name)
                    placed = True; break
                
                group_idx = (group_idx + direction) % num_groups # Thử nhóm tiếp theo
            
            if not placed: unplaced_students.append(student_name)

            # Đổi hướng đi của "zig-zag"
            if group_idx == num_groups - 1 and direction == 1: direction = -1
            elif group_idx == 0 and direction == -1: direction = 1
            else: group_idx += direction

        # Cố gắng đặt nốt những em chưa có chỗ
        for student in unplaced_students:
            groups.sort(key=len); groups[0].append(student)

        return groups

    def _create_smart_random_groups(self, num_groups):
        students_list = self.students.copy()
        random.shuffle(students_list)
        groups = [[] for _ in range(num_groups)]
        
        conflict_map = collections.defaultdict(set)
        for s in self.students_data:
            for conflict_student in s.get('dont_sit_next_to', []):
                conflict_map[s['Học sinh']].add(conflict_student)

        for student in students_list:
            placed = False
            # Sắp xếp các nhóm theo số lượng thành viên tăng dần để ưu tiên nhóm nhỏ
            sorted_groups = sorted(enumerate(groups), key=lambda x: len(x[1]))
            
            for group_idx, group in sorted_groups:
                has_conflict = any(member in conflict_map[student] for member in group)
                if not has_conflict:
                    groups[group_idx].append(student)
                    placed = True; break
            
            if not placed: # Nếu không thể tránh xung đột, đặt vào nhóm nhỏ nhất
                groups[sorted_groups[0][0]].append(student)
                
        return groups

    def _show_group_results(self, groups, criteria):
        win = ttk.Toplevel(self.root)
        win.title(f"Kết quả Chia Nhóm - Tiêu chí: {criteria.replace('_', ' ').title()}")
        win.geometry("1200x1000")

        container = ttk.Frame(win, padding=10); container.pack(fill=BOTH, expand=True)
        
        # Tạo chuỗi kết quả
        result_text = f"--- KẾT QUẢ CHIA {len(groups)} NHÓM THEO TIÊU CHÍ '{criteria.replace('_', ' ').upper()}' ---\n\n"
        student_map = {s['Học sinh']: s for s in self.students_data}
        
        for i, group in enumerate(groups):
            if not group: continue
            
            group_scores = [student_map.get(name, {}).get('DiemTB', 0) for name in group]
            avg_score = sum(group_scores) / len(group_scores) if group_scores else 0
            
            num_males = sum(1 for name in group if student_map.get(name, {}).get('Giới tính') == 'Nam')
            num_females = len(group) - num_males
            
            result_text += f"================ NHÓM {i+1} ================\n"
            result_text += f"Sĩ số: {len(group)} (Nam: {num_males}, Nữ: {num_females}) | Điểm TB: {avg_score:.2f}\n"
            result_text += "Thành viên: " + ", ".join(group) + "\n\n"
        
        # Hiển thị
        text_widget = scrolledtext.ScrolledText(container, wrap=tk.WORD, font=("Arial", 11))
        text_widget.pack(fill=BOTH, expand=True, pady=5)
        text_widget.insert(tk.END, result_text)
        text_widget.config(state=tk.DISABLED)

        # Nút xuất file
        def export_to_text():
            file_path = filedialog.asksaveasfilename(defaultextension=".txt", filetypes=[("Text Files", "*.txt")], title="Lưu kết quả chia nhóm")
            if file_path:
                with open(file_path, 'w', encoding='utf-8') as f:
                    f.write(text_widget.get('1.0', tk.END))
                messagebox.showinfo("Thành công", f"Đã lưu kết quả vào file:\n{file_path}", parent=win)

        ttk.Button(container, text="💾 Xuất ra file Text", command=export_to_text, bootstyle="info-outline").pack(pady=10)

    def _show_analysis_window(self, G, communities, isolated_students, report_text):
        win = ttk.Toplevel(self.root)
        win.title(f"Phân tích Tương tác Xã hội - Lớp {self.current_class_name}")
        win.geometry("2200x1080")
        win.transient(self.root); win.grab_set()

        pane = ttk.PanedWindow(win, orient=HORIZONTAL)
        pane.pack(fill=BOTH, expand=True)

        report_frame = ttk.Frame(pane, padding=10); pane.add(report_frame, weight=1)
        graph_frame = ttk.Frame(pane, padding=10); pane.add(graph_frame, weight=2)
        
        # Hiển thị báo cáo văn bản
        ttk.Label(report_frame, text="Kết quả Phân tích", font=("Arial", 14, "bold")).pack(pady=5)
        report_widget = scrolledtext.ScrolledText(report_frame, wrap=tk.WORD, font=("Arial", 11), relief="solid", bd=1)
        report_widget.pack(fill=BOTH, expand=True)
        report_widget.insert(tk.END, report_text)
        report_widget.config(state=tk.DISABLED)
        
        # Vẽ và nhúng biểu đồ
        fig = plt.Figure(figsize=(8, 8), dpi=100)
        ax = fig.add_subplot(111)
        
        # Sử dụng layout để các node không bị chồng chéo
        pos = nx.spring_layout(G, k=0.8, iterations=50, seed=42) 

        # Tạo màu cho các cộng đồng
        color_map = {}
        colors = plt.cm.get_cmap('viridis', len(communities))
        for i, group in enumerate(communities):
            for node in group:
                color_map[node] = colors(i)
        
        node_colors = [color_map.get(node, '#cccccc') for node in G.nodes()]

        nx.draw_networkx_edges(G, pos, ax=ax, alpha=0.6)
        nx.draw_networkx_nodes(G, pos, ax=ax, node_color=node_colors, node_size=2000)
        nx.draw_networkx_labels(G, pos, ax=ax, font_size=9, font_family='Arial', font_color='black')

        # Đánh dấu các học sinh bị cô lập
        all_nodes_in_graph = set(G.nodes())
        for student_name in isolated_students:
            if student_name not in all_nodes_in_graph:
                 G.add_node(student_name)
                 pos[student_name] = (random.uniform(-1, 1), random.uniform(-1, 1)) # Vị trí ngẫu nhiên
        
        isolated_nodes = [node for node in G.nodes() if node in isolated_students]
        if isolated_nodes:
            nx.draw_networkx_nodes(G, pos, nodelist=isolated_nodes, ax=ax, node_color='#FF6347', node_size=2000, edgecolors='red', linewidths=2)

        ax.set_title(f"Sơ đồ Tương tác Lớp {self.current_class_name}", fontsize=16)
        plt.tight_layout()

        canvas = FigureCanvasTkAgg(fig, master=graph_frame)
        canvas.draw()
        canvas.get_tk_widget().pack(side=tk.TOP, fill=tk.BOTH, expand=1)
    
    def show_student_profile(self, rect_id):
        student_info = self.seat_positions.get(rect_id)
        if not student_info: return

        win = ttk.Toplevel(self.root)
        win.title(f"Hồ sơ Học sinh - {student_info['Học sinh']}")
        win.geometry("700x1250")
        win.transient(self.root); win.grab_set()

        container = ttk.Frame(win, padding=15); container.pack(fill=BOTH, expand=True)
        
        # --- Frame thông tin cơ bản ---
        info_lf = ttk.LabelFrame(container, text=" Thông tin Cơ bản ", padding=10, bootstyle=PRIMARY)
        info_lf.pack(fill=X, pady=5)
        info_lf.columnconfigure(1, weight=1)

        ttk.Label(info_lf, text="Họ và tên:", font="-weight bold").grid(row=0, column=0, sticky=W, padx=5, pady=2)
        ttk.Label(info_lf, text=student_info['Học sinh']).grid(row=0, column=1, sticky=W, padx=5, pady=2)
        ttk.Label(info_lf, text="Giới tính:", font="-weight bold").grid(row=1, column=0, sticky=W, padx=5, pady=2)
        ttk.Label(info_lf, text=student_info.get('Giới tính', 'N/A')).grid(row=1, column=1, sticky=W, padx=5, pady=2)
        ttk.Label(info_lf, text="Điểm TB:", font="-weight bold").grid(row=2, column=0, sticky=W, padx=5, pady=2)
        ttk.Label(info_lf, text=student_info.get('DiemTB', 'N/A')).grid(row=2, column=1, sticky=W, padx=5, pady=2)
        ttk.Label(info_lf, text="Ghi chú chung:", font="-weight bold").grid(row=3, column=0, sticky=W, padx=5, pady=2)
        ttk.Label(info_lf, text=student_info.get('GhiChu', 'Không có')).grid(row=3, column=1, sticky=W, padx=5, pady=2)

        # --- Frame Điểm Thi Đua ---
        marks_lf = ttk.LabelFrame(container, text=" Tích Điểm Thi Đua ", padding=10, bootstyle=INFO)
        marks_lf.pack(fill=X, pady=10)
        
        positive_marks = tk.IntVar(value=student_info.get('positive_marks', 0))
        negative_marks = tk.IntVar(value=student_info.get('negative_marks', 0))

        ttk.Label(marks_lf, text="Điểm cộng (Phát biểu, làm tốt...):").grid(row=0, column=0, sticky=W, padx=5)
        ttk.Button(marks_lf, text="-", width=3, bootstyle="danger-outline", command=lambda: positive_marks.set(max(0, positive_marks.get() - 1))).grid(row=0, column=1, padx=(10, 2))
        ttk.Label(marks_lf, textvariable=positive_marks, font="-weight bold", width=4, anchor=CENTER).grid(row=0, column=2)
        ttk.Button(marks_lf, text="+", width=3, bootstyle="success-outline", command=lambda: positive_marks.set(positive_marks.get() + 1)).grid(row=0, column=3, padx=2)
        
        ttk.Label(marks_lf, text="Điểm trừ (Mất trật tự...):").grid(row=1, column=0, sticky=W, padx=5, pady=5)
        ttk.Button(marks_lf, text="-", width=3, bootstyle="danger-outline", command=lambda: negative_marks.set(max(0, negative_marks.get() - 1))).grid(row=1, column=1, padx=(10, 2))
        ttk.Label(marks_lf, textvariable=negative_marks, font="-weight bold", width=4, anchor=CENTER).grid(row=1, column=2)
        ttk.Button(marks_lf, text="+", width=3, bootstyle="success-outline", command=lambda: negative_marks.set(negative_marks.get() + 1)).grid(row=1, column=3, padx=2)

        # --- Frame Lịch sử chỗ ngồi ---
        history_lf = ttk.LabelFrame(container, text=" Lịch Sử Chỗ Ngồi ", padding=10)
        history_lf.pack(fill=X, pady=10)
        
        seating_history = self._get_seating_history(student_info['Học sinh'])
        if seating_history:
            history_text = "Thường ngồi cạnh nhất:\n"
            for partner, count in seating_history:
                history_text += f"- {partner} ({count} lần)\n"
        else:
            history_text = "Chưa có đủ dữ liệu lịch sử để phân tích."
        ttk.Label(history_lf, text=history_text, justify=LEFT).pack(anchor=W)

        # --- Frame Ghi chú riêng tư ---
        notes_lf = ttk.LabelFrame(container, text=" Ghi Chú Riêng Tư của Giáo Viên ", padding=10)
        notes_lf.pack(fill=BOTH, expand=True, pady=5)
        
        private_notes_text = scrolledtext.ScrolledText(notes_lf, wrap=tk.WORD, height=8, font=("Arial", 10))
        private_notes_text.pack(fill=BOTH, expand=True)
        private_notes_text.insert(tk.END, student_info.get('private_notes', ''))
        
        # --- Nút Lưu và Đóng ---
        def save_and_close():
            # Tìm đúng học sinh trong list gốc để cập nhật
            for s in self.students_data:
                if s['Học sinh'] == student_info['Học sinh']:
                    s['positive_marks'] = positive_marks.get()
                    s['negative_marks'] = negative_marks.get()
                    s['private_notes'] = private_notes_text.get('1.0', tk.END).strip()
                    break
            self._set_dirty()
            self.update_status(f"Đã cập nhật hồ sơ cho {student_info['Học sinh']}.")
            win.destroy()
        
        ttk.Button(container, text="Lưu và Đóng", command=save_and_close, bootstyle=SUCCESS).pack(pady=15)

    def _get_seating_history(self, student_name):
        if not self.current_class_id: return None
        
        self.cursor.execute("SELECT arrangement FROM arrangements_history WHERE class_id=? ORDER BY timestamp DESC LIMIT 20", (self.current_class_id,))
        history_data = self.cursor.fetchall()
        
        if not history_data: return None
        
        partner_counter = collections.Counter()
        
        for (arrangement_json,) in history_data:
            try:
                arrangement = json.loads(arrangement_json)
                if student_name not in arrangement: continue
                
                idx = arrangement.index(student_name)
                
                # Xác định bạn cùng bàn
                partner_idx = -1
                if idx % 2 == 0 and idx + 1 < len(arrangement): # Vị trí chẵn, bạn là người kế tiếp
                    partner_idx = idx + 1
                elif idx % 2 != 0: # Vị trí lẻ, bạn là người trước đó
                    partner_idx = idx - 1
                
                if partner_idx != -1:
                    partner_name = arrangement[partner_idx]
                    partner_counter[partner_name] += 1
            except (json.JSONDecodeError, ValueError):
                continue
                
        return partner_counter.most_common(3) # Lấy 3 người bạn ngồi cạnh nhiều nhất.
    
    def export_to_pdf(self):
        if not self.current_class_id or not self.students: messagebox.showwarning("Chưa có dữ liệu", "Vui lòng chọn một lớp và tải dữ liệu học sinh để xuất ra PDF."); return
        teacher_name = simpledialog.askstring("Thông tin bổ sung", "Nhập tên Giáo viên Chủ nhiệm:", parent=self.root)
        if teacher_name is None: return
        school_year = simpledialog.askstring("Thông tin bổ sung", "Nhập Năm học (VD: 2025-2026):", parent=self.root)
        if school_year is None: return
        file_path = filedialog.asksaveasfilename(title=f"Lưu sơ đồ PDF cho lớp {self.current_class_name}", defaultextension=".pdf", filetypes=[("PDF Files", "*.pdf")])
        if not file_path: return
        self.update_status("Đang tạo file PDF, vui lòng đợi...")
        try:
            pdf = PDFGenerator(orientation='L', unit='mm', format='A4', class_name=self.current_class_name, teacher_name=teacher_name, school_year=school_year)
            pdf.alias_nb_pages(); pdf.add_page()
            pdf.draw_seating_chart(students=self.students, num_teams=self.num_teams, num_tables_per_team=self.num_tables, colors_hex=self.colors)
            pdf.output(file_path)
            self.update_status(f"Đã xuất PDF thành công: {file_path}"); messagebox.showinfo("Thành công", f"Đã xuất sơ đồ lớp {self.current_class_name} ra file PDF thành công!")
        except Exception as e:
            if isinstance(e, RuntimeError) and ("TTF Font file not found" in str(e) or "DejaVuSans.ttf" in str(e)): messagebox.showerror("Lỗi Thiếu Font Chữ", "Không tìm thấy font chữ cần thiết để hỗ trợ tiếng Việt.\n\nGiải pháp:\n1. Đảm bảo font 'Times New Roman' đã được cài đặt trên máy của bạn.\n2. (Nếu cách 1 không được) Tải và đặt file 'DejaVuSans.ttf' vào cùng thư mục với ứng dụng.")
            else: messagebox.showerror("Lỗi", f"Không thể tạo file PDF: {e}")
            self.update_status("Lỗi khi tạo file PDF.")
            
    def _wrap_text(self, text, font_object, max_width):
        words = text.split(); lines = []; current_line = ""
        for word in words:
            separator = " " if current_line else ""
            test_line = current_line + separator + word
            if font_object.measure(test_line) <= max_width: current_line = test_line
            else: lines.append(current_line); current_line = word
        lines.append(current_line)
        return "\n".join(lines)

    def arrange_seats(self):
        self.canvas.delete("all"); self.seat_positions.clear(); self.text_positions.clear()
        if not self.students: 
            bbox = self.canvas.bbox("all")
            if bbox: self.canvas.config(scrollregion=bbox)
            return
        text_font = tkfont.Font(family="Arial", size=11, weight="bold")
        top_offset = self.TEACHER_DESK_HEIGHT + 100
        student_map = {s['Học sinh']: s for s in self.students_data}
        for i, student_name in enumerate(self.students):
            team_index = i // (self.num_tables * 2); table_in_team_index = (i % (self.num_tables * 2)) // 2; seat_index = i % 2
            student_info = student_map.get(student_name)
            if not student_info: continue
            team_visual_width = self.RECT_WIDTH * 2 + self.X_GAP + self.TEAM_X_GAP
            x_start = 20 + team_index * team_visual_width
            y_start = top_offset + table_in_team_index * (self.RECT_HEIGHT + self.Y_GAP)
            x = x_start + seat_index * (self.RECT_WIDTH + self.X_GAP); y = y_start
            outline_color = "#E53935" if student_info.get('GhiChu') == 'Cần ngồi trước' else ("#FFFFFF" if self.theme_var.get() else "#000000")
            outline_width = 3 if student_info.get('GhiChu') == 'Cần ngồi trước' else 1; text_color = "white" if self.theme_var.get() else "black"
            rect = self.canvas.create_rectangle(x, y, x + self.RECT_WIDTH, y + self.RECT_HEIGHT, fill=self.colors[team_index % len(self.colors)], tags="rect", outline=outline_color, width=outline_width)
            wrapped_name = self._wrap_text(student_name, text_font, self.RECT_WIDTH - 10)
            text = self.canvas.create_text(x + self.RECT_WIDTH / 2, y + self.RECT_HEIGHT / 2, text=wrapped_name, font=text_font, tags="text", fill=text_color, justify=tk.CENTER)
            self.seat_positions[rect] = student_info; self.text_positions[rect] = text
            self.canvas.tag_bind(rect, "<Button-1>", self.start_drag); self.canvas.tag_bind(rect, "<B1-Motion>", self.on_drag); self.canvas.tag_bind(rect, "<ButtonRelease-1>", self.stop_drag)
            self.canvas.tag_bind(rect, "<Enter>", lambda e, r=rect: self.show_tooltip(e, r)); self.canvas.tag_bind(rect, "<Leave>", self.hide_tooltip); self.canvas.tag_bind(rect, "<Button-3>", lambda e, r=rect: self.show_context_menu(e, r))
            # --- MỚI: Vẽ các icon ràng buộc ---
            icon_font = tkfont.Font(family="Arial", size=12)

            # Icon cho "muốn ngồi cạnh" (link)
            if student_info.get('sit_next_to'):
                self.canvas.create_text(
                    x + self.RECT_WIDTH - 12, y + 12,
                    text="🔗",
                    font=icon_font,
                    fill="#0052cc", # Màu xanh dương đậm
                    tags=(f"icon_{rect}", "icon")
                )
            # Icon cho "không muốn ngồi cạnh" (broken heart)
            if student_info.get('dont_sit_next_to'):
                self.canvas.create_text(
                    x + 12, y + 12,
                    text="💔",
                    font=icon_font,
                    fill="#d93025", # Màu đỏ đậm
                    tags=(f"icon_{rect}", "icon")
                )
            # <<--- MỚI: THÊM SỰ KIỆN DOUBLE-CLICK ĐỂ MỞ HỒ SƠ --- >>
            self.canvas.tag_bind(rect, "<Double-1>", lambda e, r=rect: self.show_student_profile(r))

        if self.num_teams > 0:
            team_1_start_x = 20; team_1_width = self.RECT_WIDTH * 2 + self.X_GAP; team_1_center_x = team_1_start_x + team_1_width / 2
            teacher_desk_x = team_1_center_x - (self.TEACHER_DESK_WIDTH / 2); teacher_desk_y = 20
            self.canvas.create_rectangle(teacher_desk_x, teacher_desk_y, teacher_desk_x + self.TEACHER_DESK_WIDTH, teacher_desk_y + self.TEACHER_DESK_HEIGHT, fill="#DEB887", outline=("white" if self.theme_var.get() else "black"))
            self.canvas.create_text(teacher_desk_x + self.TEACHER_DESK_WIDTH / 2, teacher_desk_y + self.TEACHER_DESK_HEIGHT / 2, text="Bàn Giáo Viên", font=text_font, fill=("white" if self.theme_var.get() else "black"))
        bbox = self.canvas.bbox("all")
        if bbox: self.canvas.config(scrollregion=bbox)

    def show_context_menu(self, event, rect_id):
        student_info = self.seat_positions.get(rect_id)
        if not student_info: return
        context_menu = tk.Menu(self.root, tearoff=0)
        context_menu.add_command(label=f"Ràng buộc cho: {student_info['Học sinh']}", command=lambda: self.open_constraint_window(student_info))
        try: context_menu.tk_popup(event.x_root, event.y_root)
        finally: context_menu.grab_release()
    def open_ai_report_window(self):
        if not self.current_class_id:
            messagebox.showwarning("Chưa chọn lớp", "Vui lòng chọn một lớp để xem báo cáo.")
            return

        win = ttk.Toplevel(self.root)
        win.title(f"Báo Cáo Phân Tích AI - Lớp {self.current_class_name}")
        win.geometry("2020x1000") # Tăng kích thước cửa sổ
        win.transient(self.root)
        win.grab_set()

        container = ttk.Frame(win, padding=10)
        container.pack(fill=BOTH, expand=True)

        pane = ttk.PanedWindow(container, orient=HORIZONTAL)
        pane.pack(fill=BOTH, expand=True, pady=5)
        
        # --- KHUNG BÊN TRÁI: BÁO CÁO VĂN BẢN ---
        report_frame = ttk.Frame(pane, padding=5)
        pane.add(report_frame, weight=2) # Cho text report nhỏ hơn
        
        ttk.Label(report_frame, text="Báo cáo Phân tích AI", font="-weight bold").pack(anchor=W)
        report_widget = scrolledtext.ScrolledText(report_frame, wrap=tk.WORD, font=("Arial", 10), state=tk.DISABLED)
        report_widget.pack(fill=BOTH, expand=True)
        
        # --- KHUNG BÊN PHẢI: BIỂU ĐỒ TRỰC QUAN ---
        charts_frame = ttk.Frame(pane, padding=5)
        pane.add(charts_frame, weight=3) # Cho biểu đồ lớn hơn
        
        charts_notebook = ttk.Notebook(charts_frame, bootstyle="primary")
        charts_notebook.pack(fill=BOTH, expand=True)
        
        effectiveness_tab = ttk.Frame(charts_notebook, padding=10)
        student_trend_tab = ttk.Frame(charts_notebook, padding=10)
        
        charts_notebook.add(effectiveness_tab, text=" Hiệu quả Sơ đồ ")
        charts_notebook.add(student_trend_tab, text=" Xu hướng Học sinh ")
        
        # Hiển thị thông báo chờ
        ttk.Label(effectiveness_tab, text="Nhấn 'Chạy Phân Tích' để xem biểu đồ.", bootstyle="secondary").pack(expand=True)
        ttk.Label(student_trend_tab, text="Nhấn 'Chạy Phân Tích' để xem biểu đồ.", bootstyle="secondary").pack(expand=True)

        def generate_and_show_report():
            self.update_status("AI đang phân tích dữ liệu lịch sử, vui lòng đợi...")
            win.update_idletasks()
            
            report, effectiveness_data, snapshots = self._generate_ai_report()
            
            # 1. Hiển thị báo cáo văn bản
            report_widget.config(state=tk.NORMAL)
            report_widget.delete('1.0', tk.END)
            report_widget.insert(tk.END, report)
            report_widget.config(state=tk.DISABLED)
            
            # 2. Vẽ các biểu đồ nếu có dữ liệu
            if effectiveness_data:
                self._plot_chart_effectiveness(effectiveness_data, effectiveness_tab)
            
            if snapshots:
                self._setup_student_trend_tab(snapshots, student_trend_tab)

            self.update_status("Phân tích AI hoàn tất.")
            
        ttk.Button(container, text="🚀 Chạy Phân Tích", command=generate_and_show_report, bootstyle="success").pack(pady=10)


    def _plot_chart_effectiveness(self, data, parent_frame):
            # Xóa widget cũ trong frame (nếu có)
            for widget in parent_frame.winfo_children():
                widget.destroy()
                
            fig = plt.Figure(figsize=(7, 5), dpi=100)
            ax = fig.add_subplot(111)

            dates = [f"Sơ đồ\n{item['date']}" for item in data]
            scores = [item['score'] for item in data]
            
            bars = ax.bar(dates, scores, color=['#4CAF50', '#FFC107', '#F44336'][:len(scores)])
            ax.set_title("So sánh Hiệu quả các Sơ đồ Chỗ ngồi", fontsize=14)
            ax.set_ylabel("Điểm hiệu quả (thay đổi điểm cộng/ngày)")
            ax.bar_label(bars, fmt='{:.2f}')
            
            fig.tight_layout()

            canvas = FigureCanvasTkAgg(fig, master=parent_frame)
            canvas.draw()
            canvas.get_tk_widget().pack(side=tk.TOP, fill=tk.BOTH, expand=True)

    def _setup_student_trend_tab(self, snapshots, parent_frame):
        # Xóa widget cũ trong frame (nếu có)
        for widget in parent_frame.winfo_children():
            widget.destroy()

        # Lấy danh sách học sinh duy nhất từ snapshots
        student_names = sorted(list(set(name for s in snapshots for name in s['data'].keys())))
        
        if not student_names:
            ttk.Label(parent_frame, text="Không có dữ liệu học sinh trong lịch sử.", bootstyle="secondary").pack(expand=True)
            return

        # Tạo frame chứa combobox và biểu đồ
        top_frame = ttk.Frame(parent_frame)
        top_frame.pack(fill=X, pady=5)
        
        ttk.Label(top_frame, text="Chọn học sinh để xem xu hướng:").pack(side=LEFT, padx=(0, 10))
        student_selector = ttk.Combobox(top_frame, values=student_names, state="readonly", width=30)
        student_selector.pack(side=LEFT)
        student_selector.set(student_names[0])
        
        chart_container = ttk.Frame(parent_frame)
        chart_container.pack(fill=BOTH, expand=True, pady=10)

        # Gán sự kiện và vẽ biểu đồ lần đầu
        student_selector.bind("<<ComboboxSelected>>", lambda event: self._plot_student_trend(snapshots, student_selector.get(), chart_container))
        self._plot_student_trend(snapshots, student_selector.get(), chart_container)

    def _plot_student_trend(self, snapshots, student_name, parent_frame):
        # Xóa widget cũ trong frame (nếu có)
        for widget in parent_frame.winfo_children():
            widget.destroy()

        dates = []
        points = []
        for s in snapshots:
            if student_name in s['data']:
                dates.append(s['timestamp'])
                points.append(s['data'][student_name].get('positive_marks', 0))

        if len(dates) < 2:
            ttk.Label(parent_frame, text=f"Không đủ dữ liệu để vẽ biểu đồ cho {student_name}.", bootstyle="secondary").pack(expand=True)
            return

        fig = plt.Figure(figsize=(7, 5), dpi=100)
        ax = fig.add_subplot(111)

        ax.plot(dates, points, marker='o', linestyle='-', color='#007BFF')
        ax.set_title(f"Xu hướng Tích lũy Điểm cộng - {student_name}", fontsize=14)
        ax.set_ylabel("Tổng số điểm cộng")
        ax.grid(True, linestyle='--', alpha=0.6)
        fig.autofmt_xdate() # Tự động xoay và căn chỉnh ngày
        fig.tight_layout()

        canvas = FigureCanvasTkAgg(fig, master=parent_frame)
        canvas.draw()
        canvas.get_tk_widget().pack(side=tk.TOP, fill=tk.BOTH, expand=True)
    def _generate_ai_report(self):
        self.cursor.execute("""
            SELECT timestamp, arrangement, students_data_snapshot_json 
            FROM arrangements_history 
            WHERE class_id=? AND students_data_snapshot_json IS NOT NULL 
            ORDER BY timestamp ASC
        """, (self.current_class_id,))
        
        history = self.cursor.fetchall()
        
        if len(history) < 2:
            return "Chưa có đủ dữ liệu lịch sử để phân tích. \n\nVui lòng sử dụng tính năng 'Tích điểm thi đua' và lưu sơ đồ nhiều lần để AI có dữ liệu học tập.", None, None

        # Chuyển đổi dữ liệu
        snapshots = []
        for timestamp_str, arr_json, data_json in history:
            snapshots.append({
                "timestamp": datetime.strptime(timestamp_str, '%Y-%m-%d %H:%M:%S'),
                "arrangement": json.loads(arr_json),
                "data": {s['Học sinh']: s for s in json.loads(data_json)}
            })

        # Bắt đầu tạo báo cáo
        report = f"BÁO CÁO PHÂN TÍCH AI CHO LỚP {self.current_class_name}\n"
        report += f"Dựa trên {len(snapshots)} bản ghi từ {snapshots[0]['timestamp'].strftime('%d/%m/%Y')} đến {snapshots[-1]['timestamp'].strftime('%d/%m/%Y')}\n"
        report += "="*60 + "\n\n"

        # 1. Phân tích xu hướng bất thường
        report += "1. Cảnh Báo Xu Hướng Bất Thường (Tương tác gần đây so với trước đó):\n\n"
        trends = self._analyze_student_trends(snapshots)
        if not trends:
            report += "   - Không có xu hướng bất thường nào được phát hiện.\n"
        else:
            for trend in trends:
                report += f"   - ⚠️ CHÚ Ý: Tương tác của em {trend['student']} có dấu hiệu giảm mạnh.\n"
                report += f"     (Điểm cộng gần đây: {trend['recent_rate']:.1f}/ngày, trước đó: {trend['past_rate']:.1f}/ngày)\n\n"
        
        report += "="*60 + "\n\n"
        
        # 2. Phân tích hiệu quả sơ đồ
        report += "2. Hiệu Quả của các Sơ Đồ Chỗ Ngồi (dựa trên thay đổi điểm cộng toàn lớp):\n\n"
        chart_effectiveness = self._analyze_chart_effectiveness(snapshots)
        if not chart_effectiveness:
             report += "   - Chưa đủ dữ liệu để so sánh hiệu quả các sơ đồ.\n"
        else:
            for i, chart in enumerate(chart_effectiveness):
                report += f"   - Hạng {i+1}: Sơ đồ ngày {chart['date']} (Hiệu quả: {chart['score']:.2f} điểm/ngày)\n"
            report += "\n   *Ghi chú: Sơ đồ hiệu quả là sơ đồ giúp toàn lớp có nhiều điểm cộng nhất trong thời gian áp dụng.*\n"

        report += "="*60 + "\n\n"
        
        # 3. Phân tích tương quan ẩn
        report += "3. Phân Tích Tương Quan Ẩn:\n\n"
        correlations = self._analyze_correlations(snapshots)
        report += f"   - Vị trí ngồi hiệu quả nhất: {correlations['best_position']}\n"
        if correlations['best_partners']:
            report += "   - Các cặp bạn cùng bàn hiệu quả (giúp nhau tiến bộ):\n"
            for student, partner in correlations['best_partners'].items():
                report += f"     + Em {student} tương tác tốt hơn khi ngồi cạnh em {partner}.\n"
        
        return report, chart_effectiveness, snapshots

    def _analyze_student_trends(self, snapshots, lookback_days=14):
        last_date = snapshots[-1]['timestamp']
        split_date = last_date - timedelta(days=lookback_days)
        
        student_points = collections.defaultdict(list)
        for s in snapshots:
            for name, data in s['data'].items():
                student_points[name].append((s['timestamp'], data.get('positive_marks', 0)))
        
        warnings = []
        for student, points_history in student_points.items():
            if len(points_history) < 2: continue
            
            recent_points = [(t, p) for t, p in points_history if t >= split_date]
            past_points = [(t, p) for t, p in points_history if t < split_date]

            if len(recent_points) < 2 or len(past_points) < 2: continue
            
            # Tính tốc độ tăng điểm
            past_days = (past_points[-1][0] - past_points[0][0]).days
            past_rate = (past_points[-1][1] - past_points[0][1]) / past_days if past_days > 0 else 0
            
            recent_days = (recent_points[-1][0] - recent_points[0][0]).days
            recent_rate = (recent_points[-1][1] - recent_points[0][1]) / recent_days if recent_days > 0 else 0

            if past_rate > 0.1 and recent_rate < past_rate * 0.2: # Giảm hơn 80%
                warnings.append({'student': student, 'recent_rate': recent_rate, 'past_rate': past_rate})
                
        return warnings

    def _analyze_chart_effectiveness(self, snapshots):
        chart_scores = {}
        for i in range(len(snapshots) - 1):
            s1 = snapshots[i]; s2 = snapshots[i+1]
            days_diff = (s2['timestamp'] - s1['timestamp']).days
            if days_diff == 0: continue
            
            total_points_s1 = sum(d.get('positive_marks', 0) for d in s1['data'].values())
            total_points_s2 = sum(d.get('positive_marks', 0) for d in s2['data'].values())
            
            points_gain = total_points_s2 - total_points_s1
            
            chart_key = tuple(s1['arrangement'])
            if chart_key not in chart_scores: chart_scores[chart_key] = {'total_gain': 0, 'total_days': 0, 'date': s1['timestamp'].strftime('%d/%m/%Y')}
            
            chart_scores[chart_key]['total_gain'] += points_gain
            chart_scores[chart_key]['total_days'] += days_diff
            
        results = []
        for chart, data in chart_scores.items():
            if data['total_days'] > 0:
                score = data['total_gain'] / data['total_days']
                results.append({'chart': chart, 'score': score, 'date': data['date']})
        
        return sorted(results, key=lambda x: x['score'], reverse=True)[:3] # Top 3 sơ đồ

    def _analyze_correlations(self, snapshots):
        position_points = collections.defaultdict(lambda: {'gain': 0, 'count': 0})
        partner_effects = collections.defaultdict(lambda: collections.defaultdict(list))

        for i in range(len(snapshots) - 1):
            s1 = snapshots[i]; s2 = snapshots[i+1]
            
            for j, student_name in enumerate(s1['arrangement']):
                # Phân tích vị trí
                pos_key = "Đầu lớp" if j < self.num_tables * 2 else ("Giữa lớp" if j < self.num_tables * 4 else "Cuối lớp")
                
                p1 = s1['data'].get(student_name, {}).get('positive_marks', 0)
                p2 = s2['data'].get(student_name, {}).get('positive_marks', 0)
                gain = p2 - p1
                
                position_points[pos_key]['gain'] += gain
                position_points[pos_key]['count'] += 1

                # Phân tích bạn cùng bàn
                partner_idx = j + 1 if j % 2 == 0 else j - 1
                if 0 <= partner_idx < len(s1['arrangement']):
                    partner_name = s1['arrangement'][partner_idx]
                    partner_effects[student_name][partner_name].append(gain)
        
        # Xử lý kết quả
        best_pos = "Chưa xác định"
        if position_points:
            avg_pos_points = {pos: data['gain']/data['count'] for pos, data in position_points.items() if data['count'] > 0}
            if avg_pos_points: best_pos = max(avg_pos_points, key=avg_pos_points.get)

        best_partners = {}
        for student, partners in partner_effects.items():
            avg_partner_effects = {p: sum(gains)/len(gains) for p, gains in partners.items() if gains}
            if avg_partner_effects:
                best_partner = max(avg_partner_effects, key=avg_partner_effects.get)
                if avg_partner_effects[best_partner] > 0.5: # Chỉ báo cáo nếu hiệu ứng đủ lớn
                    best_partners[student] = best_partner
        
        return {'best_position': best_pos, 'best_partners': best_partners}

    def open_constraint_window(self, student_info):
        student_name = student_info['Học sinh']
        other_students_map = {s['Học sinh']: i for i, s in enumerate(self.students_data) if s['Học sinh'] != student_name}
        other_students_list = list(other_students_map.keys())
        win = ttk.Toplevel(self.root); win.title(f"Ràng buộc cho {student_name}"); win.transient(self.root); win.grab_set()
        main_frame = ttk.Frame(win, padding=15); main_frame.pack(fill=BOTH, expand=True)
        sit_next_lf = ttk.LabelFrame(main_frame, text=" Chọn bạn MUỐN ngồi cạnh ", padding=10); sit_next_lf.pack(fill=BOTH, expand=True, pady=5)
        sit_next_lb = tk.Listbox(sit_next_lf, selectmode=tk.MULTIPLE, height=8, exportselection=False)
        sit_next_scroll = ttk.Scrollbar(sit_next_lf, orient=VERTICAL, command=sit_next_lb.yview, bootstyle='round'); sit_next_lb.config(yscrollcommand=sit_next_scroll.set); sit_next_scroll.pack(side=RIGHT, fill=Y); sit_next_lb.pack(side=LEFT, fill=BOTH, expand=True)
        dont_sit_lf = ttk.LabelFrame(main_frame, text=" Chọn bạn KHÔNG MUỐN ngồi cạnh ", padding=10); dont_sit_lf.pack(fill=BOTH, expand=True, pady=5)
        dont_sit_lb = tk.Listbox(dont_sit_lf, selectmode=tk.MULTIPLE, height=8, exportselection=False)
        dont_sit_scroll = ttk.Scrollbar(dont_sit_lf, orient=VERTICAL, command=dont_sit_lb.yview, bootstyle='round'); dont_sit_lb.config(yscrollcommand=dont_sit_scroll.set); dont_sit_scroll.pack(side=RIGHT, fill=Y); dont_sit_lb.pack(side=LEFT, fill=BOTH, expand=True)
        def on_sit_next_select(event):
            selected_indices = sit_next_lb.curselection()
            for idx in selected_indices:
                selected_name = sit_next_lb.get(idx)
                if selected_name in other_students_map: dont_sit_lb.selection_clear(other_students_list.index(selected_name))
        def on_dont_sit_select(event):
            selected_indices = dont_sit_lb.curselection()
            for idx in selected_indices:
                selected_name = dont_sit_lb.get(idx)
                if selected_name in other_students_map: sit_next_lb.selection_clear(other_students_list.index(selected_name))
        sit_next_lb.bind("<<ListboxSelect>>", on_sit_next_select); dont_sit_lb.bind("<<ListboxSelect>>", on_dont_sit_select)
        btn_frame = ttk.Frame(main_frame); btn_frame.pack(fill=X, pady=(10, 0))
        save_cmd = lambda: self._save_constraints(student_info, sit_next_lb, dont_sit_lb, win)
        ttk.Button(btn_frame, text="Lưu thay đổi", command=save_cmd, bootstyle=SUCCESS).pack(side=LEFT, expand=True, padx=5)
        ttk.Button(btn_frame, text="Hủy", command=win.destroy, bootstyle="secondary-outline").pack(side=LEFT, expand=True, padx=5)
        current_sit_next = student_info.get('sit_next_to', []); current_dont_sit = student_info.get('dont_sit_next_to', [])
        for i, s_name in enumerate(other_students_list):
            sit_next_lb.insert(tk.END, s_name); dont_sit_lb.insert(tk.END, s_name)
            if s_name in current_sit_next: sit_next_lb.selection_set(i)
            if s_name in current_dont_sit: dont_sit_lb.selection_set(i)

    def _save_constraints(self, student_info, sit_next_lb, dont_sit_lb, window):
        student_name = student_info['Học sinh']
        selected_sit_next_indices = sit_next_lb.curselection(); new_sit_next_list = {sit_next_lb.get(i) for i in selected_sit_next_indices}
        selected_dont_sit_indices = dont_sit_lb.curselection(); new_dont_sit_list = {dont_sit_lb.get(i) for i in selected_dont_sit_indices}
        for other_student in self.students_data:
            other_name = other_student['Học sinh']
            if other_name == student_name: continue
            other_sit_next = set(other_student.get('sit_next_to', [])); other_dont_sit = set(other_student.get('dont_sit_next_to', []))
            if other_name in new_sit_next_list: other_sit_next.add(student_name)
            else: other_sit_next.discard(student_name)
            if other_name in new_dont_sit_list: other_dont_sit.add(student_name)
            else: other_dont_sit.discard(student_name)
            other_student['sit_next_to'] = sorted(list(other_sit_next)); other_student['dont_sit_next_to'] = sorted(list(other_dont_sit))
        for student in self.students_data:
            if student['Học sinh'] == student_name:
                student['sit_next_to'] = sorted(list(new_sit_next_list)); student['dont_sit_next_to'] = sorted(list(new_dont_sit_list)); break
        self._set_dirty(); self.update_status(f"Đã cập nhật ràng buộc cho {student_name} và các bạn liên quan."); messagebox.showinfo("Thành công", "Đã lưu ràng buộc. Hãy nhấn 'Sắp xếp thông minh' để áp dụng.", parent=window); window.destroy()
    def pick_random_student(self):
        # ... (pick_random_student giữ nguyên)
        if not self.students:
            messagebox.showwarning("Chưa có dữ liệu", "Không có học sinh nào trong danh sách để lựa chọn.")
            return

        random_student_name = random.choice(self.students)
        self.update_status(f"Đang chọn ngẫu nhiên... Kết quả là: {random_student_name}!")

        target_rect_id = None
        for rect_id, student_info in self.seat_positions.items():
            if student_info['Học sinh'] == random_student_name:
                target_rect_id = rect_id
                break
        
        if target_rect_id:
            original_color = self.canvas.itemcget(target_rect_id, "fill")
            highlight_color = "#FFD700"

            def revert_highlight():
                self.canvas.itemconfig(target_rect_id, fill=original_color)

            self.canvas.itemconfig(target_rect_id, fill=highlight_color)
            
            text_item = self.text_positions.get(target_rect_id)
            if text_item:
                self.canvas.tag_raise(target_rect_id)
                self.canvas.tag_raise(text_item)
            
            messagebox.showinfo(
                "Học sinh được chọn",
                f"🌟 Chúc mừng em: {random_student_name}! 🌟",
                parent=self.root
            )

            self.root.after(2000, revert_highlight)
        else:
            messagebox.showinfo(
                 "Học sinh được chọn",
                f"Học sinh được chọn là: {random_student_name} (không tìm thấy trên sơ đồ)."
            )

    def toggle_theme(self):
        # ... (toggle_theme giữ nguyên)
        if self.theme_var.get():
            self.root.style.theme_use('darkly')
            self.canvas.config(bg="#303030")
        else:
            self.root.style.theme_use('litera')
            self.canvas.config(bg="white")
        if self.students: self.arrange_seats()

    def _save_state_for_undo(self):
        # ... (các hàm undo/redo giữ nguyên)
        if not self.students:
            return
        self.undo_stack.append(self.students.copy())
        self.redo_stack.clear()
        self._update_undo_redo_buttons()

    def _update_undo_redo_buttons(self):
        self.undo_btn.config(state="normal" if self.undo_stack else "disabled")
        self.redo_btn.config(state="normal" if self.redo_stack else "disabled")

    def undo(self):
        if not self.undo_stack: return
        self.redo_stack.append(self.students.copy())
        self.students = self.undo_stack.pop()
        self._set_dirty(); self._sync_data_and_ui()
        self.update_status("Đã hoàn tác hành động.")

    def redo(self):
        if not self.redo_stack: return
        self.undo_stack.append(self.students.copy())
        self.students = self.redo_stack.pop()
        self._set_dirty(); self._sync_data_and_ui()
        self.update_status("Đã làm lại hành động.")

    def _sync_data_and_ui(self):
        if self.students_data:
            student_map = {s['Học sinh']: s for s in self.students_data}
            self.students_data = [student_map[name] for name in self.students if name in student_map]
        
        self.arrange_seats()
        self._update_undo_redo_buttons()

    # <<--- MỚI: CẬP NHẬT HÀM APPLY_SORT ĐỂ SỬ DỤNG ADVANCEDSORTER --- >>
    def apply_sort(self, window):
        self._save_state_for_undo()
        sort_method = self.sort_method_var.get()
        
        base_arrangement = self.students.copy()
        
        if sort_method == "random": 
            random.shuffle(base_arrangement)
        elif sort_method == "height":
            df = pd.DataFrame(self.students_data)
            base_arrangement = df.sort_values(by='Chiều cao', ascending=False)['Học sinh'].tolist()
        elif sort_method == "gender":
            df = pd.DataFrame(self.students_data)
            males = df[df['Giới tính'] == 'Nam']['Học sinh'].tolist(); random.shuffle(males)
            females = df[df['Giới tính'] == 'Nữ']['Học sinh'].tolist(); random.shuffle(females)
            base_arrangement = []
            i, j = 0, 0
            while i < len(males) or j < len(females):
                if i < len(males): base_arrangement.append(males[i]); i += 1
                if j < len(females): base_arrangement.append(females[j]); j += 1
        
        # Luôn áp dụng sắp xếp thông minh sau các bước trên nếu được chọn
        if sort_method == "intelligent":
            # Tạo một sắp xếp cơ sở tốt (ví dụ theo chiều cao) trước khi tối ưu
            df = pd.DataFrame(self.students_data)
            base_arrangement = df.sort_values(by='Chiều cao', ascending=False)['Học sinh'].tolist()
            
            sorter = AdvancedSorter(self.students_data)
            self.students = sorter.arrange(base_arrangement)
        else:
            self.students = base_arrangement

        self._set_dirty()
        self._sync_data_and_ui()
        self.update_status(f"Đã áp dụng sắp xếp theo phương pháp: {sort_method}.")
        window.destroy()

    def stop_drag(self, event):
        # ... (stop_drag giữ nguyên)
        if not self.dragged_item: return
        drop_x = self.canvas.canvasx(event.x)
        drop_y = self.canvas.canvasy(event.y)
        overlapping_items = self.canvas.find_overlapping(drop_x, drop_y, drop_x, drop_y)
        target_item = None
        for item in overlapping_items:
            if item in self.seat_positions and item != self.dragged_item:
                target_item = item
                break
        if target_item:
            self._save_state_for_undo()
            
            dragged_info = self.seat_positions[self.dragged_item]
            target_info = self.seat_positions[target_item]
            dragged_index = self.students.index(dragged_info['Học sinh'])
            target_index = self.students.index(target_info['Học sinh'])
            self.students[dragged_index], self.students[target_index] = self.students[target_index], self.students[dragged_index]
            self.update_status(f"Đã hoán đổi vị trí: {dragged_info['Học sinh']} và {target_info['Học sinh']}.")
            self._set_dirty()
            self._sync_data_and_ui()
        else: 
            self.update_status("Thao tác kéo thả bị hủy.")
            self.arrange_seats()
            
        self.dragged_item = None

    def restore_arrangement(self, arrangement_json):
        # ... (restore_arrangement giữ nguyên)
        if not self.students_data:
            messagebox.showerror("Lỗi", "Không thể khôi phục khi chưa có danh sách học sinh.")
            return
        
        self._save_state_for_undo()
        
        restored_student_names = json.loads(arrangement_json)
        if set(restored_student_names) != {s['Học sinh'] for s in self.students_data}:
            messagebox.showwarning("Cảnh báo", "Danh sách học sinh trong lịch sử không khớp với danh sách hiện tại.")
        
        self.students = restored_student_names
        
        if self.num_teams * self.num_tables * 2 < len(self.students):
            messagebox.showwarning("Cảnh báo", f"Số ghế hiện tại không đủ.")
        
        self._set_dirty()
        self._sync_data_and_ui()
        self.update_status("Đã khôi phục sơ đồ từ lịch sử.")
        messagebox.showinfo("Thành công", "Đã khôi phục sơ đồ từ lịch sử!")
    
    def _load_class_list_to_selector(self):
        # ... (_load_class_list_to_selector giữ nguyên)
        try:
            self.cursor.execute("SELECT id, name FROM classes ORDER BY name")
            self.class_list = self.cursor.fetchall()
            class_names = [row[1] for row in self.class_list]
            self.class_selector['values'] = class_names
            if not class_names:
                self.class_selector.set("Chưa có lớp nào. Hãy thêm một lớp.")
        except sqlite3.Error as e:
            messagebox.showerror("Lỗi Database", f"Không thể tải danh sách lớp: {e}")

    def _on_class_selected(self, event=None):
        # ... (_on_class_selected giữ nguyên)
        if self.is_dirty:
            if not messagebox.askyesno("Cảnh báo", "Bạn có thay đổi chưa lưu ở lớp hiện tại. Bạn có muốn chuyển lớp và hủy các thay đổi đó?"):
                self.class_selector.set(self.current_class_name)
                return
        
        selected_name = self.class_selector.get()
        selected_class = next((c for c in self.class_list if c[1] == selected_name), None)

        if selected_class:
            self.current_class_id, self.current_class_name = selected_class
            try:
                self.cursor.execute("SELECT * FROM classes WHERE id=?", (self.current_class_id,))
                class_data = self.cursor.fetchone()
                
                self.students_data = []
                self.students = []
                
                if class_data and class_data[2]:
                    self.students_data = json.loads(class_data[2])
                if class_data and class_data[3]:
                    self.students = json.loads(class_data[3])
                
                self.num_teams = class_data[4] if class_data[4] else 1
                self.num_tables = class_data[5] if class_data[5] else 3
                
                self.team_spinbox.set(self.num_teams)
                self.table_spinbox.set(self.num_tables)
                
                self.undo_stack.clear()
                self.redo_stack.clear()
                self._set_dirty(False)
                self._sync_data_and_ui()
                self.update_status(f"Đã tải dữ liệu cho lớp: {self.current_class_name}")

            except sqlite3.Error as e:
                messagebox.showerror("Lỗi Database", f"Không thể tải dữ liệu lớp: {e}")
            except (json.JSONDecodeError, TypeError):
                messagebox.showwarning("Dữ liệu lỗi", "Dữ liệu của lớp này có thể bị lỗi. Vui lòng tải lại file Excel.")
                self._clear_canvas_and_data()

    def _add_new_class(self):
        # ... (_add_new_class giữ nguyên)
        class_name = simpledialog.askstring("Thêm Lớp Mới", "Nhập tên lớp học:", parent=self.root)
        if not class_name or not class_name.strip(): return
        class_name = class_name.strip()
        try:
            self.cursor.execute("INSERT INTO classes (name) VALUES (?)", (class_name,))
            self.conn.commit()
            self.update_status(f"Đã tạo lớp mới: {class_name}")
            self._load_class_list_to_selector()
            self.class_selector.set(class_name)
            self._on_class_selected()
        except sqlite3.IntegrityError:
            messagebox.showerror("Lỗi", f"Tên lớp '{class_name}' đã tồn tại. Vui lòng chọn tên khác.")
        except sqlite3.Error as e:
            messagebox.showerror("Lỗi Database", f"Không thể tạo lớp mới: {e}")

    def _save_class_state(self):
        # ... (_save_class_state giữ nguyên)
        if not self.current_class_id:
            messagebox.showwarning("Chưa chọn lớp", "Vui lòng chọn một lớp để lưu.")
            return

        try:
            students_data_json = json.dumps(self.students_data, ensure_ascii=False, indent=2)
            current_arrangement_json = json.dumps(self.students, ensure_ascii=False)
            self.num_teams = int(self.team_spinbox.get())
            self.num_tables = int(self.table_spinbox.get())
            
            self.cursor.execute("""
                UPDATE classes 
                SET students_data_json=?, current_arrangement_json=?, num_teams=?, num_tables=?, last_modified=CURRENT_TIMESTAMP
                WHERE id=?
            """, (students_data_json, current_arrangement_json, self.num_teams, self.num_tables, self.current_class_id))
            self.conn.commit()
            self._set_dirty(False)
            self.update_status(f"Đã lưu thành công trạng thái của lớp: {self.current_class_name}")
            messagebox.showinfo("Thành công", f"Đã lưu trạng thái của lớp '{self.current_class_name}'.")
        except (sqlite3.Error, TclError, ValueError) as e:
             messagebox.showerror("Lỗi", f"Không thể lưu trạng thái lớp: {e}")

    def _delete_class(self):
        # ... (_delete_class giữ nguyên)
        if not self.current_class_id:
            messagebox.showwarning("Chưa chọn lớp", "Vui lòng chọn một lớp để xóa.")
            return
        
        if messagebox.askyesno("Xác nhận Xóa", f"Bạn có chắc chắn muốn xóa vĩnh viễn lớp '{self.current_class_name}' và toàn bộ lịch sử của nó không?"):
            try:
                class_name_to_delete = self.current_class_name
                self.cursor.execute("DELETE FROM classes WHERE id=?", (self.current_class_id,))
                self.conn.commit()
                self._clear_canvas_and_data()
                self.current_class_id = None
                self.current_class_name = None
                self._load_class_list_to_selector()
                self.update_status(f"Đã xóa lớp: {class_name_to_delete}")
            except sqlite3.Error as e:
                messagebox.showerror("Lỗi Database", f"Không thể xóa lớp: {e}")

    def _clear_canvas_and_data(self):
        # ... (_clear_canvas_and_data giữ nguyên)
        self.students = []; self.students_data = []
        self.undo_stack.clear(); self.redo_stack.clear()
        self._sync_data_and_ui()
        
    def apply_team_table_config(self):
        # ... (apply_team_table_config giữ nguyên)
        try:
            self.num_teams = int(self.team_spinbox.get())
            self.num_tables = int(self.table_spinbox.get())
            if self.num_teams <= 0 or self.num_tables <= 0:
                raise ValueError("Số tổ và số bàn phải lớn hơn 0!")
            total_seats = self.num_teams * self.num_tables * 2
            if self.students and total_seats < len(self.students):
                messagebox.showwarning("Cảnh báo", f"Tổng số ghế ({total_seats}) nhỏ hơn số học sinh ({len(self.students)})!", title="Cấu hình không hợp lệ")
                return
            if self.students:
                self.arrange_seats()
                self._set_dirty()
            self.update_status(f"Đã thiết lập: {self.num_teams} tổ, {self.num_tables} bàn mỗi tổ.")
        except (ValueError, TclError):
            messagebox.showerror("Lỗi", "Số tổ và số bàn phải là các số hợp lệ.", title="Lỗi dữ liệu")
            self.update_status("Lỗi: Dữ liệu cấu hình không hợp lệ.")

    def show_sort_options(self):
        # ... (show_sort_options giữ nguyên)
        if not self.students_data:
            messagebox.showwarning("Cảnh báo", "Vui lòng nhập danh sách học sinh trước!", title="Chưa có dữ liệu")
            return
        
        sort_window = ttk.Toplevel(self.root)
        sort_window.title("Tùy chọn sắp xếp")
        sort_window.transient(self.root); sort_window.grab_set()
        
        container = ttk.Frame(sort_window, padding=20); container.pack(fill=BOTH, expand=True)
        ttk.Label(container, text="Chọn phương pháp sắp xếp:", font=("Arial", 12)).pack(pady=10)
        self.sort_method_var = tk.StringVar(value="intelligent")

        style = ttk.Style(); style.configure('TRadiobutton', font=('Arial', 10), padding=(0,5))
        
        ttk.Radiobutton(container, text="Thông minh (ưu tiên ràng buộc)", variable=self.sort_method_var, value="intelligent", style='TRadiobutton').pack(fill=X, padx=20, pady=2)
        ttk.Radiobutton(container, text="Ngẫu nhiên", variable=self.sort_method_var, value="random", style='TRadiobutton').pack(fill=X, padx=20, pady=2)
        ttk.Radiobutton(container, text="Xen kẽ nam-nữ", variable=self.sort_method_var, value="gender", style='TRadiobutton').pack(fill=X, padx=20, pady=2)
        ttk.Radiobutton(container, text="Theo chiều cao (cao trước, thấp sau)", variable=self.sort_method_var, value="height", style='TRadiobutton').pack(fill=X, padx=20, pady=2)
        
        button_frame = ttk.Frame(container); button_frame.pack(pady=20)
        ttk.Button(button_frame, text="Áp dụng", command=lambda: self.apply_sort(sort_window), bootstyle=SUCCESS).pack(side=LEFT, padx=10)
        ttk.Button(button_frame, text="Hủy", command=sort_window.destroy, bootstyle="secondary-outline").pack(side=LEFT, padx=10)
        
    # <<--- MỚI: CẬP NHẬT HÀM LOAD_STUDENTS ĐỂ ĐỌC RÀNG BUỘC TỪ EXCEL --- >>
    def load_students(self):
        if not self.current_class_id:
            messagebox.showwarning("Chưa chọn lớp", "Vui lòng chọn hoặc thêm một lớp trước khi tải danh sách học sinh.")
            return
        
        file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx;*.xls")])
        if not file_path: return
        try:
            df = pd.read_excel(file_path)
            required_columns = ['Học sinh', 'Giới tính', 'Chiều cao', 'DiemTB']
            if not all(col in df.columns for col in required_columns):
                raise ValueError("File Excel phải có các cột: 'Học sinh', 'Giới tính', 'Chiều cao', 'DiemTB'.\nCác cột tùy chọn: 'GhiChu', 'KhongNgoiCanh', 'NgoiCanh'.")
            
            # Chuẩn hóa các cột ràng buộc
            df['GhiChu'] = df.get('GhiChu', pd.Series(index=df.index, dtype='object')).fillna('')
            df['KhongNgoiCanh'] = df.get('KhongNgoiCanh', pd.Series(index=df.index, dtype='object')).fillna('')
            df['NgoiCanh'] = df.get('NgoiCanh', pd.Series(index=df.index, dtype='object')).fillna('')

            for index, row in df.iterrows():
                if pd.isna(row['Chiều cao']) or not isinstance(row['Chiều cao'], (int, float)): raise ValueError(f"Dòng {index + 2}: 'Chiều cao' phải là một con số.")
                if pd.isna(row['DiemTB']) or not isinstance(row['DiemTB'], (int, float)): raise ValueError(f"Dòng {index + 2}: 'DiemTB' phải là một con số.")
                if row['Giới tính'] not in ['Nam', 'Nữ']: raise ValueError(f"Dòng {index + 2}: 'Giới tính' phải là 'Nam' hoặc 'Nữ'.")

            self.students_data = df.to_dict('records')
            
            # Chuyển đổi chuỗi ràng buộc thành danh sách
            for student in self.students_data:
                dont_sit_str = student.get('KhongNgoiCanh', '')
                student['dont_sit_next_to'] = [name.strip() for name in str(dont_sit_str).split(',') if name.strip()]
                
                sit_next_str = student.get('NgoiCanh', '')
                student['sit_next_to'] = [name.strip() for name in str(sit_next_str).split(',') if name.strip()]

            self.students = df['Học sinh'].tolist()
            
            self.update_status(f"Đã tải {len(self.students)} HS cho lớp {self.current_class_name}. Nhấn 'Lưu Trạng Thái' để ghi nhớ.")
            random.shuffle(self.students)
            self.undo_stack.clear(); self.redo_stack.clear()
            self._set_dirty()
            self._sync_data_and_ui()
        except Exception as e:
            messagebox.showerror("Lỗi tải file", f"Đã xảy ra lỗi: {str(e)}", title="Lỗi")
            self.update_status(f"Lỗi tải file Excel cho lớp {self.current_class_name}.")

    # ... (Các hàm còn lại từ change_team_colors đến hết giữ nguyên không thay đổi)
    def change_team_colors(self):
        if self.num_teams == 0:
            messagebox.showinfo("Thông báo", "Vui lòng thiết lập số tổ trước.")
            return
        
        color_window = ttk.Toplevel(self.root)
        color_window.title("Đổi màu tổ")
        
        for i in range(self.num_teams):
            frame = ttk.Frame(color_window, padding=5)
            frame.pack(fill=X)
            while i >= len(self.colors): self.colors.append("#FFFFFF")
            color_preview = tk.Label(frame, text="    ", bg=self.colors[i], relief="solid", borderwidth=1)
            color_preview.pack(side=LEFT, padx=5)
            ttk.Label(frame, text=f"Màu cho tổ {i + 1}:").pack(side=LEFT, padx=5)
            ttk.Button(frame, text="Chọn màu", bootstyle="outline", command=lambda idx=i, p=color_preview: self.choose_color(idx, p)).pack(side=LEFT, padx=5)

    def view_history(self):
        if not self.current_class_id:
            messagebox.showwarning("Chưa chọn lớp", "Vui lòng chọn một lớp để xem lịch sử.")
            return

        history_window = ttk.Toplevel(self.root)
        history_window.title(f"Lịch sử sắp xếp - Lớp {self.current_class_name}")
        history_window.geometry("1050x1000")
        history_window.transient(self.root); history_window.grab_set()
        
        left_frame = ttk.Frame(history_window, padding=5); left_frame.pack(side=LEFT, fill=Y)
        right_frame = ttk.Frame(history_window, padding=5); right_frame.pack(side=RIGHT, fill=BOTH, expand=True)
        ttk.Label(left_frame, text="Các phiên đã lưu", font=("Arial", 12, "bold")).pack(pady=5)
        list_frame = ttk.Frame(left_frame); list_frame.pack(fill=BOTH, expand=True)
        history_listbox = tk.Listbox(list_frame, width=30, font=("Arial", 10)); scrollbar = ttk.Scrollbar(list_frame, orient="vertical", command=history_listbox.yview, bootstyle="round")
        history_listbox.config(yscrollcommand=scrollbar.set); scrollbar.pack(side=RIGHT, fill=Y); history_listbox.pack(side=LEFT, fill=BOTH, expand=True)
        ttk.Label(right_frame, text="Xem trước Sơ đồ", font=("Arial", 12, "bold")).pack(pady=5)
        preview_text = scrolledtext.ScrolledText(right_frame, wrap=tk.WORD, state=tk.DISABLED, font=("Arial", 10)); preview_text.pack(fill=BOTH, expand=True, pady=5)
        button_frame = ttk.Frame(right_frame); button_frame.pack(fill=X, pady=5)
        restore_btn = ttk.Button(button_frame, text="Khôi phục phiên này", state=tk.DISABLED, bootstyle=SUCCESS); restore_btn.pack(side=LEFT, padx=5, expand=True, fill=X)
        delete_btn = ttk.Button(button_frame, text="Xóa mục này", state=tk.DISABLED, bootstyle=DANGER); delete_btn.pack(side=LEFT, padx=5, expand=True, fill=X)
        ttk.Button(right_frame, text="Xóa toàn bộ lịch sử của lớp này", bootstyle="danger-outline", command=lambda: self.clear_history(history_window)).pack(fill=X, pady=10)

        self.cursor.execute("SELECT id, timestamp, arrangement FROM arrangements_history WHERE class_id=? ORDER BY timestamp DESC", (self.current_class_id,))
        arrangements_data = self.cursor.fetchall()
        
        if not arrangements_data: history_listbox.insert(tk.END, "Không có lịch sử nào."); return
        history_map = {index: {'id': arr_id, 'json': arr_json} for index, (arr_id, _, arr_json) in enumerate(arrangements_data)}
        for index, (arr_id, timestamp, _) in enumerate(arrangements_data): history_listbox.insert(tk.END, f"ID {arr_id}: {timestamp}")
        def on_history_select(event):
            selected_indices = history_listbox.curselection()
            if not selected_indices: return
            selected_index = selected_indices[0]; data = history_map.get(selected_index)
            if not data: return
            preview_text.config(state=tk.NORMAL); preview_text.delete('1.0', tk.END)
            students_list = json.loads(data['json'])
            preview_content = f"--- SƠ ĐỒ CHI TIẾT (ID: {data['id']}) ---\n\n"
            for i, student_name in enumerate(students_list):
                team = i // (self.num_tables * 2) + 1; table = (i % (self.num_tables * 2)) // 2 + 1; seat = "A" if i % 2 == 0 else "B"
                preview_content += f"Tổ {team} - Bàn {table} - Ghế {seat}: {student_name}\n"
                if seat == "B": preview_content += "-"*20 + "\n"
            preview_text.insert('1.0', preview_content); preview_text.config(state=tk.DISABLED)
            restore_btn.config(state=tk.NORMAL, command=lambda: (self.restore_arrangement(data['json']), history_window.destroy()))
            delete_btn.config(state=tk.NORMAL, command=lambda: self.delete_history_entry(data['id'], history_window))
        history_listbox.bind('<<ListboxSelect>>', on_history_select)

    def save_results(self): # Hàm này được gọi khi "Xuất ra Excel"
        if not self.current_class_id or not self.students:
            messagebox.showwarning("Cảnh báo", "Không có dữ liệu để xuất hoặc chưa chọn lớp.", title="Lỗi")
            return
        
        try:
            arrangement_to_save = json.dumps(self.students, ensure_ascii=False)
            # <<--- THAY ĐỔI QUAN TRỌNG: Lưu cả snapshot của students_data --- >>
            students_data_snapshot = json.dumps(self.students_data, ensure_ascii=False)

            self.cursor.execute("""
                INSERT INTO arrangements_history 
                (class_id, arrangement, timestamp, students_data_snapshot_json) 
                VALUES (?, ?, datetime('now', 'localtime'), ?)
            """, (self.current_class_id, arrangement_to_save, students_data_snapshot))
            self.conn.commit()
            self.update_status(f"Đã lưu một bản ghi vào lịch sử của lớp {self.current_class_name}.")
        except sqlite3.Error as e:
            messagebox.showerror("Lỗi DB", f"Không thể lưu vào lịch sử: {e}")
            return
            
        wb = Workbook(); ws = wb.active
        ws.append(["Tổ", "Bàn", "Ghế", "Học sinh", "Giới tính", "Chiều cao", "Điểm TB", "Ghi Chú"])
        for i, student_name in enumerate(self.students):
            team = i // (self.num_tables * 2) + 1; table = (i % (self.num_tables * 2)) // 2 + 1; seat = i % 2 + 1
            student_info = next((s for s in self.students_data if s['Học sinh'] == student_name), None)
            if student_info: ws.append([team, table, seat, student_info['Học sinh'], student_info.get('Giới tính'), student_info.get('Chiều cao'), student_info.get('DiemTB'), student_info.get('GhiChu')])
        
        file_path = filedialog.asksaveasfilename(
            title=f"Lưu sơ đồ lớp {self.current_class_name}",
            defaultextension=".xlsx", 
            filetypes=[("Excel Files", "*.xlsx")]
        )
        if file_path:
            try:
                wb.save(file_path)
                messagebox.showinfo("Thành công", f"Đã xuất sơ đồ lớp {self.current_class_name} ra file Excel và lưu vào lịch sử!")
            except Exception as e:
                messagebox.showerror("Lỗi", f"Không thể lưu file: {e}")

    def save_as_image(self):
        if not self.students: messagebox.showwarning("Cảnh báo", "Không có sơ đồ để chụp ảnh!", title="Lỗi"); return
        file_path = filedialog.asksaveasfilename(defaultextension=".png", filetypes=[("PNG Files", "*.png")])
        if file_path:
            x = self.canvas.winfo_rootx(); y = self.canvas.winfo_rooty()
            x1 = x + self.canvas.winfo_width(); y1 = y + self.canvas.winfo_height()
            ImageGrab.grab().crop((x, y, x1, y1)).save(file_path)
            self.update_status(f"Đã lưu ảnh sơ đồ thành công: {file_path}")
            messagebox.showinfo("Thành công", "Lưu ảnh thành công!")

    def start_drag(self, event):
        items = self.canvas.find_overlapping(event.x, event.y, event.x, event.y);
        if items:
            closest_item = items[-1]
            if "rect" in self.canvas.gettags(closest_item):
                self.dragged_item = closest_item; self.canvas.tag_raise(self.dragged_item)
                text_item = self.text_positions.get(self.dragged_item)
                if text_item: self.canvas.tag_raise(text_item)
                self.update_status(f"Đang di chuyển học sinh: {self.seat_positions[self.dragged_item]['Học sinh']}")

    def on_drag(self, event):
        if self.dragged_item:
            canvas_x = self.canvas.canvasx(event.x); canvas_y = self.canvas.canvasy(event.y)
            text_item = self.text_positions.get(self.dragged_item)
            self.canvas.coords(self.dragged_item, canvas_x - self.RECT_WIDTH/2, canvas_y - self.RECT_HEIGHT/2, canvas_x + self.RECT_WIDTH/2, canvas_y + self.RECT_HEIGHT/2)
            if text_item: self.canvas.coords(text_item, canvas_x, canvas_y)

    def show_tooltip(self, event, rect_id):
        if self.tooltip: self.tooltip.destroy()
        student_info = self.seat_positions.get(rect_id)
        if not student_info: return
        
        # <<--- MỚI: HIỂN THỊ RÀNG BUỘC TRONG TOOLTIP --- >>
        sit_next_text = ", ".join(student_info.get('sit_next_to', [])) or "Không"
        dont_sit_text = ", ".join(student_info.get('dont_sit_next_to', [])) or "Không"
        
        text = (f"Tên: {student_info['Học sinh']}\n"
                f"Giới tính: {student_info.get('Giới tính', 'N/A')}\n"
                f"Điểm TB: {student_info.get('DiemTB', 'N/A')}\n"
                f"Ghi chú: {student_info.get('GhiChu', 'Không')}\n"
                f"--- Ràng buộc ---\n"
                f"Ngồi cạnh: {sit_next_text}\n"
                f"Không ngồi cạnh: {dont_sit_text}")
                
        self.tooltip = ttk.Toplevel(self.root); self.tooltip.wm_overrideredirect(True); self.tooltip.wm_geometry(f"+{event.x_root + 15}+{event.y_root + 10}")
        ttk.Label(self.tooltip, text=text, justify=LEFT, padding=5, background="#FFFFE0", relief="solid", borderwidth=1).pack()

    def hide_tooltip(self, event):
        if self.tooltip: self.tooltip.destroy(); self.tooltip = None

    def choose_color(self, team_index, preview_label):
        color_code = colorchooser.askcolor(title=f"Chọn màu cho tổ {team_index + 1}")
        if color_code[1]:
            self.colors[team_index] = color_code[1]
            preview_label.config(bg=self.colors[team_index])
            if self.students: self.arrange_seats(); self.update_status(f"Đã đổi màu cho tổ {team_index + 1}.")
    
    def delete_history_entry(self, arrangement_id, window):
        if messagebox.askyesno("Xác nhận", f"Bạn có chắc muốn xóa vĩnh viễn mục lịch sử ID: {arrangement_id}?"):
            self.cursor.execute("DELETE FROM arrangements_history WHERE id = ?", (arrangement_id,)); self.conn.commit()
            self.update_status(f"Đã xóa mục lịch sử ID: {arrangement_id}.")
            window.destroy(); self.view_history()

    def clear_history(self, window=None):
        if not self.current_class_id: return
        if messagebox.askyesno("Xác nhận", f"Bạn có chắc chắn muốn xóa TOÀN BỘ lịch sử của lớp '{self.current_class_name}' không?"):
            self.cursor.execute("DELETE FROM arrangements_history WHERE class_id=?", (self.current_class_id,)); self.conn.commit()
            self.update_status(f"Đã xóa toàn bộ lịch sử của lớp {self.current_class_name}.")
            if window: window.destroy(); self.view_history()


if __name__ == "__main__":
    root = ttk.Window(themename="litera")
    app = SeatArrangementApp(root)
    root.mainloop()